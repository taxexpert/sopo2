"""
엑셀 생성 모듈 — 기존 매출집계 파일과 동일한 형식으로 출력
환율 적용 기준: 소포수령증 발행일(write_date) 환율
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from typing import Optional
import re
from pathlib import Path
from io import BytesIO
from copy import copy


# ── 스타일 정의 ────────────────────────────────────────────────
HEADER_FILL   = PatternFill('solid', start_color='D9E1F2', end_color='D9E1F2')
SUBHEAD_FILL  = PatternFill('solid', start_color='E2EFDA', end_color='E2EFDA')
GRAY_FILL     = PatternFill('solid', start_color='F2F2F2', end_color='F2F2F2')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

FONT_DEFAULT = Font(name='맑은 고딕', size=9)
FONT_BOLD    = Font(name='맑은 고딕', size=9, bold=True)
FONT_TITLE   = Font(name='맑은 고딕', size=11, bold=True)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center')

# exchange_rate 모듈에서 모든 환율을 1통화 단위로 정규화하므로 추가 제수는 없습니다.
RATE_DIVISOR = {}

# 제출자(판매자) 정보 기본값 — PDF에서 못 읽었을 때만 사용
DEFAULT_SUBMITTER = {
    # 제출자 정보를 읽지 못한 경우 특정 업체 정보 대신 중립적인 기본 문구를 사용합니다.
    'name': '사업자명(사업자번호)', 'biz_no': '',
    'ceo': '', 'address': '',
}

# 숫자 천단위 콤마 서식
NUM_FMT  = '#,##0'        # 정수(수량·원화)
NUM_FMT2 = '#,##0.00'     # 일반 소수/환율
NUM_FMT4 = '#,##0.0000'   # JPY·IDR·VND 1통화 단위 적용환율
HUNDRED_UNIT_CURRENCIES = {'JPY', 'IDR', 'VND'}

# 수출실적명세서의 수출신고번호는 공란, 기타영세율건수는 1로 신고합니다.
# 영세율첨부서류의 L/C 번호 칸에는 운송장번호를 입력합니다
# (v54 확정 규칙 — docs/HANDOFF_V54.md §7.2·§8. 대상 신고 시스템 확인 전까지 유지).
TRACKING_NO_PATTERN = re.compile(r"^[A-Z]{2}[A-Z0-9]{13}$", re.I)

def is_valid_tracking_no(value):
    text = re.sub(r"[^A-Za-z0-9]", "", str(value or "")).upper()
    return bool(TRACKING_NO_PATTERN.fullmatch(text))

def other_zero_rate_count_value(value=None):
    return 1

def _date_to_int(value):
    d = re.sub(r"\D", "", str(value or ""))[:8]
    return int(d) if len(d) == 8 else None


def _submitter_label(submitter: dict = None) -> str:
    """상호와 사업자번호를 중복 괄호 없이 표시합니다."""
    sub = submitter or {}
    name = str(sub.get('name') or '사업자명(사업자번호)').strip()
    biz_no = str(sub.get('biz_no') or '').strip()
    return f"{name}({biz_no})" if biz_no else name


def _applied_rate_format(currency: str) -> str:
    return NUM_FMT4 if str(currency or '').upper() in HUNDRED_UNIT_CURRENCIES else NUM_FMT2


def _applied_rate_value(currency: str, value) -> float:
    """신고/집계용 1통화 단위 환율. 구형 100통화 캐시도 안전하게 정규화합니다."""
    from .exchange_rate import round_applied_rate
    return round_applied_rate(currency, value)


def _lazada_item_date(item: dict, lazada_result: dict = None, fallback: str = "") -> str:
    """라자다 주문 Excel은 deliveredDate, 기존 PDF는 거래기간 종료일을 사용합니다.

    작성일자(write_date)는 익월일 수 있어 귀속월이 어긋나므로 최후순위로만 씁니다.
    """
    item = item or {}
    lazada_result = lazada_result or {}
    return (item.get("date") or item.get("delivered_date") or fallback
            or lazada_result.get("period_end") or lazada_result.get("write_date") or "")


def _lazada_item_rate(item: dict, currency: str, rates: dict,
                       lazada_result: dict = None, avg_rate: float = None) -> float:
    """라자다 주문 Excel은 배송완료일 일별환율, 기존 소포수령증 PDF는 기간평균환율."""
    item = item or {}
    source_kind = item.get("source_kind") or (lazada_result or {}).get("source_kind")
    item_date = item.get("date") or item.get("delivered_date")
    if source_kind == "order_excel" or item_date:
        return _get_rate(rates, currency, item_date)
    base = avg_rate if avg_rate is not None else rates.get(currency, {}).get("average", 0.0)
    return _applied_rate_value(currency, base)


def _lazada_source_sheet_name(lazada_result: dict) -> str:
    items = (lazada_result or {}).get("items", [])
    if any((it.get("source_kind") == "order_excel") for it in items):
        return "라자다(주문내역)"
    return "라자다(소포수령증)"


def _smbs_display_rate(currency: str, value):
    """환율(통화) 시트 표시용 SMBS 고시단위 값을 반환합니다."""
    if value in (None, ''):
        return value
    try:
        from .exchange_rate import smbs_source_rate
        return smbs_source_rate(currency, value)
    except (TypeError, ValueError):
        return value


def _exchange_currency_name(currency: str, name: str) -> str:
    cur = str(currency or '').upper()
    text = str(name or cur)
    if cur in HUNDRED_UNIT_CURRENCIES and '(100)' not in text:
        return f'{text} (100)'
    return text



def _months_between(start, end):
    """거래기간(start~end)이 포함하는 (연,월) 리스트."""
    def _p(x):
        d = re.sub(r'\D', '', str(x))[:8]
        return (int(d[:4]), int(d[4:6])) if len(d) >= 6 else None
    a = _p(start); b = _p(end)
    a = a or b; b = b or a
    if not a:
        return []
    if b < a:
        a, b = b, a
    out = []
    y, m = a
    while (y, m) <= b:
        out.append((y, m))
        m += 1
        if m > 12:
            m = 1; y += 1
    return out


def period_labels(shopee_results, lazada_result, qoo10_result, ebay_results=None,
                  joom_results=None, shopify_results=None, fallback=''):
    """데이터 거래기간으로 (표시용, 파일명용) 라벨 생성.
    예: 10월 / 10~12월 / 1~12월 / 3,5,9월(파일명) · 3/5/9월(표시)."""
    pairs = []
    for sd in (shopee_results or []):
        pairs.append((sd.get('period_start', ''), sd.get('period_end', '')))
    if lazada_result:
        pairs.append((lazada_result.get('period_start', ''), lazada_result.get('period_end', '')))
    if qoo10_result:
        pairs.append((qoo10_result.get('period_start', ''), qoo10_result.get('period_end', '')))
    for er in (ebay_results or []):
        pairs.append((er.get('period_start', ''), er.get('period_end', '')))
    for jr in (joom_results or []):
        pairs.append((jr.get('period_start', ''), jr.get('period_end', '')))
    for sr in (shopify_results or []):
        pairs.append((sr.get('period_start', ''), sr.get('period_end', '')))
    yms = set()
    for s_, e_ in pairs:
        yms.update(_months_between(s_, e_))
    if not yms:
        return fallback, fallback
    years = sorted(set(y for y, m in yms))

    def _fmt(y, list_sep):
        ms = sorted(m for yy, m in yms if yy == y)
        if len(ms) == 1:
            return f'{ms[0]}월'
        if ms == list(range(ms[0], ms[-1] + 1)):
            return f'{ms[0]}~{ms[-1]}월'
        return list_sep.join(str(m) for m in ms) + '월'

    if len(years) == 1:
        y = years[0]
        disp = f'{y}년 {_fmt(y, "/")}'
        fname = _fmt(y, ',')
    else:
        disp = ', '.join(f'{y}년 {_fmt(y, "/")}' for y in years)
        fname = '_'.join(f'{y}년{_fmt(y, ",")}' for y in years)
    return disp, fname


def _style(cell, font=None, fill=None, align=None, border=None, num_format=None):
    if font:      cell.font       = font
    if fill:      cell.fill       = fill
    if align:     cell.alignment  = align
    if border:    cell.border     = border
    if num_format: cell.number_format = num_format


# ── 소포수령증 표 열 그룹 (값 열 + 사이 빈 열을 병합해 깔끔하게 이어줌) ──
_RECEIPT_GROUPS_2 = [(1, 3), (4, 6), (7, 10), (11, 12), (13, 15), (16, 19)]
_RECEIPT_GROUPS_3 = [(1, 3), (4, 6), (7, 10), (11, 12), (13, 15), (16, 17), (18, 18), (19, 19)]


def _merge_row(ws, row, groups, border=None):
    """한 행에서 각 열 그룹을 병합하고(2칸 이상), 그룹 전체에 테두리를 적용.
    겹침 검사 없는 빠른 병합(대량 행 처리 속도 향상, 결과는 동일)."""
    for c1, c2 in groups:
        if c2 > c1:
            ws.merged_cells.ranges.add(
                CellRange(min_col=c1, min_row=row, max_col=c2, max_row=row))
        if border is not None:
            for col in range(c1, c2 + 1):
                ws.cell(row=row, column=col).border = border


def _get_rate(rates: dict, currency: str, date_str: str) -> float:
    """
    발행일(write_date) 기준 환율 반환.
    daily 데이터 없으면 average(수동입력값) 반환.
    date_str이 비어 있으면 average 반환.
    """
    from .exchange_rate import get_rate_for_date
    rate_data = rates.get(currency)
    if not rate_data:
        return 0.0
    # daily 데이터가 없으면 average 반환 (수동입력 모드)
    if not rate_data.get('daily'):
        return _applied_rate_value(currency, rate_data.get('average', 0.0))
    if not date_str:
        return _applied_rate_value(currency, rate_data.get('average', 0.0))
    rate = get_rate_for_date(rate_data, date_str)
    if rate == 0.0:
        rate = rate_data.get('average', 0.0)
    return _applied_rate_value(currency, rate)


# ── 환율 시트 작성 ──────────────────────────────────────────────

def write_exchange_rate_sheet(ws, rate_data: dict):
    """환율(XXX) 시트를 SMBS 원문 표시단위로 작성합니다.

    내부 계산값은 1통화 단위이지만 JPY/IDR/VND는 이 시트에서만
    서울외국환중개와 동일하게 100통화 단위 환율로 다시 표시합니다.
    """
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    if rate_data is None:
        ws['A1'] = '환율 데이터 없음 (수동 입력 필요)'
        return

    currency = str(rate_data.get('currency', '') or '').upper()
    currency_name = _exchange_currency_name(currency, rate_data.get('currency_name', ''))

    def src(value):
        return _smbs_display_rate(currency, value)

    # 월평균 전용 데이터만 있는 경우(이베이/큐텐재팬 등)
    if rate_data.get('monthly') and not rate_data.get('daily'):
        ws['A1'] = '월평균 매매기준율'
        _style(ws['A1'], font=FONT_BOLD)
        ws['A2'] = f"기간 : {rate_data.get('period', '')}"
        headers = ['년월', '통화명', '월평균환율']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
        for r, d in enumerate(rate_data.get('monthly', []), 5):
            vals = [d.get('year_month', ''), currency_name, round(float(src(d.get('rate', 0)) or 0), 2)]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=v)
                _style(
                    c, font=FONT_DEFAULT, align=CENTER if col != 2 else LEFT, border=THIN_BORDER,
                    num_format=NUM_FMT2 if col == 3 else None,
                )
        return

    ws['A1'] = '기간별 매매기준율'
    _style(ws['A1'], font=FONT_BOLD)
    ws['A2'] = f"기간 : {rate_data['period']}"

    ws['A4'] = '평균환율'
    _style(ws['A4'], font=FONT_BOLD)

    headers5 = ['평균환율', '최저치', '기록일', '최고치', '기록일', '등락폭', 'Cross Rate']
    for col, h in enumerate(headers5, 1):
        c = ws.cell(row=5, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    vals6 = [
        src(rate_data.get('average', '')),
        src(rate_data.get('min', '')),
        rate_data.get('min_date', ''),
        src(rate_data.get('max', '')),
        rate_data.get('max_date', ''),
        src(rate_data.get('range', '')),
        rate_data.get('cross_rate', ''),
    ]
    for col, v in enumerate(vals6, 1):
        c = ws.cell(row=6, column=col, value=v)
        _style(
            c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER,
            num_format=NUM_FMT2 if col in (1, 2, 4, 6, 7) and isinstance(v, (int, float)) else None,
        )

    ws['A7'] = '일별 매매기준율'
    _style(ws['A7'], font=FONT_BOLD)

    headers9 = ['날짜', '통화명', '환율', '전일대비', 'Cross Rate']
    for col, h in enumerate(headers9, 1):
        c = ws.cell(row=9, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    daily_rows = list(rate_data.get('daily', []))
    display_start = str(rate_data.get('display_start', '') or '')
    display_end = str(rate_data.get('display_end', '') or '')
    if display_start:
        daily_rows = [d for d in daily_rows if str(d.get('date', '')) >= display_start]
    if display_end:
        daily_rows = [d for d in daily_rows if str(d.get('date', '')) <= display_end]

    for r, d in enumerate(daily_rows, 10):
        vals = [d['date'], currency_name, src(d['rate']), src(d['change']), d['cross']]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            _style(
                c, font=FONT_DEFAULT, align=CENTER if col != 2 else LEFT, border=THIN_BORDER,
                num_format=NUM_FMT2 if col in (3, 4) else None,
            )

    monthly_rows = list(rate_data.get('monthly', []) or [])
    if monthly_rows:
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 24
        ws.column_dimensions['K'].width = 18

        ws.merge_cells(start_row=4, start_column=9, end_row=4, end_column=11)
        title_cell = ws.cell(row=4, column=9, value='월평균 매매기준율')
        _style(title_cell, font=FONT_BOLD, fill=SUBHEAD_FILL, align=CENTER, border=THIN_BORDER)
        for col in range(9, 12):
            ws.cell(row=4, column=col).border = THIN_BORDER

        monthly_headers = ['년월', '통화명', '월평균환율']
        for offset, header in enumerate(monthly_headers, 9):
            c = ws.cell(row=5, column=offset, value=header)
            _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

        for row_no, item in enumerate(monthly_rows, 6):
            values = [
                item.get('year_month', ''),
                currency_name,
                round(float(src(item.get('rate', 0)) or 0), 2),
            ]
            for offset, value in enumerate(values, 9):
                c = ws.cell(row=row_no, column=offset, value=value)
                _style(
                    c,
                    font=FONT_DEFAULT,
                    align=CENTER if offset != 10 else LEFT,
                    border=THIN_BORDER,
                    num_format=NUM_FMT2 if offset == 11 else None,
                )


# ── 쇼피 소포수령증 시트 작성 ───────────────────────────────────

def write_shopee_sheet(ws, shopee_data: dict, rates: dict, submitter: dict = None) -> int:
    """
    쇼피(MYR) 등 국가별 쇼피 시트 작성
    환율: 소포수령증 발행일(write_date) 기준
    """
    # S열은 원화 총합(천만 단위)이 들어가므로 13 이상이어야 ###이 나지 않습니다.
    col_widths = [16, 3, 3, 12, 3, 3, 20, 3, 3, 3, 5, 3, 5, 3, 3, 12, 3, 10, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    currency   = shopee_data.get('currency', '')
    carrier    = shopee_data.get('carrier', '주)두라로지스틱스')
    country    = shopee_data.get('country', '')
    period_end = shopee_data.get('period_end', '')
    divisor    = RATE_DIVISOR.get(currency, 1)

    # ── 행 1: 제목 헤더 ──
    sub = submitter or shopee_data.get('submitter') or DEFAULT_SUBMITTER
    ws.merge_cells('A1:J1')
    ws['A1'] = (
        f"해외배송 소포 수령증\n"
        f"Registration No. 117-81-45551\n"
        f"{sub.get('biz_no','')}\n{sub.get('ceo','')}"
    )
    _style(ws['A1'], font=FONT_BOLD, align=CENTER)
    ws.row_dimensions[1].height = 55

    ws.merge_cells('L1:S1')
    ws['L1'] = (
        f"해외배송기간: {shopee_data.get('period_start','')} ~ {period_end}\n"
        f"{sub.get('name','')}\n{sub.get('address','')}"
    )
    _style(ws['L1'], font=FONT_DEFAULT, align=LEFT)

    # ── 행 2-4: 인적사항 라벨 ──
    ws['A2'] = '사업자등록번호\n대표자 성명 거래기간'
    ws['A3'] = '상호(법인명) 작성일자'
    ws['A4'] = '사업장소재지'
    for row in [2, 3, 4]:
        _style(ws.cell(row=row, column=1), font=FONT_DEFAULT, align=LEFT)

    # ── 행 5: 거래기간, 작성일자 ──
    ws['A5'] = f"{shopee_data.get('period_start','')} ~ {period_end}"
    ws['I5'] = shopee_data.get('write_date', '')
    _style(ws['A5'], font=FONT_DEFAULT)
    _style(ws['I5'], font=FONT_DEFAULT)

    # ── 행 6: 섹션 2 제목 ──
    ws.merge_cells('A6:S6')
    ws['A6'] = '2. 해외배송 소포 수령 수량'
    _style(ws['A6'], font=FONT_BOLD, fill=SUBHEAD_FILL)

    # ── 행 7: 헤더 ──
    headers7 = {
        'A': '해외배송업체', 'D': '배송국가', 'G': '기간', 'K': '통화', 'M': '발송수량', 'P': '발송금액'
    }
    for col_letter, val in headers7.items():
        c = ws[f'{col_letter}7']
        c.value = val
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
    _merge_row(ws, 7, _RECEIPT_GROUPS_2, border=THIN_BORDER)
    for col_letter in headers7:
        _style(ws[f'{col_letter}7'], font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    # ── 행 8: 데이터 ──
    ws['A8'] = carrier
    ws['D8'] = country
    ws['G8'] = f"{shopee_data.get('period_start','')} ~ {period_end}"
    ws['K8'] = currency
    ws['M8'] = shopee_data.get('total_qty', 0)
    ws['P8'] = shopee_data.get('total_amount', 0.0)
    _merge_row(ws, 8, _RECEIPT_GROUPS_2, border=THIN_BORDER)
    for col in ['A', 'D', 'G', 'K', 'M', 'P']:
        nf = NUM_FMT if col == 'M' else (NUM_FMT2 if col == 'P' else None)
        _style(ws[f'{col}8'], font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER, num_format=nf)

    # ── 행 9: 합계 ──
    ws['M9'] = shopee_data.get('total_qty', 0)
    ws['P9'] = shopee_data.get('total_amount', 0.0)
    ws['G9'] = '합계'
    _merge_row(ws, 9, _RECEIPT_GROUPS_2, border=THIN_BORDER)
    _style(ws['G9'], font=FONT_BOLD, align=CENTER, border=THIN_BORDER)
    _style(ws['M9'], font=FONT_BOLD, align=CENTER, border=THIN_BORDER, num_format=NUM_FMT)
    _style(ws['P9'], font=FONT_BOLD, align=CENTER, border=THIN_BORDER, num_format=NUM_FMT2)

    # ── 행 10: 섹션 3 제목 ──
    ws.merge_cells('A10:O10')
    ws['A10'] = '3. 해외배송 내역'
    _style(ws['A10'], font=FONT_BOLD, fill=SUBHEAD_FILL)

    # ── 행 11: 컬럼 헤더 ──
    col_headers = {
        'A': '해외배송업체', 'D': '발행일', 'G': '운송장번호',
        'K': '도착국가', 'M': '발송수량', 'P': '수출신고금액', 'R': '환율', 'S': '원화'
    }
    for col_letter, val in col_headers.items():
        c = ws[f'{col_letter}11']
        c.value = val
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
    _merge_row(ws, 11, _RECEIPT_GROUPS_3, border=THIN_BORDER)
    for col_letter in col_headers:
        _style(ws[f'{col_letter}11'], font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    # ── 행 12+: 거래 데이터 (각 행의 발행일 기준 환율 개별 적용) ──
    row = 12
    total_krw = 0
    for tx in shopee_data.get('transactions', []):
        # 해당 거래의 발행일(tx['date']) 기준 환율 개별 조회
        tx_rate = _get_rate(rates, currency, tx['date'])
        krw = round(tx['amount'] * tx_rate / divisor)
        total_krw += krw

        ws.cell(row=row, column=1,  value=tx['carrier'])
        ws.cell(row=row, column=4,  value=tx['date'])
        ws.cell(row=row, column=7,  value=tx['tracking_no'])
        ws.cell(row=row, column=11, value=tx['country'])
        ws.cell(row=row, column=13, value=tx['qty'])
        ws.cell(row=row, column=16, value=tx['amount'])
        ws.cell(row=row, column=18, value=tx_rate)
        ws.cell(row=row, column=19, value=krw)

        _merge_row(ws, row, _RECEIPT_GROUPS_3, border=THIN_BORDER)
        for col in [1, 4, 7, 11, 13, 16, 18, 19]:
            c = ws.cell(row=row, column=col)
            nf = {13: NUM_FMT, 16: NUM_FMT2, 18: _applied_rate_format(currency), 19: NUM_FMT}.get(col)
            _style(c, font=FONT_DEFAULT, align=CENTER if col != 1 else LEFT, border=THIN_BORDER, num_format=nf)

        row += 1

    # 섹션 3 제목 오른쪽의 참고 합계 — 외화는 너비 3짜리 Q열 단독으로는 ###이 되므로
    # P10:Q10을 병합해 표시하고, 서식을 지정합니다. (섹션 제목 병합은 A10:O10이라 충돌 없음)
    ws.merge_cells('P10:Q10')
    ws['P10'] = shopee_data.get('total_amount', 0.0)
    ws['S10'] = total_krw
    _style(ws['P10'], font=FONT_DEFAULT, align=RIGHT, num_format=NUM_FMT2)
    _style(ws['S10'], font=FONT_DEFAULT, align=RIGHT, num_format=NUM_FMT)

    # ── 푸터 ──
    footer_row = row + 1
    ws.merge_cells(f'A{footer_row}:S{footer_row}')
    ws[f'A{footer_row}'] = '상기 내역은 판매자가 두라로지스틱스를 통하여 해외 배송한 내역임을 증명합니다'
    _style(ws[f'A{footer_row}'], font=FONT_DEFAULT, align=CENTER)

    footer_row += 1
    ws[f'A{footer_row}'] = '상호 (법인명)'
    ws[f'C{footer_row}'] = '두라로지스틱스'
    ws[f'H{footer_row}'] = '사업자 등록번호'

    footer_row += 1
    ws[f'A{footer_row}'] = '사업장 소재지'
    ws[f'C{footer_row}'] = '서울특별시 강서구 금낭화로 54-7 (방화동, 동해빌딩 1층)'

    footer_row += 1
    ws[f'A{footer_row}'] = '비고'
    ws[f'C{footer_row}'] = '본 증명서를 위조하거나 변조하는 등 모든 행위에 대한 책임은 판매자에게 있습니다'

    footer_row += 1
    ws[f'A{footer_row}'] = '(주)두라로지스틱스'

    return total_krw


# ── 통화별 수출신고 템플릿 시트 작성 ─────────────────────────────

def write_currency_template_sheet(ws, currency: str,
                                   shopee_data: Optional[object],
                                   lazada_items: list,
                                   rates: dict,
                                   lazada_write_date: str = '',
                                   lazada_rate_override: float = None,
                                   ebay_items: list = None,
                                   joom_items: list = None,
                                   shopify_items: list = None):
    """
    MYR, PHP, SGD 등 수출신고 프로그램용 시트 작성
    환율: 각 소포수령증 발행일(write_date) 기준

    1~5행은 플랫폼별 요약(쇼피·라자다·이베이·Joom·쇼피파이),
    6행이 헤더, 7행부터 거래 데이터입니다.
    """
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    # VND처럼 외화가 1억 단위(11자+콤마)인 통화가 있어 16 이상이 필요합니다.
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16

    # 라자다 PDF는 기간평균환율, 주문내역 Excel은 deliveredDate 일별환율을 사용합니다.
    lazada_avg_rate = _applied_rate_value(
        currency,
        lazada_rate_override if lazada_rate_override is not None
        else rates.get(currency, {}).get('average', 0.0),
    )
    divisor = RATE_DIVISOR.get(currency, 1)

    # ── 쇼피 소계: 같은 통화의 PDF가 여러 개여도 모두 합산합니다. ──
    if isinstance(shopee_data, list):
        shopee_items = [sd for sd in shopee_data if sd]
    elif shopee_data:
        shopee_items = [shopee_data]
    else:
        shopee_items = []
    shopee_items = sorted(shopee_items, key=_shopee_sort_key)

    shopee_transactions = []
    shopee_fx  = 0.0
    shopee_krw = 0
    for sd in shopee_items:
        txs = sd.get('transactions', []) or []
        if txs:
            shopee_transactions.extend(txs)
            for tx in txs:
                tx_rate = _get_rate(rates, currency, tx.get('date', ''))
                amount = float(tx.get('amount', 0) or 0)
                shopee_fx  += amount
                shopee_krw += round(amount * tx_rate / divisor)
        else:
            # 거래별 상세가 없을 때는 소계만 합산합니다.
            amount = float(sd.get('total_amount', 0) or 0)
            rate_date = sd.get('write_date') or sd.get('period_end') or ''
            tx_rate = _get_rate(rates, currency, rate_date)
            shopee_fx += amount
            shopee_krw += round(amount * tx_rate / divisor)

    shopee_transactions.sort(
        key=lambda tx: (_date_to_int(tx.get('date')) or 0, str(tx.get('tracking_no', '')))
    )

    # ── 라자다 소계 ──
    lazada_fx = 0.0
    lazada_krw = 0
    for it in lazada_items:
        amount = float(it.get('amount', 0) or 0)
        item_rate = _lazada_item_rate(it, currency, rates, avg_rate=lazada_avg_rate)
        lazada_fx += amount
        lazada_krw += round(amount * item_rate / divisor)

    # ── 이베이/린코스 소계: 발행월 기준 월평균 환율 사용 ──
    from .exchange_rate import monthly_avg_rate_for_month
    ebay_items = ebay_items or []
    ebay_fx = 0.0
    ebay_krw = 0
    for it in ebay_items:
        amount = float(it.get('amount', 0) or 0)
        month_key = it.get('month') or it.get('date') or it.get('period_end') or ''
        rate = _applied_rate_value(currency, monthly_avg_rate_for_month(rates.get(currency), month_key))
        ebay_fx += amount
        ebay_krw += round(amount * rate / divisor)

    # ── Joom / 쇼피파이 소계: 건별 기준일 일별 환율 사용 ──
    joom_items = sorted(joom_items or [], key=lambda it: (_date_to_int(it.get('date')) or 0,
                                                          str(it.get('order_id', ''))))
    shopify_items = sorted(shopify_items or [], key=lambda it: (_date_to_int(it.get('date')) or 0,
                                                                str(it.get('order_name', ''))))

    def _daily_subtotal(entries):
        fx = 0.0
        krw = 0
        for entry in entries:
            value = float(entry.get('amount', 0) or 0)
            entry_rate = _get_rate(rates, currency, entry.get('date', ''))
            fx += value
            krw += round(value * entry_rate / divisor)
        return fx, krw

    joom_fx, joom_krw = _daily_subtotal(joom_items)
    shopify_fx, shopify_krw = _daily_subtotal(shopify_items)

    total_krw = shopee_krw + lazada_krw + ebay_krw + joom_krw + shopify_krw

    # ── 행 1-5 플랫폼별 요약 ──
    summary_rows = [
        ('쇼피', shopee_fx, shopee_krw),
        ('라자다', lazada_fx, lazada_krw),
        ('이베이', ebay_fx, ebay_krw),
        ('Joom', joom_fx, joom_krw),
        ('쇼피파이', shopify_fx, shopify_krw),
    ]
    for row, (label, fx_value, krw_value) in enumerate(summary_rows, 1):
        ws.cell(row=row, column=5, value=label)
        ws.cell(row=row, column=6, value=fx_value)
        ws.cell(row=row, column=7, value=krw_value)
        for col in [5, 6, 7]:
            c = ws.cell(row=row, column=col)
            nf = NUM_FMT2 if col == 6 else (NUM_FMT if col == 7 else None)
            _style(c, font=FONT_DEFAULT, align=RIGHT, num_format=nf)

    # ── 행 6: 헤더 ──
    header_row = len(summary_rows) + 1
    headers = ['수출신고번호', '기타영세율건수', '선(기)적일자', '통화코드', '환율', '외화금액', '원화금액']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    # ── 행 7+: 쇼피 거래 (여러 PDF를 한 시트에 합쳐 날짜순 정렬) ──
    data_row = header_row + 1
    for tx in shopee_transactions:
        tx_rate  = _get_rate(rates, currency, tx.get('date', ''))
        amount   = float(tx.get('amount', 0) or 0)
        krw      = round(amount * tx_rate / divisor)
        date_int = _date_to_int(tx.get('date', ''))
        row_vals = ['', 1, date_int, currency, tx_rate, amount, krw]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=data_row, column=col, value=v)
            nf = {5: _applied_rate_format(currency), 6: NUM_FMT2, 7: NUM_FMT}.get(col)
            _style(c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER, num_format=nf)
        data_row += 1

    # ── 라자다 거래: 주문 Excel은 deliveredDate, 기존 PDF는 문서 기준일 ──
    for it in lazada_items:
        amount = float(it.get('amount', 0) or 0)
        item_date = _lazada_item_date(it, fallback=lazada_write_date)
        item_rate = _lazada_item_rate(it, currency, rates, avg_rate=lazada_avg_rate)
        krw = round(amount * item_rate / divisor)
        date_int_laz = _date_to_int(item_date)
        row_vals = ['', 1, date_int_laz, currency, item_rate, amount, krw]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=data_row, column=col, value=v)
            nf = {5: _applied_rate_format(currency), 6: NUM_FMT2, 7: NUM_FMT}.get(col)
            _style(c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER, num_format=nf)
        data_row += 1

    # ── 이베이/린코스 거래 (발행월 기준 월평균 환율 적용) ──
    for it in ebay_items:
        month_key = it.get('month') or it.get('date') or it.get('period_end') or ''
        rate = monthly_avg_rate_for_month(rates.get(currency), month_key)
        amount = float(it.get('amount', 0) or 0)
        krw = round(amount * rate / divisor)
        date_int_ebay = _date_to_int(it.get('date') or it.get('period_end') or '')
        row_vals = ['', 1, date_int_ebay, currency, rate, amount, krw]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=data_row, column=col, value=v)
            nf = {5: _applied_rate_format(currency), 6: NUM_FMT2, 7: NUM_FMT}.get(col)
            _style(c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER, num_format=nf)
        data_row += 1

    # ── Joom / 쇼피파이 거래 (건별 기준일 일별 환율 적용) ──
    for it in list(joom_items) + list(shopify_items):
        item_date = it.get('date', '')
        rate = _get_rate(rates, currency, item_date)
        amount = float(it.get('amount', 0) or 0)
        krw = round(amount * rate / divisor)
        row_vals = ['', 1, _date_to_int(item_date), currency, rate, amount, krw]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=data_row, column=col, value=v)
            nf = {5: _applied_rate_format(currency), 6: NUM_FMT2, 7: NUM_FMT}.get(col)
            _style(c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER, num_format=nf)
        data_row += 1


# ── 라자다 소포수령증 / 주문내역 시트 ────────────────────────────

def _copy_lazada_source_workbooks(ws, source_workbooks: list) -> bool:
    """라자다 국가별 원본 Excel 시트를 한 시트에 양식 그대로 세로로 이어 붙입니다.

    첫 번째 파일은 1행부터, 다음 파일은 한 줄을 비운 뒤 바로 아래에 붙입니다.
    셀 값·수식·서식·병합·행높이·열너비·숨김 상태를 가능한 범위에서 유지합니다.
    """
    valid_sources = [src for src in (source_workbooks or []) if src.get('content')]
    if not valid_sources:
        return False

    next_start_row = 1
    first_sheet = True
    for src_info in valid_sources:
        raw = src_info.get('content')
        if not isinstance(raw, (bytes, bytearray)):
            continue
        try:
            src_wb = openpyxl.load_workbook(BytesIO(raw), data_only=False)
        except Exception:
            continue

        requested = src_info.get('sheet_name')
        src_ws = src_wb[requested] if requested in src_wb.sheetnames else src_wb.worksheets[0]
        if not first_sheet:
            next_start_row += 1  # 원본 파일 블록 사이 빈 행
        row_offset = next_start_row - 1

        # 열 너비와 숨김 상태. 파일마다 차이가 있으면 더 넓은 값을 유지합니다.
        for col_idx in range(1, src_ws.max_column + 1):
            letter = get_column_letter(col_idx)
            src_dim = src_ws.column_dimensions[letter]
            dst_dim = ws.column_dimensions[letter]
            if src_dim.width is not None:
                dst_dim.width = max(dst_dim.width or 0, src_dim.width)
            if src_dim.hidden:
                dst_dim.hidden = True
            if src_dim.bestFit:
                dst_dim.bestFit = True
            if src_dim.outlineLevel:
                dst_dim.outlineLevel = src_dim.outlineLevel

        # 셀 값/수식과 개별 셀 서식을 복사합니다.
        for src_row in range(1, src_ws.max_row + 1):
            dst_row = row_offset + src_row
            src_row_dim = src_ws.row_dimensions[src_row]
            dst_row_dim = ws.row_dimensions[dst_row]
            if src_row_dim.height is not None:
                dst_row_dim.height = src_row_dim.height
            dst_row_dim.hidden = src_row_dim.hidden
            dst_row_dim.outlineLevel = src_row_dim.outlineLevel
            dst_row_dim.collapsed = src_row_dim.collapsed

            for col_idx in range(1, src_ws.max_column + 1):
                src_cell = src_ws.cell(src_row, col_idx)
                dst_cell = ws.cell(dst_row, col_idx, value=src_cell.value)
                if src_cell.has_style:
                    dst_cell._style = copy(src_cell._style)
                if src_cell.number_format:
                    dst_cell.number_format = src_cell.number_format
                if src_cell.alignment:
                    dst_cell.alignment = copy(src_cell.alignment)
                if src_cell.protection:
                    dst_cell.protection = copy(src_cell.protection)
                if src_cell.comment is not None:
                    dst_cell.comment = copy(src_cell.comment)
                if src_cell.hyperlink:
                    dst_cell._hyperlink = copy(src_cell.hyperlink)

        # 병합은 셀 복사 후에 행 오프셋을 적용하여 복원합니다.
        for merged in src_ws.merged_cells.ranges:
            ws.merge_cells(
                start_row=merged.min_row + row_offset,
                start_column=merged.min_col,
                end_row=merged.max_row + row_offset,
                end_column=merged.max_col,
            )

        if first_sheet:
            ws.freeze_panes = src_ws.freeze_panes
            ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
            if src_ws.sheet_format.defaultRowHeight is not None:
                ws.sheet_format.defaultRowHeight = src_ws.sheet_format.defaultRowHeight
            if src_ws.sheet_format.defaultColWidth is not None:
                ws.sheet_format.defaultColWidth = src_ws.sheet_format.defaultColWidth

        next_start_row = row_offset + src_ws.max_row + 1
        first_sheet = False

    return not first_sheet


# 미반영 사유 한글 표기 (시트 구분 열에 사용) — 부분일치 순서대로 검사합니다.
# 라자다 status는 'refund_completed' 같은 변형이 있어 짧은 키(refund/return/cancel)도 둡니다.
SKIP_REASON_LABELS = {
    'negative': '음수금액(검토)',
    'refund': '전액환불', 'void': '취소', 'unfulfilled': '미배송',
    'cancel': '취소', 'return': '반품',
}
SKIP_FONT = Font(name='맑은 고딕', size=9, color='C00000')


def _skip_reason_label(reason):
    text = str(reason or '').lower()
    for key, label in SKIP_REASON_LABELS.items():
        if key in text:
            return label
    return text or '미반영'


def write_lazada_order_sheet(ws, lazada_data: dict, rates: dict):
    """라자다 국가별 주문내역 Excel 원본을 양식 그대로 아래로 이어 붙입니다.

    미반영(취소·반품) 건은 통화별 상세 시트(라자다(XXX))에 정상 건과 같은 형식의
    행으로 표시되므로 여기서는 원본을 그대로 보존만 합니다.
    """
    if _copy_lazada_source_workbooks(ws, lazada_data.get('source_workbooks', [])):
        return

    # 구버전 파싱 결과처럼 원본 파일 바이트가 없을 때만 기존 집계형 표를 사용합니다.
    widths = {
        'A': 13, 'B': 10, 'C': 18, 'D': 20, 'E': 22, 'F': 18,
        'G': 10, 'H': 16, 'I': 12, 'J': 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws['A1'] = '라자다 주문내역 집계'
    _style(ws['A1'], font=FONT_TITLE, align=CENTER)
    ws.merge_cells('A1:J1')
    headers = ['배송완료일', '도착국', '주문번호', '주문상품번호', '송장번호', '배송사',
               '통화', 'Paid Price', '적용환율', '원화금액']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    for idx, it in enumerate(sorted(lazada_data.get('items', []), key=lambda item: (
            item.get('date') or '', item.get('currency') or '', item.get('tracking_no') or '')), 1):
        cur = it.get('currency', '')
        amount = float(it.get('amount', 0) or 0)
        rate = _lazada_item_rate(it, cur, rates, lazada_result=lazada_data)
        krw = round(amount * rate)
        vals = [
            _lazada_item_date(it, lazada_data), it.get('destination', ''), it.get('order_number', ''),
            it.get('order_item_id', ''), it.get('tracking_no', ''), it.get('carrier', ''),
            cur, amount, rate, krw,
        ]
        row = idx + 3
        for col, value in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=value)
            nf = {8: NUM_FMT2, 9: _applied_rate_format(cur), 10: NUM_FMT}.get(col)
            _style(c, font=FONT_DEFAULT, align=RIGHT if col in (8, 9, 10) else CENTER,
                   border=THIN_BORDER, num_format=nf)


def write_lazada_receipt_sheet(ws, lazada_data: dict, rates: dict, submitter: dict = None):
    """라자다 PDF 소포수령증 또는 주문내역 Excel 원본 시트."""
    if any(it.get('source_kind') == 'order_excel' for it in lazada_data.get('items', [])):
        return write_lazada_order_sheet(ws, lazada_data, rates)
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15

    carrier    = lazada_data.get('carrier', '용성종합물류')
    period_end = lazada_data.get('period_end', '')

    # 헤더
    ws['A1'] = (
        f'YONG SUNG LOGISTICS CO., LTD.\n'
        f'ROOM 1215, TOWER A 152, MAGOKSEO-RO, GANGSEO-GU, SEOUL, KOREA\n'
        f'TEL: 82-2-2664-4032  FAX: 82-2-2664-3815\n'
        f'E-mail : admin@yslogic.co.kr    URL : http://www.yslogic.co.kr'
    )
    _style(ws['A1'], font=FONT_DEFAULT, align=LEFT)
    ws.row_dimensions[1].height = 55

    ws['A2'] = '해외화물 소포 수령증'
    _style(ws['A2'], font=FONT_TITLE, align=CENTER)

    ws['A3'] = '1.   제출자 인적 사항'
    _style(ws['A3'], font=FONT_BOLD)

    sub = submitter or lazada_data.get('submitter') or DEFAULT_SUBMITTER
    info_rows = [
        ('사업자등록번호', sub.get('biz_no', ''), '상호(법인명)', sub.get('name', '')),
        ('성명(대표자)',   sub.get('ceo', ''),    '사업장소재지', sub.get('address', '')),
        ('거래기간',
         f"{lazada_data.get('period_start','')} – {period_end}",
         '작성일자', lazada_data.get('write_date', '')),
    ]
    for r, (k1, v1, k2, v2) in enumerate(info_rows, 4):
        ws.cell(row=r, column=1, value=k1)
        ws.cell(row=r, column=4, value=v1)
        ws.cell(row=r, column=9, value=k2)
        ws.cell(row=r, column=11, value=v2)

    ws['A7'] = '2.   해외 배송 내역서'
    _style(ws['A7'], font=FONT_BOLD)
    ws['A8'] = '발행사유'
    ws['B8'] = f'{carrier}를 통해 해외로 수출한 내역 증명'

    # 헤더행
    header_row = 9
    for col, h in enumerate(['서비스', '해외배송업체', '출발', '도착', '발송번호', '발송수량', '금액'], 1):
        c = ws.cell(row=header_row, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    # 데이터
    for r, it in enumerate(lazada_data.get('items', []), header_row + 1):
        row_vals = [
            it.get('service', '라자다'),
            it.get('carrier', carrier),
            it.get('origin', 'KR'),
            it.get('destination', ''),
            it.get('tracking_no', ''),
            f"{it.get('qty', '')}건",
            f"{it.get('amount', '')}({it.get('currency', '')})",
        ]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            _style(c, font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER)



def write_lazada_currency_detail_sheet(ws, currency: str, items: list, rates: dict, lazada_result: dict):
    """통화별 라자다 상세 시트. 원본파일 열 없이 deliveredDate 일별 환율을 사용합니다.

    미반영(취소·반품·전액환불) 건도 정상 건과 같은 형식의 행으로 표시하되
    비고 열과 회색 배경으로 구분하고, 합계에는 넣지 않습니다.
    """
    widths = {'A': 8, 'B': 13, 'C': 18, 'D': 20, 'E': 22, 'F': 18, 'G': 16, 'H': 12, 'I': 16, 'J': 12, 'K': 50}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws['A1'] = f'라자다 {currency} 상세'
    _style(ws['A1'], font=FONT_TITLE, align=CENTER)
    ws.merge_cells('A1:K1')
    headers = ['No', '배송완료일', '주문번호', '주문상품번호', '송장번호', '배송사', '외화금액', '환율', '원화금액', '비고', '상품명']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    avg_rate = None
    if lazada_result and lazada_result.get('period_start') and lazada_result.get('period_end'):
        from .exchange_rate import avg_rate_for_period
        avg_rate = avg_rate_for_period(rates.get(currency), lazada_result.get('period_start'), lazada_result.get('period_end'))

    total_fx = 0.0
    total_krw = 0
    row = 3
    for idx, it in enumerate(sorted(items, key=lambda x: (x.get('date') or '', x.get('tracking_no') or '')), 1):
        amount = float(it.get('amount', 0) or 0)
        rate = _lazada_item_rate(it, currency, rates, lazada_result, avg_rate)
        krw = round(amount * rate)
        vals = [idx, _lazada_item_date(it, lazada_result), it.get('order_number',''), it.get('order_item_id',''),
                it.get('tracking_no',''), it.get('carrier',''), amount, rate, krw, '', it.get('item_name','')]
        row += 1
        for col, value in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=value)
            nf = {7: NUM_FMT2, 8: _applied_rate_format(currency), 9: NUM_FMT}.get(col)
            _style(c, font=FONT_DEFAULT,
                   align=RIGHT if col in (7,8,9) else (LEFT if col == 11 else CENTER),
                   border=THIN_BORDER, num_format=nf)
        total_fx += amount
        total_krw += krw

    # 반영 건 바로 아래 합계, 그 아래 미반영 건을 사유(환불→반품→취소) 순으로 나열합니다.
    row += 1
    ws.cell(row, 1, '합계(매출 반영)')
    ws.cell(row, 7, total_fx)
    ws.cell(row, 9, total_krw)
    for col in range(1, 12):
        c = ws.cell(row, col)
        nf = {7: NUM_FMT2, 9: NUM_FMT}.get(col)
        _style(c, font=FONT_BOLD, fill=GRAY_FILL, align=RIGHT if col in (7,9) else CENTER, border=THIN_BORDER, num_format=nf)

    def _reason_rank(reason):
        text = str(reason or '').lower()
        if 'refund' in text:
            return 0
        if 'return' in text:
            return 1
        return 2  # cancel 등

    skipped = [it for it in (lazada_result or {}).get('skipped_items', [])
               if it.get('currency') == currency]
    skipped.sort(key=lambda x: (_reason_rank(x.get('skip_reason')),
                                x.get('date') or '', x.get('order_number') or ''))
    if skipped:
        row += 1  # 합계와 미반영 그룹 사이 한 줄 공백
        next_no = len(items)
        for it in skipped:
            next_no += 1
            row += 1
            vals = [next_no, it.get('date',''), it.get('order_number',''), it.get('order_item_id',''),
                    it.get('tracking_no',''), it.get('carrier',''), float(it.get('amount',0) or 0), None, None,
                    _skip_reason_label(it.get('skip_reason')), it.get('item_name','')]
            for col, value in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=value)
                nf = {7: NUM_FMT2}.get(col)
                _style(c, font=SKIP_FONT if col == 10 else FONT_DEFAULT, fill=GRAY_FILL,
                       align=RIGHT if col in (7,8,9) else (LEFT if col == 11 else CENTER),
                       border=THIN_BORDER, num_format=nf)


# ── 이베이/린코스 소포수령증 시트 ───────────────────────────────

def write_ebay_receipt_sheet(ws, ebay_data: dict, rates: dict, submitter: dict = None):
    """이베이 - 린코스 해외배송 소포 수령증 시트"""
    from .exchange_rate import monthly_avg_rate_for_month

    for col, width in {'A': 16, 'B': 18, 'C': 18, 'D': 12, 'E': 10, 'F': 14, 'G': 12, 'H': 14}.items():
        ws.column_dimensions[col].width = width

    sub = submitter or ebay_data.get('submitter') or DEFAULT_SUBMITTER
    ws['A1'] = '해외배송 소포 수령증 - 이베이'
    _style(ws['A1'], font=FONT_TITLE, align=CENTER)
    ws.merge_cells('A1:H1')

    info = [
        ('사업자등록번호', sub.get('biz_no',''), '상호(법인명)', sub.get('name','')),
        ('성명(대표자)', sub.get('ceo',''), '사업장소재지', sub.get('address','')),
        ('거래기간', f"{ebay_data.get('period_start','')} ~ {ebay_data.get('period_end','')}", '작성일자', ebay_data.get('write_date','')),
    ]
    r = 3
    for k1, v1, k2, v2 in info:
        vals = [k1, v1, k2, v2]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            _style(cell, font=FONT_BOLD if c in [1,3] else FONT_DEFAULT, fill=HEADER_FILL if c in [1,3] else None, align=CENTER if c in [1,3] else LEFT, border=THIN_BORDER)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='2. 해외배송 소포 수령증')
    _style(ws.cell(row=r, column=1), font=FONT_BOLD)
    r += 1
    headers = ['해외배송업체', '배송국가/Service', '현지송장번호', '통화단위', '발송수량', '신고금액']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        _style(cell, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
    for item in ebay_data.get('summary_items', []):
        r += 1
        vals = [item.get('carrier',''), item.get('service',''), item.get('tracking_no',''), item.get('currency',''), item.get('qty',0), item.get('amount',0)]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            nf = NUM_FMT2 if c == 6 else (NUM_FMT if c == 5 else None)
            _style(cell, font=FONT_DEFAULT, align=CENTER if c != 6 else RIGHT, border=THIN_BORDER, num_format=nf)

    r += 2
    ws.cell(row=r, column=1, value='3. 해외배송 내역서')
    _style(ws.cell(row=r, column=1), font=FONT_BOLD)
    r += 1
    headers = ['발행월', '해외배송업체', '배송국가/Service', '통화단위', '발송수량', '신고금액', '월평균환율', '원화금액']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        _style(cell, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
    for item in ebay_data.get('items', []):
        r += 1
        cur = item.get('currency','')
        rate = _applied_rate_value(cur, monthly_avg_rate_for_month(rates.get(cur), item.get('month') or ''))
        div = RATE_DIVISOR.get(cur, 1)
        amount = float(item.get('amount',0) or 0)
        krw = round(amount * rate / div)
        vals = [item.get('month',''), item.get('carrier',''), item.get('service',''), cur, item.get('qty',0), amount, rate, krw]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            nf = {5: NUM_FMT, 6: NUM_FMT2, 7: _applied_rate_format(cur), 8: NUM_FMT}.get(c)
            _style(cell, font=FONT_DEFAULT, align=CENTER if c not in [6,8] else RIGHT, border=THIN_BORDER, num_format=nf)

# ── Joom(에이치3네트웍스) 소포수령증 시트 ───────────────────────

def write_joom_sheet(ws, joom_data: dict, rates: dict, submitter: dict = None):
    """Joom [상품 수령 및 운송 확인증] 시트.

    ⚠️ 출력 필수항목 (사용자 지정 2026-07-30 — 지시 없이 삭제·변경 금지):
    서비스 / 발송날짜 / Order ID / 도착국가 / 접수번호 / 금액 / 적용환율 / 원화금액 / 상품명
    + 조회기간 해외배송 합계 행 + PDF 합계 표기 행.
    건별 발송날짜의 일별 매매기준율을 적용합니다.
    """
    widths = {'A': 16, 'B': 12, 'C': 14, 'D': 10, 'E': 26, 'F': 12, 'G': 12, 'H': 14, 'I': 50}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    sub = submitter or joom_data.get('submitter') or DEFAULT_SUBMITTER
    ws['A1'] = '상품 수령 및 운송 확인증 - Joom'
    _style(ws['A1'], font=FONT_TITLE, align=CENTER)
    ws.merge_cells('A1:I1')

    info = [
        ('사업자등록번호', sub.get('biz_no', ''), '법인명', sub.get('name', '')),
        ('대표자 성명', sub.get('ceo', ''), '법인 주소', sub.get('address', '')),
        ('조회기간', f"{joom_data.get('period_start','')} ~ {joom_data.get('period_end','')}",
         '작성일자', joom_data.get('write_date', '')),
    ]
    r = 3
    for k1, v1, k2, v2 in info:
        for c, v in enumerate([k1, v1, k2, v2], 1):
            cell = ws.cell(row=r, column=c, value=v)
            _style(cell, font=FONT_BOLD if c in (1, 3) else FONT_DEFAULT,
                   fill=HEADER_FILL if c in (1, 3) else None,
                   align=CENTER if c in (1, 3) else LEFT, border=THIN_BORDER)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='2. 해외배송 내역서')
    _style(ws.cell(row=r, column=1), font=FONT_BOLD)
    r += 1
    ws.cell(row=r, column=1, value='발행목적')
    ws.cell(row=r, column=2, value=f"{joom_data.get('carrier') or 'H3 NETWORKS'}를 통해 해외로 수출한 내역 증명")
    r += 1

    headers = ['서비스', '발송날짜', 'Order ID', '도착국가', '접수번호',
               '금액', '적용환율', '원화금액', '상품명']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        _style(cell, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    total_fx = 0.0
    total_krw = 0
    for item in joom_data.get('items', []):
        r += 1
        cur = item.get('currency', 'USD')
        rate = _get_rate(rates, cur, item.get('date', ''))
        amount = float(item.get('amount', 0) or 0)
        krw = round(amount * rate / RATE_DIVISOR.get(cur, 1))
        total_fx += amount
        total_krw += krw
        vals = [item.get('service', '해외배송서비스'), item.get('date', ''), item.get('order_id', ''),
                item.get('destination', ''), item.get('tracking_no', ''),
                amount, rate, krw, item.get('item_name', '')]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            nf = {6: NUM_FMT2, 7: _applied_rate_format(cur), 8: NUM_FMT}.get(c)
            _style(cell, font=FONT_DEFAULT,
                   align=RIGHT if c in (6, 7, 8) else (LEFT if c == 9 else CENTER),
                   border=THIN_BORDER, num_format=nf)

    r += 1
    ws.cell(row=r, column=1, value='조회기간 해외배송 합계')
    ws.cell(row=r, column=6, value=total_fx)
    ws.cell(row=r, column=8, value=total_krw)
    for c in range(1, 10):
        nf = {6: NUM_FMT2, 8: NUM_FMT}.get(c)
        _style(ws.cell(row=r, column=c), font=FONT_BOLD, fill=GRAY_FILL,
               align=RIGHT if c in (6, 8) else CENTER, border=THIN_BORDER, num_format=nf)

    # PDF 인쇄 합계와 나란히 표시해 순간 대조가 가능하게 합니다.
    declared = joom_data.get('declared_total') or {}
    if declared:
        r += 1
        ws.cell(row=r, column=1, value='PDF 합계 표기')
        ws.cell(row=r, column=6, value=sum(float(v or 0) for v in declared.values()))
        _style(ws.cell(row=r, column=1), font=FONT_DEFAULT, align=CENTER)
        _style(ws.cell(row=r, column=6), font=FONT_DEFAULT, align=RIGHT, num_format=NUM_FMT2)
        if joom_data.get('total_mismatch'):
            ws.cell(row=r, column=7, value='⚠️ 건별 합계와 불일치')
            _style(ws.cell(row=r, column=7), font=SKIP_FONT)


# ── 쇼피파이 주문내역 시트 ──────────────────────────────────────

def write_shopify_sheet(ws, shopify_data: dict, rates: dict):
    """쇼피파이 주문내역 시트.

    행 배치(사용자 지정 규칙):
      1) 매출 반영 건(paid·partially_refunded)을 먼저 모아서 표시
      2) 바로 아래 합계(매출 반영)
      3) 그 아래 refunded(전액환불) 건들
      4) 그 아래 voided(취소) 건들, 마지막으로 미배송 건들
    미반영 행은 구분 열 + 회색 배경으로 표시하고 환율·원화는 비웁니다.
    """
    widths = {'A': 6, 'B': 12, 'C': 12, 'D': 13, 'E': 18, 'F': 8, 'G': 12, 'H': 10, 'I': 14, 'J': 50}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    store = str(shopify_data.get('store') or '').strip()
    ws['A1'] = f'쇼피파이 주문내역 - {store}' if store else '쇼피파이 주문내역'
    _style(ws['A1'], font=FONT_TITLE, align=CENTER)
    ws.merge_cells('A1:J1')

    headers = ['No', '주문번호', '구분', '배송완료일', '결제상태', '통화', '외화금액', '환율', '원화금액', '상품명']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    def _order_key(it):
        return it.get('row_index') if it.get('row_index') is not None else 10 ** 9

    def _write_row(row, no, it, counted):
        cur = it.get('currency', 'USD')
        amount = float(it.get('amount', 0) or 0)
        if counted:
            rate = _get_rate(rates, cur, it.get('date', ''))
            krw = round(amount * rate / RATE_DIVISOR.get(cur, 1))
        else:
            rate = krw = None
        vals = [no, it.get('order_name', ''),
                '' if counted else _skip_reason_label(it.get('skip_reason')),
                it.get('date', ''), it.get('financial_status', ''), cur, amount, rate, krw,
                it.get('item_name', '')]
        for col, value in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=value)
            nf = {7: NUM_FMT2, 8: _applied_rate_format(cur), 9: NUM_FMT}.get(col)
            _style(c,
                   font=SKIP_FONT if (not counted and col == 3) else FONT_DEFAULT,
                   fill=None if counted else GRAY_FILL,
                   align=RIGHT if col in (7, 8, 9) else (LEFT if col == 10 else CENTER),
                   border=THIN_BORDER, num_format=nf)
        return amount, krw

    # 1) 매출 반영 건
    total_fx = 0.0
    total_krw = 0
    row = 3
    no = 0
    for it in sorted(shopify_data.get('items', []), key=_order_key):
        row += 1
        no += 1
        amount, krw = _write_row(row, no, it, counted=True)
        total_fx += amount
        total_krw += krw

    # 2) 합계(매출 반영)
    row += 1
    ws.cell(row, 1, '합계(매출 반영)')
    ws.cell(row, 7, total_fx)
    ws.cell(row, 9, total_krw)
    for col in range(1, 11):
        c = ws.cell(row, col)
        nf = {7: NUM_FMT2, 9: NUM_FMT}.get(col)
        _style(c, font=FONT_BOLD, fill=GRAY_FILL, align=RIGHT if col in (7, 9) else CENTER,
               border=THIN_BORDER, num_format=nf)

    # 3) 미반영 건 — refunded → voided → 미배송 순으로 그룹
    reason_rank = {'refunded': 0, 'voided': 1, 'unfulfilled': 2}
    skipped = sorted(shopify_data.get('skipped_items', []),
                     key=lambda it: (reason_rank.get(str(it.get('skip_reason')), 9), _order_key(it)))
    if skipped:
        row += 1  # 합계와 미반영 그룹 사이 한 줄 공백
        for it in skipped:
            row += 1
            no += 1
            _write_row(row, no, it, counted=False)



# ── 큐텐 소포수령증 시트 ────────────────────────────────────────

def write_qoo10_sheet(ws, qoo10_data: Optional[dict], jpy_rate: float, submitter: dict = None):
    """
    큐텐(소포수령증) 시트
    jpy_rate: 큐텐 반기말(6월/12월) 공식 월평균환율의 대표값 (1엔 기준)
    """
    # 금액(JPY)·원화가 수천만 단위까지 가므로 명시 너비가 없으면 ###이 됩니다.
    for _col, _w in {'A': 16, 'B': 14, 'C': 12, 'D': 12, 'E': 18,
                     'F': 10, 'G': 14, 'H': 12, 'I': 14}.items():
        ws.column_dimensions[_col].width = _w

    ws['A1'] = '해외배송 소포 수령증'
    _style(ws['A1'], font=FONT_TITLE, align=CENTER)

    ws['A3'] = '1.제출자 인적사항'
    _style(ws['A3'], font=FONT_BOLD)

    sub = submitter or DEFAULT_SUBMITTER
    ws['A5'] = '사업자등록번호'; ws['B5'] = sub.get('biz_no', '')
    ws['C5'] = '상호（법인명）'; ws['D5'] = sub.get('name', '')
    ws['A6'] = '성명 （대표자）'; ws['B6'] = sub.get('ceo', '')
    ws['C6'] = '사업장소재지'; ws['D6'] = sub.get('address', '')
    ws['A7'] = '거래기간'

    if qoo10_data:
        period = f"{qoo10_data.get('period_start','')} ~ {qoo10_data.get('period_end','')}"
        ws['B7'] = period

    ws['A9'] = '2.해외배송 소포 수령증'
    _style(ws['A9'], font=FONT_BOLD)

    headers = ['판매처', '해외배송업체', '배송국가', '송장번호', '수량', '비고']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=11, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    if qoo10_data:
        ws.cell(row=12, column=1, value='Qoo10')
        ws.cell(row=12, column=2, value='국제로지스틱')
        ws.cell(row=12, column=3, value='일본')
        ws.cell(row=12, column=4, value=qoo10_data.get('tracking_no', ''))
        ws.cell(row=12, column=5, value=f"{qoo10_data.get('qty', '')} 건")

        ws.cell(row=13, column=1, value='당기 해외배송 합계')
        ws.cell(row=13, column=5, value=f"{qoo10_data.get('qty', '')} 건")

        ws['A15'] = '3. 해외배송 내역서'
        _style(ws['A15'], font=FONT_BOLD)
        ws['A17'] = '발행사유'; ws['B17'] = '국제로지스틱을 통해 해외로 수출한 내역 증명'

        detail_headers = ['판매처', '해외배송업체', '출발', '도착', '발송번호', '발송수량', '금액 (JPY)', '적용환율', '원화금액']
        for col, h in enumerate(detail_headers, 1):
            c = ws.cell(row=18, column=col, value=h)
            _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

        entries = qoo10_data.get('entries') or [{
            'tracking_no': qoo10_data.get('tracking_no', ''),
            'qty':    qoo10_data.get('qty', 0),
            'amount': qoo10_data.get('amount', 0),
            'rate':   jpy_rate,
            'krw':    round(qoo10_data.get('amount', 0) * jpy_rate),
        }]

        ws.cell(row=17, column=8, value='반기말 월평균환율 (1엔 기준)')

        r = 19
        total_jpy = 0
        total_krw = 0
        total_qty = 0
        for e in entries:
            e_rate = _applied_rate_value('JPY', e.get('rate', jpy_rate))
            e_amt  = e.get('amount', 0)
            e_krw  = e.get('krw', round(e_amt * e_rate))
            e_qty  = e.get('qty', 0)
            ws.cell(row=r, column=1, value='Qoo10')
            ws.cell(row=r, column=2, value='국제로지스틱')
            ws.cell(row=r, column=3, value='KR')
            ws.cell(row=r, column=4, value='JP')
            ws.cell(row=r, column=5, value=e.get('tracking_no', ''))
            ws.cell(row=r, column=6, value=f"{e_qty} 건")
            ws.cell(row=r, column=7, value=e_amt)
            ws.cell(row=r, column=8, value=e_rate)
            ws.cell(row=r, column=9, value=e_krw)
            for col in range(1, 10):
                nf = NUM_FMT if col in (7, 9) else (_applied_rate_format('JPY') if col == 8 else None)
                _style(ws.cell(row=r, column=col), font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER, num_format=nf)
            total_jpy += e_amt
            total_krw += e_krw
            total_qty += e_qty
            r += 1

        ws.cell(row=r, column=1, value='당기 해외배송 합계')
        ws.cell(row=r, column=6, value=f"{total_qty} 건")
        ws.cell(row=r, column=7, value=total_jpy)
        ws.cell(row=r, column=8, value=_applied_rate_value('JPY', jpy_rate))
        ws.cell(row=r, column=9, value=total_krw)
        for col in range(1, 10):
            nf = NUM_FMT if col in (7, 9) else (_applied_rate_format('JPY') if col == 8 else None)
            _style(ws.cell(row=r, column=col), font=FONT_BOLD, align=CENTER, border=THIN_BORDER, num_format=nf)
    else:
        ws['A12'] = '⚠️ 큐텐 데이터 없음 — STEP 2에서 수동 입력하세요'
        _style(ws['A12'], font=Font(name='맑은 고딕', size=9, color='FF0000'))


# ── 총집계 시트 ─────────────────────────────────────────────────

def write_summary_sheet(ws, shopee_totals: dict, lazada_totals: dict,
                         qoo10_data: Optional[dict], jpy_rate: float,
                         year_month: str, submitter: dict = None,
                         ebay_totals: dict = None, joom_totals: dict = None,
                         shopify_totals: dict = None):
    """총집계 시트 작성.

    실제 집계된 플랫폼만 표시합니다. 예를 들어 이베이 자료만 처리한 경우
    쇼피/라자다/큐텐 구역은 만들지 않습니다. 각 플랫폼 안에서도 실제로
    집계된 통화만 행으로 출력합니다.
    """
    NUM  = '#,##0'
    NUM2 = '#,##0.00'
    ws.column_dimensions['B'].width = 16
    # 외화 열은 VND 1억 단위(콤마 포함 11자 이상)까지 수용해야 합니다.
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    sub = submitter or DEFAULT_SUBMITTER
    ws['B1'] = _submitter_label(sub)
    _style(ws['B1'], font=FONT_TITLE)
    ws['D2'] = year_month
    _style(ws['D2'], font=FONT_BOLD)

    def _sub(r, label):
        c = ws.cell(row=r, column=2, value=label)
        _style(c, font=FONT_BOLD, fill=SUBHEAD_FILL, align=CENTER)

    def _hdr3(r, h1, h2, h3):
        for col, val in [(2, h1), (3, h2), (4, h3)]:
            c = ws.cell(row=r, column=col, value=val)
            _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    def _datarow(r, name, fx, krw):
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=fx)
        ws.cell(row=r, column=4, value=krw)
        _style(ws.cell(row=r, column=2), font=FONT_DEFAULT, align=LEFT,  border=THIN_BORDER)
        _style(ws.cell(row=r, column=3), font=FONT_DEFAULT, align=RIGHT, border=THIN_BORDER, num_format=NUM2)
        _style(ws.cell(row=r, column=4), font=FONT_DEFAULT, align=RIGHT, border=THIN_BORDER, num_format=NUM)

    def _totalrow(r, krw):
        ws.cell(row=r, column=2, value='총합')
        ws.cell(row=r, column=4, value=krw)
        _style(ws.cell(row=r, column=2), font=FONT_BOLD, align=LEFT,  border=THIN_BORDER, fill=GRAY_FILL)
        _style(ws.cell(row=r, column=3), border=THIN_BORDER, fill=GRAY_FILL)
        _style(ws.cell(row=r, column=4), font=FONT_BOLD, align=RIGHT, border=THIN_BORDER, fill=GRAY_FILL, num_format=NUM)

    COUNTRY_NAMES = {
        'MYR': '말레이시아(MYR)', 'PHP': '필리핀(PHP)',
        'SGD': '싱가폴(SGD)', 'THB': '태국(THB)',
        'TWD': '대만(TWD)', 'VND': '베트남(VND)', 'IDR': '인도네시아(IDR)',
        'BRL': '브라질(BRL)', 'MXN': '멕시코(MXN)',
        'USD': '미국(USD)', 'EUR': '유로(EUR)', 'GBP': '영국(GBP)',
        'CAD': '캐나다(CAD)', 'AUD': '호주(AUD)',
    }

    # 실제 집계된 통화만 대상으로 삼습니다. totals 딕셔너리는 실제 입력
    # 자료가 있을 때만 통화 키가 생성되므로 값이 0인 정상 거래도 유지됩니다.
    shopee_totals = shopee_totals or {}
    lazada_totals = lazada_totals or {}
    ebay_totals = ebay_totals or {}
    joom_totals = joom_totals or {}
    # 쇼피파이는 스토어 단위로 파일이 오므로 '스토어(통화)' 키를 그대로 표시합니다.
    shopify_totals = shopify_totals or {}

    row = 5
    written_platforms = 0

    if shopee_totals:
        _sub(row, '쇼피')
        _hdr3(row + 1, '국가', '외화', '원화')
        data_row = row + 2
        total_krw = 0
        for cur in _ordered_currencies(shopee_totals.keys()):
            data = shopee_totals.get(cur, {})
            fx = data.get('fx', 0.0)
            krw = data.get('krw', 0)
            total_krw += krw
            _datarow(data_row, COUNTRY_NAMES.get(cur, cur), fx, krw)
            data_row += 1
        _totalrow(data_row, total_krw)
        row = data_row + 2
        written_platforms += 1

    if lazada_totals:
        _sub(row, '라자다')
        _hdr3(row + 1, '국가', '외화', '원화')
        data_row = row + 2
        total_krw = 0
        for cur in _ordered_currencies(lazada_totals.keys()):
            data = lazada_totals.get(cur, {})
            fx = data.get('fx', 0.0)
            krw = data.get('krw', 0)
            total_krw += krw
            _datarow(data_row, COUNTRY_NAMES.get(cur, cur), fx, krw)
            data_row += 1
        _totalrow(data_row, total_krw)
        row = data_row + 2
        written_platforms += 1

    if ebay_totals:
        _sub(row, '이베이')
        _hdr3(row + 1, '통화', '외화', '원화')
        data_row = row + 2
        total_krw = 0
        for cur in _ordered_currencies(ebay_totals.keys()):
            data = ebay_totals.get(cur, {})
            fx = data.get('fx', 0.0)
            krw = data.get('krw', 0)
            total_krw += krw
            _datarow(data_row, COUNTRY_NAMES.get(cur, cur), fx, krw)
            data_row += 1
        _totalrow(data_row, total_krw)
        row = data_row + 2
        written_platforms += 1

    if joom_totals:
        _sub(row, 'Joom')
        _hdr3(row + 1, '통화', '외화', '원화')
        data_row = row + 2
        total_krw = 0
        for cur in _ordered_currencies(joom_totals.keys()):
            data = joom_totals.get(cur, {})
            krw = data.get('krw', 0)
            total_krw += krw
            _datarow(data_row, COUNTRY_NAMES.get(cur, cur), data.get('fx', 0.0), krw)
            data_row += 1
        _totalrow(data_row, total_krw)
        row = data_row + 2
        written_platforms += 1

    if shopify_totals:
        _sub(row, '쇼피파이')
        _hdr3(row + 1, '스토어(통화)', '외화', '원화')
        data_row = row + 2
        total_krw = 0
        for key in sorted(shopify_totals.keys()):
            data = shopify_totals.get(key, {})
            krw = data.get('krw', 0)
            total_krw += krw
            _datarow(data_row, key, data.get('fx', 0.0), krw)
            data_row += 1
        _totalrow(data_row, total_krw)
        row = data_row + 2
        written_platforms += 1

    if qoo10_data:
        _sub(row, '큐텐')
        _hdr3(row + 1, '외화', '평균환율', '원화')
        jpy_amount = qoo10_data.get('amount', 0)
        krw = qoo10_data.get('total_krw') or round(jpy_amount * jpy_rate)
        # 평균환율 = 실효환율(원화÷외화) — 1엔 기준
        eff_rate = _applied_rate_value('JPY', krw / jpy_amount if jpy_amount else jpy_rate)
        data_row = row + 2
        ws.cell(row=data_row, column=2, value=jpy_amount)
        ws.cell(row=data_row, column=3, value=eff_rate)
        ws.cell(row=data_row, column=4, value=krw)
        _style(ws.cell(row=data_row, column=2), font=FONT_DEFAULT, align=RIGHT, border=THIN_BORDER, num_format=NUM)
        _style(ws.cell(row=data_row, column=3), font=FONT_DEFAULT, align=RIGHT, border=THIN_BORDER, num_format=_applied_rate_format('JPY'))
        _style(ws.cell(row=data_row, column=4), font=FONT_DEFAULT, align=RIGHT, border=THIN_BORDER, num_format=NUM)
        written_platforms += 1

    # 방어 로직: 데이터 없는 상태에서도 워크북 구조가 깨지지 않게 안내만 표시합니다.
    if written_platforms == 0:
        ws['B5'] = '집계된 플랫폼이 없습니다.'
        _style(ws['B5'], font=FONT_DEFAULT)



# ── 사용 데이터 기준 시트 정리 유틸 ───────────────────────────────
PREFERRED_CURRENCY_ORDER = ['MYR', 'PHP', 'SGD', 'THB', 'TWD', 'VND', 'IDR', 'BRL', 'MXN', 'USD', 'EUR', 'GBP', 'CAD', 'AUD', 'JPY']
SHOPEE_SHEET_NAMES = {
    'MYR': '쇼피(MYR)', 'PHP': '쇼피(PHP)', 'SGD': '쇼피(SGD)',
    'THB': '쇼피(THB)', 'TWD': '쇼피(TWD)', 'VND': '쇼피(VND)', 'IDR': '쇼피(IDR)',
    'BRL': '쇼피(BRL)', 'MXN': '쇼피(MXN)',
}
LAZADA_CURRENCY_ORDER = ['MYR', 'PHP', 'SGD', 'VND', 'IDR']


def _ordered_currencies(values):
    """PREFERRED_CURRENCY_ORDER 기준으로 통화코드를 정렬합니다."""
    values = {str(v or '').upper() for v in values if v}
    return [cur for cur in PREFERRED_CURRENCY_ORDER if cur in values] + sorted(values - set(PREFERRED_CURRENCY_ORDER))


def _shopee_sort_key(sd):
    """같은 통화의 쇼피 PDF가 여러 개일 때 기간순으로 정렬합니다."""
    return (
        _date_to_int(sd.get('period_start')) or 0,
        _date_to_int(sd.get('period_end')) or 0,
        _date_to_int(sd.get('write_date')) or 0,
    )


def _shopee_items_for_currency(shopee_results, currency):
    """해당 통화의 쇼피 PDF 결과를 기간순으로 반환합니다."""
    return sorted(
        [s for s in (shopee_results or [])
         if s.get('currency') == currency and _has_shopee_data(s)],
        key=_shopee_sort_key,
    )


def _base_shopee_sheet_name(currency):
    return SHOPEE_SHEET_NAMES.get(currency, f'쇼피({currency})')


def _shopee_sheet_names_for_results(shopee_results, shopee_currencies):
    """
    쇼피 원본 PDF 시트명을 생성합니다.
    같은 통화의 PDF가 여러 개면 쇼피(MYR)(1), 쇼피(MYR)(2)처럼 번호를 붙입니다.
    한 개뿐이면 기존처럼 쇼피(MYR)를 유지합니다.
    """
    out = []
    for cur in shopee_currencies:
        items = _shopee_items_for_currency(shopee_results, cur)
        base = _base_shopee_sheet_name(cur)
        if len(items) <= 1:
            if items:
                out.append(base)
        else:
            out.extend(f'{base}({idx})' for idx, _sd in enumerate(items, 1))
    return out


def _has_shopee_data(sd):
    if not sd:
        return False
    if sd.get('transactions'):
        return True
    return bool(sd.get('currency') and (sd.get('total_qty', 0) or sd.get('total_amount', 0)))


def _has_lazada_data(lazada_result):
    return bool(lazada_result and lazada_result.get('items'))


def _has_ebay_data(ebay_result):
    return bool(ebay_result and ebay_result.get('items'))


def _ebay_sheet_names_for_results(ebay_results):
    items = [er for er in (ebay_results or []) if _has_ebay_data(er)]
    if not items:
        return []
    if len(items) == 1:
        return ['이베이']
    return [f'이베이({i})' for i, _ in enumerate(items, 1)]


def _has_joom_data(joom_result):
    return bool(joom_result and joom_result.get('items'))


def _joom_sheet_names_for_results(joom_results):
    items = [jr for jr in (joom_results or []) if _has_joom_data(jr)]
    if not items:
        return []
    if len(items) == 1:
        return ['Joom']
    return [f'Joom({i})' for i, _ in enumerate(items, 1)]


def _has_shopify_data(shopify_result):
    return bool(shopify_result and shopify_result.get('items'))


def _shopify_sheet_names_for_results(shopify_results):
    """스토어명으로 시트를 만들고 중복되면 번호를 붙입니다."""
    names = []
    used = {}
    for sr in (shopify_results or []):
        if not _has_shopify_data(sr):
            continue
        # 엑셀 시트명에 쓸 수 없는 문자와 31자 제한을 정리합니다.
        store = re.sub(r'[\[\]:*?/\\]', '_', str(sr.get('store') or '').strip()) or '주문내역'
        base = f'쇼피파이({store})'[:31]
        used[base] = used.get(base, 0) + 1
        names.append(base if used[base] == 1 else f'{base[:28]}({used[base]})')
    return names


def _shopify_total_key(shopify_result, currency):
    store = str((shopify_result or {}).get('store') or '').strip() or '주문내역'
    return f'{store}({currency})'


def _has_qoo10_data(qoo10_result):
    if not qoo10_result:
        return False
    if qoo10_result.get('entries'):
        return True
    return bool(qoo10_result.get('qty', 0) or qoo10_result.get('amount', 0) or qoo10_result.get('tracking_no'))


def _infer_used_sources_and_currencies(shopee_results, lazada_result, qoo10_result, ebay_results=None,
                                       joom_results=None, shopify_results=None):
    """
    실제 입력 데이터가 있는 소스/통화만 추려냅니다.
    이 결과를 기준으로 불필요한 쇼피/라자다/큐텐/이베이/Joom/쇼피파이/환율/통화시트를 만들지 않습니다.
    """
    shopee_currencies = {sd.get('currency') for sd in (shopee_results or []) if _has_shopee_data(sd)}
    lazada_currencies = set()
    if _has_lazada_data(lazada_result):
        lazada_currencies = {it.get('currency') for it in lazada_result.get('items', []) if it.get('currency')}
    ebay_currencies = set()
    for er in (ebay_results or []):
        if _has_ebay_data(er):
            ebay_currencies.update(it.get('currency') for it in er.get('items', []) if it.get('currency'))
    joom_currencies = set()
    for jr in (joom_results or []):
        if _has_joom_data(jr):
            joom_currencies.update(it.get('currency') for it in jr.get('items', []) if it.get('currency'))
    shopify_currencies = set()
    for sr in (shopify_results or []):
        if _has_shopify_data(sr):
            shopify_currencies.update(it.get('currency') for it in sr.get('items', []) if it.get('currency'))
    qoo10_used = _has_qoo10_data(qoo10_result)
    used_currencies = (set(shopee_currencies) | set(lazada_currencies) | set(ebay_currencies)
                       | set(joom_currencies) | set(shopify_currencies))
    if qoo10_used:
        used_currencies.add('JPY')
    return {
        'shopee_currencies': _ordered_currencies(shopee_currencies),
        'lazada_currencies': _ordered_currencies(lazada_currencies),
        'ebay_currencies': _ordered_currencies(ebay_currencies),
        'joom_currencies': _ordered_currencies(joom_currencies),
        'shopify_currencies': _ordered_currencies(shopify_currencies),
        'qoo10_used': qoo10_used,
        'used_currencies': _ordered_currencies(used_currencies),
    }


def _prune_workbook_sheets(wb, keep_sheet_names):
    """혹시 생성 과정에서 남은 불필요 시트를 최종적으로 삭제합니다."""
    keep = set(keep_sheet_names)
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in keep:
            del wb[sheet_name]

# ── 월별집계 시트 ───────────────────────────────────────────────



def _qoo10_reporting_date(entry=None, result=None):
    """큐텐 신고/집계 기준일: 거래기간 종료일, 없으면 해당 반기말."""
    entry = entry or {}
    result = result or {}
    period_end = entry.get('period_end') or result.get('period_end') or ''
    digits = re.sub(r'\D', '', str(period_end))[:8]
    if len(digits) == 8:
        return f"{digits[:4]}-{digits[4:6]}-{digits[6:8]}"

    base = (entry.get('period_start') or result.get('period_start')
            or entry.get('write_date') or result.get('write_date') or '')
    digits = re.sub(r'\D', '', str(base))[:8]
    if len(digits) >= 6:
        year = digits[:4]
        month = int(digits[4:6])
        return f"{year}-06-30" if month <= 6 else f"{year}-12-31"
    return ''

def _qoo10_reporting_month(entry=None, result=None):
    """큐텐 거래기간 기준 반기말 월(YYYY-06 또는 YYYY-12)을 반환합니다.

    큐텐만 예외적으로 반기말에 환율을 모아 적용하므로, 신고 기준일이
    10월이어도 환율 조회월은 12월(하반기말)입니다. app.py의 수집 로직과
    같은 월을 돌려줘야 '수집은 12월, 조회는 10월'로 어긋나지 않습니다.
    """
    report_date = _qoo10_reporting_date(entry, result)
    digits = re.sub(r"\D", "", str(report_date or ""))[:8]
    if len(digits) >= 6:
        year = digits[:4]
        month = int(digits[4:6])
        return f"{year}-06" if month <= 6 else f"{year}-12"
    return ""


def _month_label_from_date(value):
    """날짜 문자열/숫자에서 'YYYY년 MM월' 라벨을 반환합니다."""
    d = re.sub(r"\D", "", str(value or ""))[:8]
    if len(d) >= 6:
        return f"{d[:4]}년 {d[4:6]}월"
    return "날짜미상"


def write_monthly_summary_sheet(ws, shopee_results: list, lazada_result: Optional[dict],
                                qoo10_result: Optional[dict], rates: dict,
                                lazada_avg_rates: dict = None,
                                lazada_write_date: str = '',
                                jpy_rate: float = 0.0,
                                submitter: dict = None,
                                ebay_results: list = None,
                                joom_results: list = None,
                                shopify_results: list = None):
    """
    월별집계 시트 작성.
    기준일은 각 문서의 수출실적/통화 시트에 들어가는 선(기)적일자와 동일하게 봅니다.
    - 쇼피: 거래별 발행일(tx['date'])
    - 라자다 주문 Excel: deliveredDate / 기존 PDF: 거래기간 종료일(없으면 작성일자)
    - 큐텐: 입력 건별 거래기간 종료일(상반기 6월 말/하반기 12월 말)
    """
    NUM = '#,##0'
    NUM2 = '#,##0.00'
    lazada_avg_rates = lazada_avg_rates or {}

    for col, width in {'A': 13, 'B': 12, 'C': 10, 'D': 10, 'E': 16, 'F': 16}.items():
        ws.column_dimensions[col].width = width

    sub = submitter or DEFAULT_SUBMITTER
    ws['A1'] = f"월별집계 - {_submitter_label(sub)}"
    _style(ws['A1'], font=FONT_TITLE)
    ws.merge_cells('A1:F1')

    headers = ['월', '구분', '통화코드', '건수', '외화금액', '원화금액']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

    rows = []

    # 쇼피: 거래별 발행일 기준
    for sd in shopee_results or []:
        cur = sd.get('currency', '')
        if not cur:
            continue
        div = RATE_DIVISOR.get(cur, 1)
        txs = sd.get('transactions') or []
        if txs:
            for tx in txs:
                amount = float(tx.get('amount', 0) or 0)
                qty = int(tx.get('qty', 0) or 0)
                rate = _get_rate(rates, cur, tx.get('date', ''))
                krw = round(amount * rate / div)
                rows.append({
                    'month': _month_label_from_date(tx.get('date', '')),
                    'source': '쇼피', 'currency': cur, 'qty': qty,
                    'fx': amount, 'krw': krw,
                })
        else:
            amount = float(sd.get('total_amount', 0) or 0)
            qty = int(sd.get('total_qty', 0) or 0)
            rate_date = sd.get('write_date') or sd.get('period_end') or ''
            rate = _get_rate(rates, cur, rate_date)
            krw = round(amount * rate / div)
            rows.append({
                'month': _month_label_from_date(rate_date),
                'source': '쇼피', 'currency': cur, 'qty': qty,
                'fx': amount, 'krw': krw,
            })

    # 라자다: 주문 Excel은 deliveredDate 기준, 기존 PDF는 문서 기준일
    if lazada_result and lazada_result.get('items'):
        fallback_date = lazada_write_date or lazada_result.get('period_end') or lazada_result.get('write_date') or ''
        for it in lazada_result.get('items', []):
            cur = it.get('currency', '')
            if not cur:
                continue
            amount = float(it.get('amount', 0) or 0)
            qty = int(it.get('qty', 0) or 0)
            div = RATE_DIVISOR.get(cur, 1)
            date_value = _lazada_item_date(it, lazada_result, fallback_date)
            rate = _lazada_item_rate(
                it, cur, rates, lazada_result, lazada_avg_rates.get(cur)
            )
            krw = round(amount * rate / div)
            rows.append({
                'month': _month_label_from_date(date_value),
                'source': '라자다', 'currency': cur, 'qty': qty,
                'fx': amount, 'krw': krw,
            })

    # 이베이/린코스: 발행월 기준 월평균 환율
    from .exchange_rate import monthly_avg_rate_for_month
    for er in ebay_results or []:
        for it in er.get('items', []):
            cur = it.get('currency', '')
            if not cur:
                continue
            amount = float(it.get('amount', 0) or 0)
            qty = int(it.get('qty', 0) or 0)
            div = RATE_DIVISOR.get(cur, 1)
            rate = _applied_rate_value(cur, monthly_avg_rate_for_month(rates.get(cur), it.get('month') or it.get('date') or ''))
            krw = round(amount * rate / div)
            rows.append({
                'month': _month_label_from_date((it.get('month') or '').replace('-', '') + '01'),
                'source': '이베이', 'currency': cur, 'qty': qty,
                'fx': amount, 'krw': krw,
            })

    # Joom / 쇼피파이: 건별 기준일(발송날짜 · Fulfilled at) 일별 환율
    for source_label, results in (('Joom', joom_results or []), ('쇼피파이', shopify_results or [])):
        for res in results:
            for it in res.get('items', []):
                cur = it.get('currency', '')
                if not cur:
                    continue
                amount = float(it.get('amount', 0) or 0)
                qty = int(it.get('qty', 1) or 0)
                div = RATE_DIVISOR.get(cur, 1)
                item_date = it.get('date', '')
                rate = _get_rate(rates, cur, item_date)
                rows.append({
                    'month': _month_label_from_date(item_date),
                    'source': source_label, 'currency': cur, 'qty': qty,
                    'fx': amount, 'krw': round(amount * rate / div),
                })

    # 큐텐: 입력 건별 거래기간 종료일 기준
    if qoo10_result and qoo10_result.get('entries'):
        for e in qoo10_result.get('entries', []):
            amount = float(e.get('amount', 0) or 0)
            qty = int(e.get('qty', 0) or 0)
            rate = e.get('rate', jpy_rate)
            krw = e.get('krw', round(amount * rate))
            date_value = _qoo10_reporting_date(e, qoo10_result)
            rows.append({
                'month': _month_label_from_date(date_value),
                'source': '큐텐', 'currency': 'JPY', 'qty': qty,
                'fx': amount, 'krw': krw,
            })

    # 집계: 월 + 구분 + 통화
    grouped = {}
    for item in rows:
        key = (item['month'], item['source'], item['currency'])
        if key not in grouped:
            grouped[key] = {'qty': 0, 'fx': 0.0, 'krw': 0}
        grouped[key]['qty'] += item['qty']
        grouped[key]['fx'] += item['fx']
        grouped[key]['krw'] += item['krw']

    # 출력
    r = 4
    grand_qty = 0
    grand_fx_by_currency = {}
    grand_krw = 0
    months = sorted({k[0] for k in grouped.keys()})
    source_order = {'쇼피': 1, '라자다': 2, '이베이': 3, 'Joom': 4, '쇼피파이': 5, '큐텐': 6}

    for month in months:
        month_qty = 0
        month_krw = 0
        month_fx_by_currency = {}
        keys = sorted(
            [k for k in grouped.keys() if k[0] == month],
            key=lambda x: (source_order.get(x[1], 99), PREFERRED_CURRENCY_ORDER.index(x[2]) if x[2] in PREFERRED_CURRENCY_ORDER else 99, x[2])
        )
        for _month, source, cur in keys:
            data = grouped[(_month, source, cur)]
            vals = [month, source, cur, data['qty'], data['fx'], data['krw']]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=v)
                nf = {4: NUM, 5: NUM2, 6: NUM}.get(col)
                _style(c, font=FONT_DEFAULT, align=CENTER if col <= 4 else RIGHT, border=THIN_BORDER, num_format=nf)
            month_qty += data['qty']
            month_fx_by_currency[cur] = month_fx_by_currency.get(cur, 0.0) + data['fx']
            month_krw += data['krw']
            grand_qty += data['qty']
            grand_fx_by_currency[cur] = grand_fx_by_currency.get(cur, 0.0) + data['fx']
            grand_krw += data['krw']
            r += 1

        # 월 합계: 외화는 통화가 여러 개일 수 있으므로 원화 합계 중심으로 표시
        ws.cell(row=r, column=1, value=month)
        ws.cell(row=r, column=2, value='월 합계')
        ws.cell(row=r, column=4, value=month_qty)
        ws.cell(row=r, column=6, value=month_krw)
        for col in range(1, 7):
            c = ws.cell(row=r, column=col)
            nf = {4: NUM, 6: NUM}.get(col)
            _style(c, font=FONT_BOLD, fill=GRAY_FILL, align=CENTER if col <= 4 else RIGHT, border=THIN_BORDER, num_format=nf)
        r += 1

    # 전체 총합
    ws.cell(row=r, column=1, value='전체')
    ws.cell(row=r, column=2, value='전체 총합')
    ws.cell(row=r, column=4, value=grand_qty)
    ws.cell(row=r, column=6, value=grand_krw)
    for col in range(1, 7):
        c = ws.cell(row=r, column=col)
        nf = {4: NUM, 6: NUM}.get(col)
        _style(c, font=FONT_BOLD, fill=SUBHEAD_FILL, align=CENTER if col <= 4 else RIGHT, border=THIN_BORDER, num_format=nf)

    # 참고: 통화별 전체 외화 합계
    r += 2
    ws.cell(row=r, column=1, value='통화별 외화 합계')
    _style(ws.cell(row=r, column=1), font=FONT_BOLD)
    r += 1
    for cur in _ordered_currencies(grand_fx_by_currency.keys()):
        ws.cell(row=r, column=2, value=cur)
        ws.cell(row=r, column=5, value=grand_fx_by_currency.get(cur, 0.0))
        _style(ws.cell(row=r, column=2), font=FONT_DEFAULT, align=CENTER, border=THIN_BORDER)
        _style(ws.cell(row=r, column=5), font=FONT_DEFAULT, align=RIGHT, border=THIN_BORDER, num_format=NUM2)
        r += 1

# ── 전체 엑셀 생성 ───────────────────────────────────────────────

def generate_excel(
    shopee_results: list,      # [parse_shopee_pdf() 결과, ...]
    lazada_result:  Optional[dict],   # parse_lazada_pdf() 결과
    qoo10_result:   Optional[dict],   # parse_qoo10_pdf() 결과
    rates:          dict,      # fetch_all_currencies() 결과
    output_path:    str,
    year:           int,
    month:          int,
    ebay_results:   Optional[list] = None,
    joom_results:   Optional[list] = None,
    shopify_results: Optional[list] = None,
):
    """전체 엑셀 파일 생성"""
    wb = Workbook()
    wb.remove(wb.active)

    ebay_results = ebay_results or []
    joom_results = joom_results or []
    shopify_results = shopify_results or []
    usage = _infer_used_sources_and_currencies(shopee_results, lazada_result, qoo10_result, ebay_results,
                                               joom_results=joom_results, shopify_results=shopify_results)
    shopee_currencies = usage['shopee_currencies']
    lazada_currencies = usage['lazada_currencies']
    qoo10_used = usage['qoo10_used']
    ebay_currencies = usage.get('ebay_currencies', [])
    used_currencies = usage['used_currencies']

    # ── 라자다 기준일 추출 (period_end → write_date fallback: 귀속월 유지) ──
    if lazada_result:
        lazada_write_date = (lazada_result.get('period_end', '')
                             or lazada_result.get('write_date', ''))
    else:
        lazada_write_date = ''

    from .exchange_rate import avg_rate_for_period, monthly_avg_rate_for_month

    # ── 라자다 거래기간 평균환율 (통화별) ──
    lazada_avg_rates = {}
    if lazada_result:
        _lp_s = lazada_result.get('period_start', '')
        _lp_e = lazada_result.get('period_end', '')
        for _it in lazada_result.get('items', []):
            _lc = _it.get('currency', '')
            if _lc and _lc not in lazada_avg_rates:
                lazada_avg_rates[_lc] = avg_rate_for_period(rates.get(_lc), _lp_s, _lp_e)

    # ── 큐텐 JPY 환율: 반기말(6월/12월) 공식 월평균 매매기준율 사용 ──
    jpy_rate_data = rates.get('JPY')
    jpy_rate = 0.0
    if jpy_rate_data:
        # 표시·폴백용 대표값. 실제 계산은 아래에서 건별 반기말 월평균환율을 사용합니다.
        monthly_values = [
            _applied_rate_value('JPY', row.get('rate', 0))
            for row in (jpy_rate_data.get('monthly', []) or [])
            if float(row.get('rate', 0) or 0) > 0
        ]
        if monthly_values:
            jpy_rate = monthly_values[0]
        else:
            jpy_rate = _applied_rate_value('JPY', jpy_rate_data.get('monthly_average', 0))

    # 큐텐 신고/집계 기준일은 작성일이 아니라 거래기간 종료일입니다.
    qoo10_report_date = ''
    if qoo10_result:
        qoo10_report_date = _qoo10_reporting_date({}, qoo10_result)

        # ── 큐텐 건별 월평균환율·원화 계산 (entries) ──
        q_entries = qoo10_result.get('entries')
        if not q_entries:
            q_entries = [{
                'tracking_no': qoo10_result.get('tracking_no', ''),
                'qty':         qoo10_result.get('qty', 0),
                'amount':      qoo10_result.get('amount', 0),
                'write_date':  qoo10_result.get('write_date', ''),
                'period_start': qoo10_result.get('period_start', ''),
                'period_end':   qoo10_result.get('period_end', ''),
            }]
        q_total_krw = 0
        for e in q_entries:
            report_month = _qoo10_reporting_month(e, qoo10_result)
            r = _applied_rate_value('JPY', monthly_avg_rate_for_month(jpy_rate_data, report_month))
            e['rate'] = r
            e['rate_month'] = report_month
            e['rate_source'] = 'SMBS_MON_AVG_OFFICIAL'
            e['krw'] = round(float(e.get('amount', 0) or 0) * e['rate'])
            q_total_krw += e['krw']
        qoo10_result['entries'] = q_entries
        qoo10_result['amount'] = sum(float(e.get('amount', 0) or 0) for e in q_entries)
        qoo10_result['qty'] = sum(int(e.get('qty', 0) or 0) for e in q_entries)
        qoo10_result['total_krw'] = q_total_krw
        # 여러 반기 자료가 함께 있으면 총집계 표시는 외화금액 가중 실효환율로 표시합니다.
        if qoo10_result['amount']:
            jpy_rate = _applied_rate_value('JPY', q_total_krw / qoo10_result['amount'])

    # ── 제출자(판매자) 정보: PDF에서 자동 추출, 없으면 기본값 ──
    report_submitter = None
    for sd in shopee_results:
        if sd.get('submitter') and sd['submitter'].get('name'):
            report_submitter = sd['submitter']
            break
    if report_submitter is None and lazada_result and lazada_result.get('submitter', {}).get('name'):
        report_submitter = lazada_result['submitter']
    if report_submitter is None:
        for er in ebay_results:
            if er.get('submitter') and er['submitter'].get('name'):
                report_submitter = er['submitter']
                break
    if report_submitter is None:
        for jr in joom_results:
            if jr.get('submitter') and jr['submitter'].get('name'):
                report_submitter = jr['submitter']
                break
    if report_submitter is None and qoo10_result and qoo10_result.get('submitter', {}).get('name'):
        report_submitter = qoo10_result['submitter']
    if report_submitter is None:
        report_submitter = DEFAULT_SUBMITTER

    # ── 총집계 ──────────────────────────────────────────────
    ws_summary = wb.create_sheet('총집계')
    shopee_totals = {}
    lazada_totals = {}
    ebay_totals = {}
    joom_totals = {}
    shopify_totals = {}

    for sd in shopee_results:
        cur = sd.get('currency', '')
        if not cur:
            continue
        # 각 거래의 발행일(tx['date']) 기준 환율로 개별 계산 후 합산
        div = RATE_DIVISOR.get(cur, 1)
        total_fx  = 0.0
        total_krw = 0
        for tx in sd.get('transactions', []):
            tx_rate    = _get_rate(rates, cur, tx['date'])
            total_fx  += tx['amount']
            total_krw += round(tx['amount'] * tx_rate / div)
        # 거래 내역 없으면 total_amount 사용 (fallback)
        if not sd.get('transactions'):
            rate_date = sd.get('write_date', '') or sd.get('period_end', '')
            rate = _get_rate(rates, cur, rate_date)
            total_fx  = sd.get('total_amount', 0.0)
            total_krw = round(total_fx * rate / div)
        if cur not in shopee_totals:
            shopee_totals[cur] = {'fx': 0.0, 'krw': 0}
        shopee_totals[cur]['fx'] += total_fx
        shopee_totals[cur]['krw'] += total_krw

    if lazada_result:
        for it in lazada_result.get('items', []):
            cur = it.get('currency', '')
            if not cur:
                continue
            amount = float(it.get('amount', 0) or 0)
            rate = _lazada_item_rate(
                it, cur, rates, lazada_result, lazada_avg_rates.get(cur)
            )
            div = RATE_DIVISOR.get(cur, 1)
            krw = round(amount * rate / div)
            if cur not in lazada_totals:
                lazada_totals[cur] = {'fx': 0.0, 'krw': 0}
            lazada_totals[cur]['fx'] += amount
            lazada_totals[cur]['krw'] += krw

    # 이베이/린코스: 발행월 기준 월평균 매매기준율 사용
    for er in ebay_results:
        for it in er.get('items', []):
            cur = it.get('currency', '')
            if not cur:
                continue
            rate = _applied_rate_value(cur, monthly_avg_rate_for_month(rates.get(cur), it.get('month') or it.get('date') or ''))
            div = RATE_DIVISOR.get(cur, 1)
            amount = float(it.get('amount', 0) or 0)
            krw = round(amount * rate / div)
            if cur not in ebay_totals:
                ebay_totals[cur] = {'fx': 0.0, 'krw': 0}
            ebay_totals[cur]['fx'] += amount
            ebay_totals[cur]['krw'] += krw

    # Joom: 발송날짜 일별 매매기준율
    for jr in joom_results:
        for it in jr.get('items', []):
            cur = it.get('currency', '')
            if not cur:
                continue
            rate = _get_rate(rates, cur, it.get('date', ''))
            amount = float(it.get('amount', 0) or 0)
            krw = round(amount * rate / RATE_DIVISOR.get(cur, 1))
            joom_totals.setdefault(cur, {'fx': 0.0, 'krw': 0})
            joom_totals[cur]['fx'] += amount
            joom_totals[cur]['krw'] += krw

    # 쇼피파이: Fulfilled at 일별 매매기준율 (스토어 × 통화로 표시)
    for sr in shopify_results:
        for it in sr.get('items', []):
            cur = it.get('currency', '')
            if not cur:
                continue
            rate = _get_rate(rates, cur, it.get('date', ''))
            amount = float(it.get('amount', 0) or 0)
            krw = round(amount * rate / RATE_DIVISOR.get(cur, 1))
            key = _shopify_total_key(sr, cur)
            shopify_totals.setdefault(key, {'fx': 0.0, 'krw': 0})
            shopify_totals[key]['fx'] += amount
            shopify_totals[key]['krw'] += krw

    period_label, _ = period_labels(shopee_results, lazada_result, qoo10_result, ebay_results=ebay_results,
                                    joom_results=joom_results, shopify_results=shopify_results,
                                    fallback=f'{year}년 {month:02d}월')

    write_summary_sheet(ws_summary, shopee_totals, lazada_totals,
                        qoo10_result, jpy_rate,
                        period_label, submitter=report_submitter,
                        ebay_totals=ebay_totals, joom_totals=joom_totals,
                        shopify_totals=shopify_totals)

    # ── 월별집계 ─────────────────────────────────────────────
    # 총집계 바로 오른쪽에 배치합니다.
    ws_monthly = wb.create_sheet('월별집계', 1)
    write_monthly_summary_sheet(ws_monthly, shopee_results, lazada_result, qoo10_result,
                                rates, lazada_avg_rates=lazada_avg_rates,
                                lazada_write_date=lazada_write_date,
                                jpy_rate=jpy_rate, submitter=report_submitter,
                                ebay_results=ebay_results,
                                joom_results=joom_results,
                                shopify_results=shopify_results)

    # ── 통화별 수출신고 템플릿 시트
    # 실제 쇼피/라자다/큐텐 데이터가 있는 통화만 생성합니다.
    for cur in used_currencies:
        if cur == 'JPY':
            continue
        ws = wb.create_sheet(cur)
        sd = _shopee_items_for_currency(shopee_results, cur)
        lazada_items = []
        if _has_lazada_data(lazada_result):
            lazada_items = [it for it in lazada_result.get('items', [])
                            if it.get('currency') == cur]
        ebay_items = []
        for er in ebay_results:
            ebay_items.extend([it for it in er.get('items', []) if it.get('currency') == cur])
        joom_items = []
        for jr in joom_results:
            joom_items.extend([it for it in jr.get('items', []) if it.get('currency') == cur])
        shopify_items = []
        for sr in shopify_results:
            shopify_items.extend([it for it in sr.get('items', []) if it.get('currency') == cur])
        write_currency_template_sheet(ws, cur, sd, lazada_items, rates,
                                      lazada_write_date=lazada_write_date,
                                      lazada_rate_override=lazada_avg_rates.get(cur),
                                      ebay_items=ebay_items,
                                      joom_items=joom_items,
                                      shopify_items=shopify_items)

    # ── JPY 수출신고 시트 (큐텐 데이터가 있을 때만 생성) ──
    if qoo10_used:
        ws_jpy = wb.create_sheet('JPY')
        # 반기 합산 JPY는 외화·원화가 수천만 단위까지 가므로 기본 너비(8.43)로는 ###이 됩니다.
        for _col, _w in {'A': 14, 'B': 14, 'C': 12, 'D': 9, 'E': 10, 'F': 14, 'G': 14}.items():
            ws_jpy.column_dimensions[_col].width = _w
        headers = ['수출신고번호', '기타영세율건수', '선(기)적일자', '통화코드', '환율', '외화금액', '원화금액']
        for col, h in enumerate(headers, 1):
            c = ws_jpy.cell(row=4, column=col, value=h)
            _style(c, font=FONT_BOLD, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
        ws_jpy.cell(row=1, column=5, value='큐텐')
        ws_jpy.cell(row=1, column=6, value=qoo10_result.get('amount', 0)).number_format = NUM_FMT
        ws_jpy.cell(row=1, column=7, value=qoo10_result.get('total_krw', 0)).number_format = NUM_FMT
        _jr = 5
        for e in qoo10_result.get('entries', []):
            report_date = _qoo10_reporting_date(e, qoo10_result) or qoo10_report_date
            date_str = ''
            if report_date:
                try:
                    date_str = int(re.sub(r'\D', '', str(report_date))[:8])
                except ValueError:
                    date_str = ''
            tracking = e.get('tracking_no') or qoo10_result.get('tracking_no', '')
            ws_jpy.cell(row=_jr, column=1, value='')
            ws_jpy.cell(row=_jr, column=2, value=1)
            ws_jpy.cell(row=_jr, column=3, value=date_str or None)
            ws_jpy.cell(row=_jr, column=4, value='JPY')
            ws_jpy.cell(row=_jr, column=5, value=_applied_rate_value('JPY', e.get('rate', jpy_rate))).number_format = _applied_rate_format('JPY')
            ws_jpy.cell(row=_jr, column=6, value=e.get('amount', 0)).number_format = NUM_FMT
            ws_jpy.cell(row=_jr, column=7, value=e.get('krw', 0)).number_format = NUM_FMT
            _jr += 1

        # ── 큐텐(소포수령증) ──
        ws_q10 = wb.create_sheet('큐텐(소포수령증)')
        write_qoo10_sheet(ws_q10, qoo10_result, jpy_rate, submitter=report_submitter)

    # ── 쇼피 원본 PDF별 시트
    # 같은 통화의 PDF가 여러 개면 쇼피(MYR)(1), 쇼피(MYR)(2)처럼 기간순으로 분리합니다.
    for cur in shopee_currencies:
        items = _shopee_items_for_currency(shopee_results, cur)
        base_sheet_name = _base_shopee_sheet_name(cur)
        for idx, sd in enumerate(items, 1):
            sheet_name = f'{base_sheet_name}({idx})' if len(items) > 1 else base_sheet_name
            ws = wb.create_sheet(sheet_name)
            write_shopee_sheet(ws, sd, rates, submitter=sd.get('submitter') or report_submitter)

    # ── 라자다 원본/주문내역 + 통화별 상세
    if _has_lazada_data(lazada_result):
        lazada_source_sheet = _lazada_source_sheet_name(lazada_result)
        ws_laz = wb.create_sheet(lazada_source_sheet)
        write_lazada_receipt_sheet(ws_laz, lazada_result, rates, submitter=report_submitter)

        for cur in lazada_currencies:
            ws = wb.create_sheet(f'라자다({cur})')
            items = [it for it in lazada_result.get('items', []) if it.get('currency') == cur]
            if items:
                write_lazada_currency_detail_sheet(ws, cur, items, rates, lazada_result)

    # ── 이베이/린코스 원본 PDF별 시트
    ebay_sheet_names = _ebay_sheet_names_for_results(ebay_results)
    for idx, er in enumerate([x for x in ebay_results if _has_ebay_data(x)], 0):
        ws_ebay = wb.create_sheet(ebay_sheet_names[idx])
        write_ebay_receipt_sheet(ws_ebay, er, rates, submitter=er.get('submitter') or report_submitter)

    # ── Joom 원본 PDF별 시트
    joom_sheet_names = _joom_sheet_names_for_results(joom_results)
    for idx, jr in enumerate([x for x in joom_results if _has_joom_data(x)]):
        ws_joom = wb.create_sheet(joom_sheet_names[idx])
        write_joom_sheet(ws_joom, jr, rates, submitter=jr.get('submitter') or report_submitter)

    # ── 쇼피파이 스토어별 주문내역 시트
    shopify_sheet_names = _shopify_sheet_names_for_results(shopify_results)
    for idx, sr in enumerate([x for x in shopify_results if _has_shopify_data(x)]):
        ws_shopify = wb.create_sheet(shopify_sheet_names[idx])
        write_shopify_sheet(ws_shopify, sr, rates)

    # ── 환율 시트
    # 실제 데이터가 있는 통화만 생성합니다.
    for cur in used_currencies:
        ws = wb.create_sheet(f'환율({cur})')
        write_exchange_rate_sheet(ws, rates.get(cur))

    # ── 최종 안전장치: 불필요한 시트 삭제
    keep_sheets = {'총집계', '월별집계'}
    keep_sheets.update(cur for cur in used_currencies if cur != 'JPY')
    if qoo10_used:
        keep_sheets.update({'JPY', '큐텐(소포수령증)'})
    keep_sheets.update(_shopee_sheet_names_for_results(shopee_results, shopee_currencies))
    if _has_lazada_data(lazada_result):
        keep_sheets.add(_lazada_source_sheet_name(lazada_result))
        keep_sheets.update(f'라자다({cur})' for cur in lazada_currencies)
    keep_sheets.update(_ebay_sheet_names_for_results(ebay_results))
    keep_sheets.update(_joom_sheet_names_for_results(joom_results))
    keep_sheets.update(_shopify_sheet_names_for_results(shopify_results))
    keep_sheets.update(f'환율({cur})' for cur in used_currencies)
    _prune_workbook_sheets(wb, keep_sheets)

    wb.save(output_path)
    print(f'  ✅ 엑셀 저장 완료: {output_path}')
    return output_path
