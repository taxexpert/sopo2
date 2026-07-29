# -*- coding: utf-8 -*-
"""라자다 주문내역 Excel 파서.

라자다 Seller Center 주문내역 xlsx에서 다음 필드를 사용합니다.
- deliveredDate: 매출/선적 기준일
- paidPrice: 외화 매출금액

한 주문에 상품이 여러 개이면 주문번호가 같아도 각 행의 paidPrice를 각각 매출로
반영합니다. 즉, 주문번호 기준으로 임의 중복 제거하지 않습니다.
"""

from __future__ import annotations

import re
import warnings
from pathlib import Path
from typing import Iterable, Optional

import pandas as pd
from openpyxl import load_workbook


COUNTRY_INFO = {
    "vietnam": ("VN", "VND"),
    "viet nam": ("VN", "VND"),
    "베트남": ("VN", "VND"),
    "singapore": ("SG", "SGD"),
    "싱가포르": ("SG", "SGD"),
    "thailand": ("TH", "THB"),
    "태국": ("TH", "THB"),
    "malaysia": ("MY", "MYR"),
    "말레이시아": ("MY", "MYR"),
    "philippines": ("PH", "PHP"),
    "philippine": ("PH", "PHP"),
    "필리핀": ("PH", "PHP"),
    "indonesia": ("ID", "IDR"),
    "인도네시아": ("ID", "IDR"),
    "taiwan": ("TW", "TWD"),
    "대만": ("TW", "TWD"),
}


def _norm_header(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _text(value) -> str:
    return str(value or "").strip()


def _number(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _date(value) -> str:
    if value in (None, ""):
        return ""
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return ""
    return parsed.strftime("%Y-%m-%d")


def _country_info(value: str, filename: str = "") -> tuple[str, str, str]:
    combined = f"{_text(value)} {_text(filename)}".lower()
    for token, (country_code, currency) in COUNTRY_INFO.items():
        if token.lower() in combined:
            country_name = _text(value) or token
            return country_code, currency, country_name
    return "", "", _text(value)


def _find_header_row(ws, max_scan_rows: int = 20):
    """deliveredDate와 paidPrice가 함께 있는 헤더행을 찾습니다."""
    for row_no in range(1, min(ws.max_row, max_scan_rows) + 1):
        headers = [_norm_header(ws.cell(row_no, col).value) for col in range(1, ws.max_column + 1)]
        if "delivereddate" in headers and "paidprice" in headers:
            return row_no, {name: idx + 1 for idx, name in enumerate(headers) if name}
    return None, {}


def is_lazada_order_excel(path: str | Path) -> bool:
    """파일이 라자다 주문내역 Excel인지 헤더만 확인합니다."""
    path = Path(path)
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = load_workbook(path, read_only=False, data_only=True)
        return any(_find_header_row(ws)[0] for ws in wb.worksheets)
    except Exception:
        return False


def parse_lazada_order_excel(path: str | Path) -> dict:
    """라자다 주문내역 xlsx를 기존 lazada_result 구조로 변환합니다."""
    path = Path(path)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(path, read_only=False, data_only=True)

    selected = None
    header_row = None
    header_map = None
    for ws in wb.worksheets:
        found_row, found_map = _find_header_row(ws)
        if found_row:
            selected = ws
            header_row = found_row
            header_map = found_map
            break
    if selected is None:
        raise ValueError("deliveredDate와 paidPrice 열이 있는 라자다 주문내역 시트를 찾지 못했습니다.")

    def value(row_no: int, key: str):
        col = header_map.get(_norm_header(key))
        return selected.cell(row_no, col).value if col else None

    items = []
    skipped_no_date = 0
    skipped_no_amount = 0
    # 환불·취소 공통 정책: 전액환불/반품/취소 상태는 매출 미반영.
    # 사유별 집계와 함께, 시트에 정상 건과 같은 형식의 행으로 표시할 상세도 보존합니다.
    skipped_by_reason = {}
    skipped_items = []

    for row_no in range(header_row + 1, selected.max_row + 1):
        delivered_date = _date(value(row_no, "deliveredDate"))
        amount = _number(value(row_no, "paidPrice"))
        status_key = _text(value(row_no, "status")).lower()
        if any(k in status_key for k in ("refund", "return", "cancel")):
            reason = status_key or "canceled"
            b = skipped_by_reason.setdefault(reason, {"count": 0, "fx": 0.0})
            b["count"] += 1
            b["fx"] = round(b["fx"] + float(amount or 0), 2)
            shipping_country = _text(value(row_no, "shippingCountry"))
            destination, currency, _name = _country_info(shipping_country, path.name)
            skipped_items.append({
                "date": delivered_date,
                "order_number": _text(value(row_no, "orderNumber")),
                "order_item_id": _text(value(row_no, "orderItemId")),
                "tracking_no": _text(value(row_no, "trackingCode")),
                "carrier": _text(value(row_no, "shippingProvider")),
                "currency": currency,
                "amount": float(amount or 0),
                "skip_reason": reason,
                "source_file": path.name,
            })
            continue
        if not delivered_date:
            skipped_no_date += 1
            continue
        if amount == 0:
            skipped_no_amount += 1
            continue

        shipping_country = _text(value(row_no, "shippingCountry"))
        destination, currency, country_name = _country_info(shipping_country, path.name)
        if not currency:
            raise ValueError(
                f"{path.name} {row_no}행의 국가/통화를 판단하지 못했습니다. "
                f"shippingCountry={shipping_country!r}"
            )

        tracking_no = _text(value(row_no, "trackingCode"))
        order_number = _text(value(row_no, "orderNumber"))
        order_item_id = _text(value(row_no, "orderItemId"))
        if not tracking_no:
            tracking_no = order_number or order_item_id

        carrier = _text(value(row_no, "shippingProvider")) or "라자다"
        items.append({
            "service": "라자다",
            "carrier": carrier,
            "origin": "KR",
            "destination": destination,
            "destination_name": country_name,
            "tracking_no": tracking_no,
            "qty": 1,
            "amount": amount,
            "currency": currency,
            "date": delivered_date,
            "delivered_date": delivered_date,
            "order_number": order_number,
            "order_item_id": order_item_id,
            "item_name": _text(value(row_no, "itemName")),
            "seller_sku": _text(value(row_no, "sellerSku")),
            "status": _text(value(row_no, "status")),
            "source_file": path.name,
            "source_row": row_no,
            "source_kind": "order_excel",
        })

    if not items:
        raise ValueError("유효한 deliveredDate와 paidPrice가 있는 라자다 주문행이 없습니다.")

    items.sort(key=lambda it: (it.get("date", ""), it.get("tracking_no", ""), it.get("order_item_id", "")))
    dates = [it["date"] for it in items if it.get("date")]
    carriers = sorted({it.get("carrier") for it in items if it.get("carrier")})
    currencies = sorted({it.get("currency") for it in items if it.get("currency")})

    return {
        "type": "lazada",
        "source_kind": "order_excel",
        "source_files": [path.name],
        # 원본 주문내역 시트를 결과 워크북의 라자다(주문내역) 시트에
        # 양식 그대로 이어 붙일 수 있도록 원본 파일 바이트와 대상 시트명을 보관합니다.
        "source_workbooks": [{
            "filename": path.name,
            "sheet_name": selected.title,
            "content": path.read_bytes(),
        }],
        "carrier": carriers[0] if len(carriers) == 1 else "라자다",
        "period_start": min(dates),
        "period_end": max(dates),
        "write_date": max(dates),
        "items": items,
        "currencies": currencies,
        "row_count": len(items),
        "total_qty": sum(int(it.get("qty", 0) or 0) for it in items),
        "total_amount_by_currency": {
            cur: sum(float(it.get("amount", 0) or 0) for it in items if it.get("currency") == cur)
            for cur in currencies
        },
        "skipped_no_date": skipped_no_date,
        "skipped_no_amount": skipped_no_amount,
        "skipped_by_reason": skipped_by_reason,
        "skipped_items": skipped_items,
    }


def merge_lazada_results(results: Iterable[Optional[dict]]) -> Optional[dict]:
    """PDF/Excel 등 여러 라자다 결과를 하나로 합칩니다."""
    valid = [r for r in results if r and r.get("items")]
    if not valid:
        return None

    items = []
    source_files = []
    source_workbooks = []
    source_kinds = set()
    submitter = {}
    # 파일별 미반영 사유 집계·상세를 유실 없이 합칩니다 (조용한 제외 금지).
    skipped_by_reason = {}
    skipped_items = []
    skipped_no_date = 0
    skipped_no_amount = 0
    for result in valid:
        skipped_items.extend(result.get("skipped_items", []))
        items.extend(result.get("items", []))
        source_files.extend(result.get("source_files", []))
        source_workbooks.extend(result.get("source_workbooks", []))
        if result.get("source_kind"):
            source_kinds.add(result.get("source_kind"))
        sub = result.get("submitter") or {}
        if not submitter.get("name") and sub.get("name"):
            submitter = sub
        skipped_no_date += int(result.get("skipped_no_date", 0) or 0)
        skipped_no_amount += int(result.get("skipped_no_amount", 0) or 0)
        for reason, data in (result.get("skipped_by_reason") or {}).items():
            b = skipped_by_reason.setdefault(reason, {"count": 0, "fx": 0.0})
            b["count"] += int(data.get("count", 0) or 0)
            b["fx"] = round(b["fx"] + float(data.get("fx", 0) or 0), 2)

    items.sort(key=lambda it: (
        it.get("date") or "9999-99-99",
        it.get("currency") or "",
        it.get("tracking_no") or "",
        it.get("order_item_id") or "",
    ))

    item_dates = [it.get("date") for it in items if it.get("date")]
    starts = [r.get("period_start") for r in valid if r.get("period_start")] + item_dates
    ends = [r.get("period_end") for r in valid if r.get("period_end")] + item_dates
    write_dates = [r.get("write_date") for r in valid if r.get("write_date")]
    carriers = sorted({it.get("carrier") for it in items if it.get("carrier")})

    return {
        "type": "lazada",
        "source_kind": next(iter(source_kinds)) if len(source_kinds) == 1 else "mixed",
        "source_files": list(dict.fromkeys(source_files)),
        "source_workbooks": source_workbooks,
        "carrier": carriers[0] if len(carriers) == 1 else "라자다",
        "period_start": min(starts) if starts else "",
        "period_end": max(ends) if ends else "",
        "write_date": max(write_dates or ends) if (write_dates or ends) else "",
        "submitter": submitter,
        "items": items,
        "row_count": len(items),
        "skipped_by_reason": skipped_by_reason,
        "skipped_items": skipped_items,
        "skipped_no_date": skipped_no_date,
        "skipped_no_amount": skipped_no_amount,
    }
