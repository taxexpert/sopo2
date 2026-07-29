# -*- coding: utf-8 -*-
"""쇼피파이(Shopify) 주문내역(orders export) 파서.

Shopify 관리자에서 내려받은 `<스토어명> orders <기간>.csv` (또는 .xlsx)를 읽습니다.
같은 폴더에 함께 오는 `transaction`, `payout` 파일은 참고자료이며 집계에 쓰지 않습니다.

매출 인식 규칙
- 금액: `Total` 열 (상품가 + 배송비 + 세금)
- 기준일: `Fulfilled at` 의 날짜 부분 → 이 날짜의 일별 매매기준율 적용
- 제외: 같은 주문의 2번째 이후 라인아이템 행(`Total` 이 비어 있음)

환불·취소 처리 규칙 (전 플랫폼 공통 정책)
- partially_refunded → 반영 (`Total` 전액)
- voided            → 미반영 (사유: voided)
- refunded (전액환불) → 미반영 (사유: refunded)
- 그 외 `Fulfilled at` 없음 → 미반영 (사유: unfulfilled)
미반영 건은 사유별 건수·금액으로 집계해 `skipped_by_reason` 에 담습니다.
"""

from __future__ import annotations

import csv
import io
import re
import warnings
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook


# orders export 판별에 쓰는 필수 열 (정규화 후 비교)
REQUIRED_HEADERS = {"name", "financialstatus", "fulfilledat", "currency", "total"}
ORDERS_HINT_HEADERS = {"lineitemquantity", "lineitemname", "lineitemprice", "subtotal"}

# 숫자로 변환할 열
NUMERIC_HEADERS = {
    "subtotal", "shipping", "taxes", "total", "discountamount",
    "lineitemquantity", "lineitemprice", "lineitemcompareatprice",
    "refundedamount", "outstandingbalance", "lineitemdiscount", "duties",
    "tax1value", "tax2value", "tax3value", "tax4value", "tax5value",
}

# 날짜/시간으로 변환할 열
DATETIME_HEADERS = {
    "paidat", "fulfilledat", "createdat", "cancelledat", "nextpaymentdueat",
}

DATETIME_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%Y/%m/%d %H:%M:%S",
    "%Y/%m/%d %H:%M",
    "%Y/%m/%d",
)


def _norm_header(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_number(value):
    text = _text(value).replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except (TypeError, ValueError):
        return None
    return int(number) if number.is_integer() and abs(number) < 1e15 else number


def _split_datetime(value):
    """(엑셀에 기록할 값, 'YYYY-MM-DD') 튜플을 반환합니다.

    '2026-01-27 22:15:00 +0900' 처럼 시간대가 붙은 값은 Excel 이 날짜로 인식하지
    못하므로 원문 문자열을 그대로 두고 날짜만 따로 추출합니다.
    """
    if isinstance(value, datetime):
        return value, value.strftime("%Y-%m-%d")

    text = _text(value)
    if not text:
        return None, ""

    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    date_str = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}" if m else ""

    has_offset = bool(re.search(r"[+-]\d{4}$|[A-Z]{3,4}$", text))
    if not has_offset:
        for fmt in DATETIME_FORMATS:
            try:
                return datetime.strptime(text, fmt), date_str
            except ValueError:
                continue
    return text, date_str


def _read_csv_rows(path: Path) -> list:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            text = raw.decode(encoding)
        except UnicodeDecodeError:
            continue
        return list(csv.reader(io.StringIO(text, newline="")))
    return list(csv.reader(io.StringIO(raw.decode("utf-8", errors="replace"), newline="")))


def _read_xlsx_rows(path: Path) -> list:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        if any(REQUIRED_HEADERS.issubset({_norm_header(c) for c in (row or [])}) for row in rows[:20]):
            return rows
    return [list(row) for row in wb.worksheets[0].iter_rows(values_only=True)] if wb.worksheets else []


def _load_rows(path: Path) -> list:
    if path.suffix.lower() == ".csv":
        return _read_csv_rows(path)
    return _read_xlsx_rows(path)


def _find_header_row(rows: Iterable[list]) -> int:
    for idx, row in enumerate(list(rows)[:20]):
        normalized = {_norm_header(c) for c in (row or [])}
        if REQUIRED_HEADERS.issubset(normalized) and normalized & ORDERS_HINT_HEADERS:
            return idx
    return -1


def is_shopify_orders_file(path: str | Path) -> bool:
    """파일이 쇼피파이 orders export 인지 헤더만 보고 판별합니다."""
    path = Path(path)
    if path.suffix.lower() not in {".csv", ".xlsx", ".xlsm"}:
        return False
    try:
        return _find_header_row(_load_rows(path)) >= 0
    except Exception:
        return False


def store_name_from_filename(filename: str) -> str:
    """'FRESORY orders 26년.csv' → 'FRESORY'"""
    stem = Path(str(filename)).stem.strip()
    m = re.split(r"\s*orders?\s*", stem, maxsplit=1, flags=re.I)
    name = m[0].strip(" _-")
    return name or stem


def parse_shopify_orders(path: str | Path, store: str = "") -> dict:
    """쇼피파이 orders export 를 매출집계용 결과 dict 로 변환합니다."""
    path = Path(path)
    rows = _load_rows(path)
    header_row = _find_header_row(rows)
    if header_row < 0:
        raise ValueError(
            f"{path.name} 은 쇼피파이 주문내역(orders) 형식이 아닙니다. "
            "Name / Financial Status / Fulfilled at / Currency / Total 열을 확인해 주세요."
        )

    headers = [_text(c) for c in rows[header_row]]
    normalized = [_norm_header(c) for c in headers]
    index = {name: i for i, name in enumerate(normalized) if name}

    i_name = index.get("name")
    i_status = index.get("financialstatus")
    i_fulfilled = index.get("fulfilledat")
    i_currency = index.get("currency")
    i_total = index.get("total")

    data_rows = []
    row_flags = []
    items = []
    skipped_blank = 0
    # 미반영 주문을 사유별로 집계합니다: {사유: {'count': n, 'fx': 외화합계}}
    skipped_by_reason = {}

    def _skip(reason, amount, currency):
        b = skipped_by_reason.setdefault(reason, {"count": 0, "fx": 0.0, "currency": currency})
        b["count"] += 1
        b["fx"] = round(b["fx"] + float(amount or 0), 2)

    for raw_row in rows[header_row + 1:]:
        raw_row = list(raw_row or [])
        if len(raw_row) < len(headers):
            raw_row += [None] * (len(headers) - len(raw_row))
        if not any(_text(c) for c in raw_row):
            continue

        values = []
        fulfilled_date = ""
        for col, cell in enumerate(raw_row):
            key = normalized[col] if col < len(normalized) else ""
            text = _text(cell)
            if not text:
                values.append(None)
                continue
            if key in NUMERIC_HEADERS and not isinstance(cell, datetime):
                number = _to_number(cell)
                values.append(number if number is not None else text)
            elif key in DATETIME_HEADERS:
                excel_value, date_str = _split_datetime(cell)
                values.append(excel_value)
                if key == "fulfilledat":
                    fulfilled_date = date_str
            else:
                values.append(cell if not isinstance(cell, str) else text)

        row_index = len(data_rows)
        data_rows.append(values)

        amount = _to_number(values[i_total]) if i_total is not None else None
        order_name = _text(values[i_name]) if i_name is not None else ""
        currency = (_text(values[i_currency]).upper() if i_currency is not None else "") or "USD"
        status = _text(values[i_status]) if i_status is not None else ""

        if not amount:
            # 같은 주문의 2번째 이후 라인아이템 행 — 금액이 비어 있어 매출이 아닙니다.
            skipped_blank += 1
            row_flags.append({"index": row_index, "date": "", "counted": False, "reason": "blank"})
            continue

        status_key = status.strip().lower()
        if status_key == "refunded":
            # 전액환불 — 배송 여부와 무관하게 매출 미반영.
            _skip("refunded", amount, currency)
            row_flags.append({"index": row_index, "date": fulfilled_date, "counted": False, "reason": "refunded"})
            continue
        if not fulfilled_date:
            # 배송이 이루어지지 않은 건 — voided(취소)와 그 외(unfulfilled)를 구분해 집계합니다.
            reason = "voided" if status_key == "voided" else "unfulfilled"
            _skip(reason, amount, currency)
            row_flags.append({"index": row_index, "date": "", "counted": False, "reason": reason})
            continue

        row_flags.append({"index": row_index, "date": fulfilled_date, "counted": True})
        items.append({
            "platform": "쇼피파이",
            "store": store or store_name_from_filename(path.name),
            "order_name": order_name,
            "date": fulfilled_date,
            "fulfilled_at": _text(raw_row[i_fulfilled]) if i_fulfilled is not None else "",
            "financial_status": status,
            "currency": currency,
            "qty": 1,
            "amount": float(amount),
            "row_index": row_index,
            "source_file": path.name,
            "rate_basis": "daily",
        })

    if not items:
        raise ValueError(
            f"{path.name} 에서 매출로 인식할 주문이 없습니다. "
            "Fulfilled at 과 Total 이 채워진 행이 있는지 확인해 주세요."
        )

    dates = sorted(it["date"] for it in items if it.get("date"))
    total_by_currency = {}
    for it in items:
        cur = it["currency"]
        total_by_currency[cur] = round(total_by_currency.get(cur, 0.0) + it["amount"], 2)

    return {
        "type": "shopify",
        "platform": "쇼피파이",
        "store": store or store_name_from_filename(path.name),
        "carrier": "",
        "submitter": {"name": "", "biz_no": "", "ceo": "", "address": ""},
        "source_file": path.name,
        "sheet_title": Path(path.name).stem,
        "headers": headers,
        "rows": data_rows,
        "row_flags": row_flags,
        "items": items,
        "period_start": dates[0] if dates else "",
        "period_end": dates[-1] if dates else "",
        "write_date": dates[-1] if dates else "",
        "currencies": sorted(total_by_currency),
        "total_by_currency": total_by_currency,
        "row_count": len(items),
        "skipped_by_reason": skipped_by_reason,
        "skipped_unfulfilled": sum(v["count"] for k, v in skipped_by_reason.items()
                                   if k in ("voided", "unfulfilled")),
        "skipped_blank": skipped_blank,
    }


def merge_shopify_results(results: Iterable[Optional[dict]]) -> list:
    """스토어별 결과를 스토어명 순으로 정렬해 반환합니다."""
    valid = [r for r in results if r and r.get("items")]
    return sorted(valid, key=lambda r: (str(r.get("store") or ""), str(r.get("source_file") or "")))
