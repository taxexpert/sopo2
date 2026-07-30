# -*- coding: utf-8 -*-
"""빠른 회귀 체크 — 루프/수정 후 1분 안에 도는 경량 판정기.

전체 검증(정답지 건별 대조)과 달리, samples/*/expected/summary.json 의 확정 숫자와
문서 자체 합계(2중 체크)만 대조합니다. 환율은 data/ 캐시만 사용하며 SMBS 에 접속하지
않습니다. 캐시가 비어 있으면 먼저 `cp samples/fixtures/*.csv data/` 를 실행하세요.

사용법:
    python tools/quick_check.py            # 전체 케이스
    python tools/quick_check.py shopee     # 특정 케이스만

검사 항목 (docs/CHECKLIST.md 의 자동화 부분):
  [A] 원화 합계 = expected_total_krw               (원화를 제대로 파악했는가)
  [B] 반영 + 사유별 미반영 + 라인아이템 = 전체 행   (input 전 건이 설명되는가)
  [C] 문서 자체 합계 vs 건별 합산                   (서로 다른 2가지 체크 방식)
  [D] 열 너비 오버플로(###) 검사                    (숫자 셀이 열 너비보다 넓지 않은가)
"""

import sys
import glob
import json
import contextlib
import io
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
import pandas as pd  # noqa: E402
from modules.pdf_parser import parse_pdf  # noqa: E402
from modules.lazada_order_parser import parse_lazada_order_excel  # noqa: E402
from modules.shopify_parser import parse_shopify_orders  # noqa: E402
from modules.exchange_rate import (  # noqa: E402
    load_rate_cache, load_monthly_rate_cache, get_rate_for_date,
    round_applied_rate, normalize_smbs_rate,
)

S = ROOT / 'samples'
FAILS = []


def report(case, name, ok, detail=''):
    mark = 'OK ' if ok else 'FAIL'
    print(f'  [{mark}] {name}' + (f' — {detail}' if detail else ''))
    if not ok:
        FAILS.append(f'{case}: {name} {detail}')


def quiet_parse(path):
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        return parse_pdf(str(path))


def load_expected(case):
    p = S / case / 'expected' / 'summary.json'
    return json.loads(p.read_text(encoding='utf-8')) if p.exists() else None


class CachedRates:
    """data/ 캐시만 쓰는 초경량 환율 조회 (SMBS 접속 없음)."""

    def __init__(self):
        self.daily = load_rate_cache()
        self.monthly = load_monthly_rate_cache()

    def on(self, currency, date_str):
        d = self.daily[self.daily['currency'] == currency]
        d = d[d['date'] <= pd.to_datetime(str(date_str).replace('.', '-'))]
        if d.empty:
            raise RuntimeError(f'{currency} {date_str} 일별 환율이 캐시에 없습니다. '
                               'cp samples/fixtures/*.csv data/ 를 먼저 실행하세요.')
        return round_applied_rate(currency, float(d.sort_values('date')['rate'].iloc[-1]))

    def month(self, currency, month_key):
        m = self.monthly[(self.monthly['currency'] == currency)
                         & (self.monthly['year_month'] == month_key)]
        if m.empty:
            raise RuntimeError(f'{currency} {month_key} 월평균이 캐시에 없습니다.')
        return round_applied_rate(currency, normalize_smbs_rate(currency, float(m['rate'].iloc[0])))


def shopee_case_totals(pdfs, rates):
    totals = {}
    for f in pdfs:
        r = quiet_parse(f)
        cur = r['currency']
        tx_fx = round(sum(t['amount'] for t in r['transactions']), 2)
        tx_qty = sum(t['qty'] for t in r['transactions'])
        # [C] 문서 자체 합계행 vs 건별 합산
        report(f.name, f'{f.name[:34]} 자체합계 대조', abs(tx_fx - r['total_amount']) < 0.01 and tx_qty == r['total_qty'],
               f'내역 {tx_fx:,.2f}/{tx_qty}건 vs 합계행 {r["total_amount"]:,.2f}/{r["total_qty"]}건')
        b = totals.setdefault(cur, {'fx': 0.0, 'krw': 0, 'count': 0, 'qty': 0})
        for t in r['transactions']:
            b['count'] += 1
            b['qty'] += t['qty']
            b['fx'] = round(b['fx'] + t['amount'], 2)
            b['krw'] += round(t['amount'] * rates.on(cur, t['date']))
    return totals


def compare_currency_block(case, ours, expected_block):
    ok_all = True
    for cur, e in (expected_block or {}).items():
        o = ours.get(cur)
        if e.get('krw') is None:
            ok = o is not None and abs(o['fx'] - e['fx']) < 0.01
            report(case, f'{cur} 외화', ok, f'앱 {o["fx"] if o else None} vs 기대 {e["fx"]}')
        else:
            ok = (o is not None and abs(o['fx'] - e['fx']) < 0.01 and o['krw'] == e['krw'])
            report(case, f'{cur} 외화·원화', ok,
                   f'앱 {o["fx"] if o else None}/{o["krw"] if o else None} vs 기대 {e["fx"]}/{e["krw"]}')
        ok_all = ok_all and ok
    return ok_all


def check_shopee(rates):
    exp = load_expected('shopee')
    print('== shopee (쏭샵)')
    totals = shopee_case_totals(sorted((S / 'shopee/input').glob('*.pdf')), rates)
    compare_currency_block('shopee', totals, exp['expected']['by_currency'])
    total = sum(v['krw'] for v in totals.values())
    report('shopee', '[A] 원화 총합', total == exp['expected_total_krw'],
           f'{total:,} vs {exp["expected_total_krw"]:,}')


def check_lazada(rates):
    exp = load_expected('lazada')
    print('== lazada (수니네: 쇼피+라자다+큐텐)')
    sh = shopee_case_totals(sorted((S / 'lazada/input').glob('수니네_*.pdf')), rates)
    compare_currency_block('lazada/쇼피', sh, exp['expected']['shopee']['by_currency'])

    lz = {}
    for f in sorted((S / 'lazada/input').glob('라자다_*.xlsx')):
        r = parse_lazada_order_excel(f)
        # [B] 반영 + 미반영 사유 집계 존재 확인
        for it in r['items']:
            cur = it['currency']
            b = lz.setdefault(cur, {'fx': 0.0, 'krw': 0, 'count': 0, 'qty': 0})
            b['count'] += 1
            b['qty'] += it['qty']
            b['fx'] = round(b['fx'] + it['amount'], 2)
            b['krw'] += round(it['amount'] * rates.on(cur, it['date']))
    compare_currency_block('lazada/라자다', lz, exp['expected']['lazada']['by_currency'])

    k = quiet_parse(S / 'lazada/input/큐텐(KSE)_해외소포수령증_26년 상반기.pdf')
    jpy = rates.month('JPY', '2026-06')
    kq = {'JPY': {'fx': k['amount'], 'krw': round(k['amount'] * jpy), 'count': 1, 'qty': k['qty']}}
    compare_currency_block('lazada/큐텐', kq, exp['expected']['qoo10']['by_currency'])

    total = (sum(v['krw'] for v in sh.values()) + sum(v['krw'] for v in lz.values())
             + kq['JPY']['krw'])
    report('lazada', '[A] 원화 총합', total == exp['expected_total_krw'],
           f'{total:,} vs {exp["expected_total_krw"]:,}')


def check_qoo10(rates):
    exp = load_expected('qoo10')
    print('== qoo10 (라라블리)')
    q = quiet_parse(next((S / 'qoo10/input').glob('*.pdf')))
    jpy = rates.month('JPY', '2026-06')
    ours = {'JPY': {'fx': q['amount'], 'krw': round(q['amount'] * jpy), 'count': 1, 'qty': q['qty']}}
    compare_currency_block('qoo10', ours, exp['expected']['by_currency'])
    e_qty = exp['expected']['by_currency']['JPY']['qty']
    report('qoo10', '[C] 건수 대조', q['qty'] == e_qty, f'{q["qty"]} vs {e_qty}')


def check_shopify(rates):
    exp = load_expected('shopify')
    print('== shopify (정지민)')
    totals = {}
    for f in sorted((S / 'shopify/input').glob('*orders*.csv')):
        r = parse_shopify_orders(f)
        # [B] 완전성: 반영 + 사유별 미반영 + 라인아이템 = 전체 행
        skipped = sum(v['count'] for v in r['skipped_by_reason'].values())
        explained = r['row_count'] + skipped + r['skipped_blank']
        report('shopify', f'[B] {r["store"]} 전건 설명', explained == len(r['rows']),
               f'반영 {r["row_count"]} + 미반영 {skipped} + 라인아이템 {r["skipped_blank"]} = {explained} / 전체 {len(r["rows"])}')
        for it in r['items']:
            cur = it['currency']
            b = totals.setdefault(cur, {'fx': 0.0, 'krw': 0, 'count': 0, 'qty': 0})
            b['count'] += 1
            b['qty'] += it['qty']
            b['fx'] = round(b['fx'] + it['amount'], 2)
            b['krw'] += round(it['amount'] * rates.on(cur, it['date']))
    compare_currency_block('shopify', totals, exp['expected']['by_currency'])
    total = sum(v['krw'] for v in totals.values())
    report('shopify', '[A] 원화 총합', total == exp['expected_total_krw'],
           f'{total:,} vs {exp["expected_total_krw"]:,}')


def _overflow_cells(workbook_path):
    """숫자 셀의 서식 표시폭이 열 너비(병합 폭 합산)를 넘는 셀을 찾습니다 (### 위험)."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    found = []
    for ws in wb.worksheets:
        merged_width = {}
        for rng in ws.merged_cells.ranges:
            width = sum(ws.column_dimensions[get_column_letter(c)].width or 8.43
                        for c in range(rng.min_col, rng.max_col + 1))
            merged_width[(rng.min_row, rng.min_col)] = width
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, (int, float)):
                    continue
                width = merged_width.get((cell.row, cell.column))
                if width is None:
                    width = ws.column_dimensions[cell.column_letter].width or 8.43
                nf = cell.number_format or 'General'
                decimals = 4 if '0.0000' in nf else (2 if '0.00' in nf else 0)
                if 'General' in nf:
                    text = f'{cell.value:,}' if isinstance(cell.value, int) else f'{cell.value:,.2f}'
                else:
                    text = f'{cell.value:,.{decimals}f}'
                if len(text) > width - 0.7:
                    found.append(f'{ws.title}!{cell.coordinate} 값 {text} / 너비 {width:.0f}')
    return found


def check_overflow(rates):
    """[D] 대표 케이스(수니네 혼합 + 큐텐 케이팝피버 수동값)로 워크북을 생성해 ### 스캔."""
    from modules.excel_writer import generate_excel
    from modules.lazada_order_parser import merge_lazada_results
    from modules.shopify_parser import merge_shopify_results

    print('== [D] 열 너비 오버플로(###) 검사')
    rate_map = {}
    for cur in ['MYR', 'PHP', 'SGD', 'THB', 'TWD', 'VND', 'BRL', 'USD']:
        daily = rates.daily[rates.daily['currency'] == cur]
        rate_map[cur] = {
            'currency': cur, 'currency_name': cur, 'period': '', 'average': 0.0,
            'min': 0.0, 'min_date': '', 'max': 0.0, 'max_date': '', 'range': 0.0,
            'cross_rate': 0.0, 'display_start': '', 'display_end': '',
            'daily': [{'date': d.strftime('%Y.%m.%d'), 'rate': float(r), 'change': 0, 'cross': 0}
                      for d, r in zip(daily['date'], daily['rate'])],
        }
    jpy_monthly = rates.monthly[rates.monthly['currency'] == 'JPY']
    rate_map['JPY'] = {
        'currency': 'JPY', 'currency_name': '일본 엔 (JPY)', 'period': '',
        'average': 0.0, 'daily': [],
        'monthly': [{'year_month': m, 'rate': float(r)}
                    for m, r in zip(jpy_monthly['year_month'], jpy_monthly['rate'])],
    }

    shopee = [quiet_parse(f) for f in sorted((S / 'lazada/input').glob('수니네_*.pdf'))]
    lazada = merge_lazada_results([parse_lazada_order_excel(f)
                                   for f in sorted((S / 'lazada/input').glob('라자다_*.xlsx'))])
    shopify = merge_shopify_results([parse_shopify_orders(f)
                                     for f in sorted((S / 'shopify/input').glob('*orders*.csv'))])
    joom = [quiet_parse(next((S / 'joom/input').glob('*.pdf')))]
    k = quiet_parse(S / 'lazada/input/큐텐(KSE)_해외소포수령증_26년 상반기.pdf')
    k['entries'] = [{'period_start': k['period_start'], 'period_end': k['period_end'],
                     'tracking_no': k['tracking_no'], 'qty': k['qty'],
                     'amount': k['amount'], 'write_date': k['write_date']}]

    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / 'overflow_check.xlsx'
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            generate_excel(shopee_results=shopee, lazada_result=lazada, qoo10_result=k,
                           rates=rate_map, output_path=str(out),
                           joom_results=joom, shopify_results=shopify,
                           year=2026, month=6)
        found = _overflow_cells(out)
    report('overflow', '[D] ### 위험 셀 0건', not found,
           f'{len(found)}건: ' + ' / '.join(found[:5]) if found else '전 시트 통과')


def check_joom_case(case):
    exp = load_expected(case)
    print(f'== {case}')
    j = quiet_parse(next((S / case / 'input').glob('*.pdf')))
    # [C] PDF 인쇄 합계 vs 건별 합산 (문서 자체 2중 체크)
    report(case, '[C] PDF 선언합계 대조', not j['total_mismatch'],
           f'선언 {j["declared_total"]} vs 건별 {j["total_by_currency"]}')
    compare_currency_block(case, {c: {'fx': v, 'krw': None, 'count': 0, 'qty': 0}
                                  for c, v in j['total_by_currency'].items()},
                           exp['expected']['by_currency'])


CHECKS = {
    'shopee': check_shopee, 'lazada': check_lazada, 'qoo10': check_qoo10,
    'shopify': check_shopify,
    'joom': lambda rates: check_joom_case('joom'),
    'joom2': lambda rates: check_joom_case('joom2'),
    'overflow': check_overflow,
}


def main():
    wanted = sys.argv[1:] or list(CHECKS)
    rates = CachedRates()
    for case in wanted:
        if case not in CHECKS:
            print(f'알 수 없는 케이스: {case} (가능: {", ".join(CHECKS)})')
            return 2
        CHECKS[case](rates)
    print()
    if FAILS:
        print(f'❌ 실패 {len(FAILS)}건:')
        for f in FAILS:
            print('  -', f)
        return 1
    print('✅ 전체 통과')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
