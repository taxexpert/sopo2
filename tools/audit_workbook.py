# -*- coding: utf-8 -*-
"""산출물 자기완결 감사 — 생성된 산출물만 열어서 판정하는 검증기.

`quick_check.py` 와 관측 지점이 다릅니다. 입력 파싱도 정답지도 믿지 않고,
매출집계 워크북과 신고서류 2종의 **셀 값만** 읽어서 서로 대사합니다.
`samples/*/expected/` 는 읽지도 쓰지도 않습니다.

사용법:
    python tools/audit_workbook.py --build qoo10        # 케이스 조립 → 3종 생성 → 감사
    python tools/audit_workbook.py --build all
    python tools/audit_workbook.py <매출집계.xlsx> [--zero a.xlsx b.xlsx] [--export c.xlsx]

검사 항목 (docs/VERIFICATION.md):
  A-4  통화 시트 1~5행 플랫폼 요약 vs 데이터행 합
  A-5  총집계 = 월별집계 = 통화별시트 = 수출실적 = 영세율 = 독립재계산 (6중, 허용오차 0원)
  B-2  환율(통화) 시트 표시값(100통화) 경유 원화 검산
  C-3  추적번호 중복 (경고 등급 — 한 소포 다수 라인아이템이 정상일 수 있음)
  D-1  시트 구성 · 빈 시트 · keep_sheets 잔여 시트
  D-2  헤더 위치 — 통화 시트 6행, JPY(큐텐) 시트만 4행
  D-3  신고서류 필수 필드 + 서식 템플릿 반영 여부
  D-4  쇼피파이 시트가 원본 CSV 전 건을 설명하는가
  E-2  외화는 있는데 환율/원화가 0인 행 (조용한 오답)

종료코드: 0 통과 / 1 불일치 / 2 감사기 자체 오류(구조 붕괴)
"""

import argparse
import contextlib
import csv
import io
import re
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

# 리포트에 —·✅ 가 들어가므로 cp949 콘솔에서도 깨지지 않게 UTF-8로 고정합니다.
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl  # noqa: E402

from modules.pdf_parser import parse_pdf  # noqa: E402
from modules.lazada_order_parser import (  # noqa: E402
    is_lazada_order_excel, parse_lazada_order_excel, merge_lazada_results,
)
from modules.shopify_parser import (  # noqa: E402
    is_shopify_orders_file, parse_shopify_orders, merge_shopify_results,
)
from modules.exchange_rate import load_rate_cache, load_monthly_rate_cache  # noqa: E402

S = ROOT / 'samples'
FORMS = ROOT / 'forms'

CURRENCY_SHEET_RE = re.compile(r'^[A-Z]{3}$')
DECL_HEADERS = ['수출신고번호', '기타영세율건수', '선(기)적일자', '통화코드', '환율', '외화금액', '원화금액']
SHOPIFY_HEADERS = ['No', '주문번호', '구분', '배송완료일', '결제상태', '통화', '외화금액', '환율', '원화금액', '상품명']
HUNDRED_UNIT = {'JPY', 'IDR', 'VND'}

CASES = ['shopee', 'lazada', 'qoo10', 'shopify', 'joom', 'joom2']


# ── 리포트 ──────────────────────────────────────────────────────

class Report:
    """sales-summary-verifier 리포트 포맷을 그대로 쓰는 결과 축적기."""

    def __init__(self, scope=''):
        self.scope = scope
        self.lines = []
        self.n_mismatch = 0
        self.n_warn = 0
        self.n_pass = 0
        self.struct = []

    def mismatch(self, where, expected, actual, diff, basis, cause):
        self.n_mismatch += 1
        self.lines.append(f'  [불일치] {where} — 기대 {expected} / 실제 {actual} / 차이 {diff}\n'
                          f'      근거: {basis}\n'
                          f'      추정원인: {cause}')

    def warn(self, where, expected, actual, diff, basis, cause):
        self.n_warn += 1
        self.lines.append(f'  [경고] {where} — 기대 {expected} / 실제 {actual} / 차이 {diff}\n'
                          f'      근거: {basis}\n'
                          f'      추정원인: {cause}')

    def ok(self, label, detail=''):
        self.n_pass += 1
        self.lines.append(f'  [일치] {label}' + (f' — {detail}' if detail else ''))

    def skip(self, label, reason):
        self.lines.append(f'  [건너뜀] {label} — {reason}')

    def broken(self, message):
        """감사기가 판정을 내릴 수 없는 구조 붕괴 — exit 2."""
        self.struct.append(message)
        self.lines.append(f'  [구조붕괴] {message}')

    def render(self):
        head = f'== {self.scope}' if self.scope else '=='
        return '\n'.join([head] + self.lines)


def num(v):
    """숫자 셀만 float로. 문자열·None은 None."""
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def fmt(v):
    if v is None:
        return 'None'
    if isinstance(v, float):
        return f'{v:,.4f}'.rstrip('0').rstrip('.')
    if isinstance(v, int):
        return f'{v:,}'
    return str(v)


# ── 워크북 열기 가드 ─────────────────────────────────────────────

def open_book(path, rep):
    """data_only=False 로 열고 수식 도입을 감지합니다.

    총합 셀은 현재 파이썬 리터럴이므로 값 비교가 그대로 성립합니다. 누군가
    Excel 수식으로 바꾸면 data_only=True 는 None 을 돌려줘 감사기가 조용히
    전부 통과하게 되므로, 여기서 명시적으로 막습니다.
    """
    wb = openpyxl.load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    rep.broken(f'{Path(path).name}!{ws.title}!{cell.coordinate} 수식 발견 '
                               f'({cell.value[:30]}) — 값 비교 감사가 무력화됩니다')
                    return wb
    return wb


# ── 총집계 ──────────────────────────────────────────────────────

PLATFORM_HEADS = {'쇼피', '라자다', '이베이', 'Joom', '쇼피파이', '큐텐'}


def read_summary(wb, rep):
    """총집계 시트에서 원화 총합과 플랫폼별 소계를 읽습니다.

    큐텐 구역만 `_totalrow`('총합' 행)를 쓰지 않고 데이터 행만 있으므로
    (excel_writer.py:1536-1550) 특수처리해야 합니다. 이걸 빼면 큐텐 금액이
    통째로 누락돼 감사기가 늘 위양성 FAIL 을 냅니다.
    """
    if '총집계' not in wb.sheetnames:
        rep.broken('총집계 시트가 없습니다')
        return None
    ws = wb['총집계']
    subtotals = {}
    total = 0
    current = None
    qoo10_head = None
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 2).value
        if label in PLATFORM_HEADS:
            current = label
            if label == '큐텐':
                qoo10_head = r
            continue
        if label == '총합':
            krw = num(ws.cell(r, 4).value)
            if krw is None:
                rep.broken(f'총집계!D{r} 총합 행의 원화가 숫자가 아닙니다: {ws.cell(r, 4).value!r}')
                return None
            subtotals[current or f'행{r}'] = int(krw)
            total += int(krw)
    if qoo10_head is not None:
        krw = num(ws.cell(qoo10_head + 2, 4).value)
        if krw is None:
            rep.broken(f'총집계 큐텐 구역(D{qoo10_head + 2})의 원화가 숫자가 아닙니다 — '
                       '큐텐은 총합 행이 없어 데이터 행을 직접 읽습니다')
            return None
        subtotals['큐텐'] = int(krw)
        total += int(krw)
    return {'total_krw': total, 'subtotals': subtotals}


# ── 월별집계 ────────────────────────────────────────────────────

def read_monthly(wb, rep):
    """월별집계 시트를 읽고 시트 내부 3중 자체 대조까지 수행합니다."""
    if '월별집계' not in wb.sheetnames:
        rep.broken('월별집계 시트가 없습니다')
        return None
    ws = wb['월별집계']
    headers = [ws.cell(3, c).value for c in range(1, 7)]
    expect = ['월', '구분', '통화코드', '건수', '외화금액', '원화금액']
    if headers != expect:
        rep.mismatch('월별집계!A3:F3', expect, headers, '헤더 불일치',
                     'VERIFICATION D-1 (시트 구성)', 'excel_writer.py:1804 headers')
        return None

    data_krw = 0
    data_qty = 0
    month_krw = 0
    month_qty = 0
    grand_krw = None
    grand_qty = None
    fx_by_cur = {}
    for r in range(4, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if b == '전체 총합':
            grand_krw = num(ws.cell(r, 6).value)
            grand_qty = num(ws.cell(r, 4).value)
            continue
        if b == '월 합계':
            month_krw += int(num(ws.cell(r, 6).value) or 0)
            month_qty += int(num(ws.cell(r, 4).value) or 0)
            continue
        cur = ws.cell(r, 3).value
        krw = num(ws.cell(r, 6).value)
        if a and isinstance(cur, str) and cur and krw is not None:
            data_krw += int(krw)
            data_qty += int(num(ws.cell(r, 4).value) or 0)

    if grand_krw is None:
        rep.broken("월별집계에 '전체 총합' 행이 없습니다")
        return None
    grand_krw = int(grand_krw)

    # 시트 자체 2중 체크 3개 — 완전 일치
    for label, got in (('월 합계 합산', month_krw), ('데이터행 합산', data_krw)):
        if got != grand_krw:
            rep.mismatch(f'월별집계 {label}', grand_krw, got, got - grand_krw,
                         'VERIFICATION A-5 (월별집계 내부 정합)',
                         'excel_writer.py:1930-1972 write_monthly_summary_sheet 누산')
        else:
            rep.ok(f'월별집계 {label} = 전체 총합', f'{grand_krw:,}원')
    if grand_qty is not None and int(grand_qty) != month_qty:
        rep.mismatch('월별집계 건수', int(grand_qty), month_qty, month_qty - int(grand_qty),
                     'VERIFICATION A-5 (건수 정합)', 'excel_writer.py:1956/1967')

    return {'grand_krw': grand_krw, 'grand_qty': int(grand_qty or 0),
            'data_qty': data_qty, 'fx_by_currency': fx_by_cur}


# ── 통화별 수출신고 시트 ─────────────────────────────────────────

def read_currency_sheets(wb, rep):
    """통화 시트(A~G)를 읽고 D-2·D-3·A-4·B-2(행내부)·E-2 를 판정합니다."""
    out = {}
    for ws in wb.worksheets:
        if not CURRENCY_SHEET_RE.match(ws.title):
            continue
        cur = ws.title
        header_row = 4 if cur == 'JPY' else 6

        # D-2 헤더 위치 — 어긋나면 수출신고 프로그램 업로드가 실패한다
        headers = [ws.cell(header_row, c).value for c in range(1, 8)]
        if headers != DECL_HEADERS:
            rep.mismatch(f'{cur}!A{header_row}:G{header_row}', DECL_HEADERS, headers,
                         '헤더 행 위치/문자열 불일치',
                         f'VERIFICATION D-2 (통화 시트 {header_row}행 헤더)',
                         'excel_writer.py:670-675 (비JPY) / 2248-2251 (JPY)')
            continue
        if [ws.cell(header_row + 1, c).value for c in range(1, 8)] == DECL_HEADERS:
            rep.mismatch(f'{cur}!{header_row + 1}행', '데이터', '헤더 중복', '헤더가 두 번',
                         'VERIFICATION D-2', 'excel_writer 헤더 작성부')

        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            d = ws.cell(r, 4).value
            fx = num(ws.cell(r, 6).value)
            if d != cur or fx is None:
                continue
            rows.append({
                'row': r,
                'export_no': ws.cell(r, 1).value,
                'other_count': ws.cell(r, 2).value,
                'ship_date': ws.cell(r, 3).value,
                'rate': num(ws.cell(r, 5).value),
                'fx': fx,
                'krw': num(ws.cell(r, 7).value),
            })
        if not rows:
            rep.mismatch(f'{cur} 시트', '데이터 행 1건 이상', '0건', '빈 시트',
                         'VERIFICATION D-1 (빈 시트 금지)', 'excel_writer.py:2217-2240')
            continue

        # D-3 필수 필드
        bad_count = [x['row'] for x in rows if x['other_count'] != 1]
        bad_export = [x['row'] for x in rows if x['export_no'] not in (None, '')]
        bad_date = [x['row'] for x in rows
                    if not (isinstance(x['ship_date'], int) and 19000101 <= x['ship_date'] <= 21001231)]
        for label, bad, basis in (
            ('기타영세율건수=1', bad_count, 'VERIFICATION D-3'),
            ('수출신고번호 공란', bad_export, 'VERIFICATION D-3'),
            ('선(기)적일자 8자리', bad_date, 'VERIFICATION D-3'),
        ):
            if bad:
                rep.mismatch(f'{cur} {label}', '전건 충족', f'{len(bad)}건 위반(행 {bad[:5]})',
                             len(bad), basis, 'extra_docs/excel_writer 신고행 작성부')

        # E-2 조용한 오답 — 외화는 있는데 환율/원화가 0
        silent = [x['row'] for x in rows if x['fx'] > 0 and (not x['rate'] or not x['krw'])]
        if silent:
            rep.mismatch(f'{cur} 환율/원화 0원 행', '0건', f'{len(silent)}건(행 {silent[:5]})',
                         len(silent), 'VERIFICATION E-2 (조용한 오답 금지)',
                         'excel_writer.py:234-235 _get_rate 가 통화 미보유 시 0.0 반환')

        # B-2 행 내부 검산 — 시트 E열은 이미 1통화 환율이므로 나눗셈이 없어야 한다
        off = [(x['row'], x['krw'], round(x['fx'] * (x['rate'] or 0)))
               for x in rows if abs((x['krw'] or 0) - round(x['fx'] * (x['rate'] or 0))) > 1]
        if off:
            r0, got, exp = off[0]
            rep.mismatch(f'{cur}!G{r0} (외 {len(off) - 1}건)', exp, got, (got or 0) - exp,
                         'VERIFICATION B-2 (원화 = 외화 × 1통화환율)',
                         'excel_writer 통화 시트 krw 계산 / 100통화 이중 환산')
        else:
            rep.ok(f'{cur} 행 내부 원화 = 외화 × 환율', f'{len(rows)}행')

        sum_krw = sum(int(x['krw'] or 0) for x in rows)
        sum_fx = sum(x['fx'] for x in rows)

        # A-4 자체 요약(1~5행 또는 JPY 1행) vs 데이터행 합
        if cur == 'JPY':
            head_fx = num(ws.cell(1, 6).value)
            head_krw = num(ws.cell(1, 7).value)
            labels = [(ws.cell(1, 5).value, head_fx, head_krw)]
        else:
            labels = [(ws.cell(r, 5).value, num(ws.cell(r, 6).value), num(ws.cell(r, 7).value))
                      for r in range(1, 6)]
        head_krw_sum = sum(int(k or 0) for _l, _f, k in labels)
        head_fx_sum = sum(f or 0.0 for _l, f, _k in labels)
        if head_krw_sum != sum_krw:
            rep.mismatch(f'{cur} 시트 요약 원화 합', sum_krw, head_krw_sum,
                         head_krw_sum - sum_krw, 'VERIFICATION A-4 (자체 요약 대조)',
                         'excel_writer.py:651-664 플랫폼별 소계 누산')
        else:
            rep.ok(f'{cur} 요약 원화 합 = 데이터행 합', f'{sum_krw:,}원')
        if abs(head_fx_sum - sum_fx) > 0.01:
            rep.mismatch(f'{cur} 시트 요약 외화 합', round(sum_fx, 2), round(head_fx_sum, 2),
                         round(head_fx_sum - sum_fx, 4), 'VERIFICATION A-4',
                         'excel_writer.py:651-664')

        out[cur] = {'rows': rows, 'sum_krw': sum_krw, 'sum_fx': sum_fx, 'header_row': header_row}
    return out


# ── 환율 시트 경유 B-2 검산 ──────────────────────────────────────

def read_rate_sheet(wb, cur):
    """환율(XXX) 시트에서 SMBS 표시단위(100통화) 값을 읽습니다."""
    name = f'환율({cur})'
    if name not in wb.sheetnames:
        return None
    ws = wb[name]
    a1 = str(ws.cell(1, 1).value or '')
    if a1.startswith('월평균'):
        table = {}
        for r in range(5, ws.max_row + 1):
            ym = ws.cell(r, 1).value
            rate = num(ws.cell(r, 3).value)
            if ym and rate:
                table[str(ym).replace('.', '-')[:7]] = rate
        return {'kind': 'monthly', 'table': table}
    if a1.startswith('기간별'):
        table = {}
        for r in range(10, ws.max_row + 1):
            d = ws.cell(r, 1).value
            rate = num(ws.cell(r, 3).value)
            if d and rate:
                table[re.sub(r'\D', '', str(d))[:8]] = rate
        return {'kind': 'daily', 'table': table}
    return None


def _sample_rows(rows, k=5):
    """첫 k / 중간 k / 마지막 k / 최대금액 k 를 표본으로 뽑습니다."""
    if len(rows) <= 4 * k:
        return rows
    mid = len(rows) // 2
    picked = rows[:k] + rows[mid:mid + k] + rows[-k:] + sorted(rows, key=lambda x: -x['fx'])[:k]
    seen, out = set(), []
    for x in picked:
        if x['row'] not in seen:
            seen.add(x['row'])
            out.append(x)
    return out


def check_b2_via_rate_sheet(wb, sheets, rep):
    """환율 시트의 100통화 표시값을 경유해 원화를 다시 검산합니다 (B-2 정본)."""
    for cur, info in sorted(sheets.items()):
        rs = read_rate_sheet(wb, cur)
        if rs is None:
            rep.skip(f'{cur} B-2 환율시트 검산', f'환율({cur}) 시트를 읽을 수 없음')
            continue
        table = rs['table']
        if not table:
            rep.skip(f'{cur} B-2 환율시트 검산', f'환율({cur}) 시트에 값이 없음')
            continue
        div = 100.0 if cur in HUNDRED_UNIT else 1.0
        keys = sorted(table)
        bad = []
        fallback = 0
        for x in _sample_rows(info['rows']):
            d = str(x['ship_date'] or '')
            if rs['kind'] == 'monthly':
                key = f'{d[:4]}-{d[4:6]}'
                disp = table.get(key)
            else:
                disp = table.get(d)
                if disp is None:
                    prev = [k for k in keys if k <= d]
                    if prev:
                        disp = table[prev[-1]]
                        fallback += 1
            if disp is None:
                continue
            expect = round(x['fx'] * disp / div)
            if abs((x['krw'] or 0) - expect) > 1:
                bad.append((x['row'], x['krw'], expect, disp))
        if bad:
            r0, got, exp, disp = bad[0]
            rep.mismatch(f'{cur}!G{r0} (표본 위반 {len(bad)}건)', exp, got, (got or 0) - exp,
                         f'VERIFICATION B-2 (환율({cur}) 표시값 {fmt(disp)} ÷ {int(div)})',
                         'exchange_rate.normalize_smbs_rate / round_applied_rate 100통화 처리')
        else:
            rep.ok(f'{cur} B-2 환율시트 경유 검산', f'표본 {len(_sample_rows(info["rows"]))}행'
                   + (f', 직전영업일 폴백 {fallback}건' if fallback else ''))
        if fallback and fallback > len(_sample_rows(info['rows'])) * 0.3:
            rep.warn(f'{cur} 환율 직전영업일 폴백 비율', '표본의 30% 이하',
                     f'{fallback}건', fallback, 'VERIFICATION B-4 (휴일 ffill)',
                     '환율 시트 표시 범위(display_start/end)가 거래일을 못 덮는지 확인')


# ── 시트 구성 (D-1) ─────────────────────────────────────────────

def check_d1(wb, rep):
    names = wb.sheetnames
    for must in ('총집계', '월별집계'):
        if must not in names:
            rep.broken(f'{must} 시트 누락')
    curs = [n for n in names if CURRENCY_SHEET_RE.match(n)]
    rate_sheets = [n for n in names if n.startswith('환율(')]
    for cur in curs:
        if f'환율({cur})' not in names:
            rep.mismatch(f'환율({cur}) 시트', '존재', '없음', 1,
                         'VERIFICATION D-1 (사용 통화별 환율 시트)',
                         'excel_writer.py:2320-2322 / keep_sheets')
    if 'JPY' in curs and '큐텐(소포수령증)' not in names:
        rep.mismatch('큐텐(소포수령증) 시트', '존재', '없음', 1,
                     'VERIFICATION D-1', 'excel_writer.py:2275-2276')
    empty = [ws.title for ws in wb.worksheets if ws.max_row <= 1 and ws.max_column <= 1]
    if empty:
        rep.mismatch('빈 시트', '0개', f'{empty}', len(empty),
                     'VERIFICATION D-1 (빈 시트 금지)', 'excel_writer 시트 생성부')
    else:
        rep.ok('D-1 시트 구성', f'{len(names)}개 시트 (통화 {len(curs)} / 환율 {len(rate_sheets)})')
    return curs


# ── 신고서류 (D-3) ──────────────────────────────────────────────

def _template_struct(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    return {
        'sheetnames': list(wb.sheetnames),
        'max_column': ws.max_column,
        'merges': {str(r) for r in ws.merged_cells.ranges},
    }


def read_export_performance(path, rep):
    """수출실적명세서 — 서식 템플릿 반영 여부 + 필수 필드 + 합계."""
    wb = open_book(path, rep)
    ws = wb.active
    tpl = FORMS / '수출실적명세서 양식.xlsx'
    if tpl.exists():
        t = _template_struct(tpl)
        if list(wb.sheetnames) != t['sheetnames']:
            rep.mismatch(f'{path.name} 시트 구성', t['sheetnames'], list(wb.sheetnames),
                         '서식 템플릿 미반영',
                         'VERIFICATION D-3 (서식 템플릿 반영)',
                         'extra_docs.py:337 _find_template / 폴백 _new 경로')
        else:
            rep.ok('수출실적명세서 서식 템플릿 반영', f'시트 {t["sheetnames"]}')

    headers = [ws.cell(1, c).value for c in range(1, 8)]
    if headers != DECL_HEADERS:
        rep.mismatch(f'{path.name}!A1:G1', DECL_HEADERS, headers, '헤더 불일치',
                     'VERIFICATION D-3', 'extra_docs.py:348')
    rows = []
    for r in range(2, ws.max_row + 1):
        fx = num(ws.cell(r, 6).value)
        if fx is None:
            continue
        rows.append({'row': r, 'other_count': ws.cell(r, 2).value,
                     'ship_date': ws.cell(r, 3).value, 'cur': ws.cell(r, 4).value,
                     'rate': num(ws.cell(r, 5).value), 'fx': fx,
                     'krw': num(ws.cell(r, 7).value)})
    if not rows:
        rep.broken(f'{path.name} 데이터 행 0건')
        return None
    off = [x['row'] for x in rows if abs((x['krw'] or 0) - round(x['fx'] * (x['rate'] or 0))) > 1]
    if off:
        rep.mismatch(f'{path.name} 원화 = 외화 × 환율', '전건 충족',
                     f'{len(off)}건 위반(행 {off[:5]})', len(off),
                     'VERIFICATION D-3', 'extra_docs.py:355-364')
    # _clear_rows 잔존 오염
    last = rows[-1]['row']
    leftover = [r for r in range(last + 1, ws.max_row + 1)
                if any(ws.cell(r, c).value not in (None, '') for c in range(1, 8))]
    if leftover:
        rep.mismatch(f'{path.name} 데이터 이후 잔존 값', '0행', f'{len(leftover)}행({leftover[:3]})',
                     len(leftover), 'VERIFICATION D-3',
                     'extra_docs.py:319 _clear_rows 범위 누락')
    total = sum(int(x['krw'] or 0) for x in rows)
    rep.ok('수출실적명세서 읽음', f'{len(rows)}행 / 원화 {total:,}')
    return {'rows': rows, 'sum_krw': total}


ZERO_COLS = {'구분': 1, '서류명': 2, '발급자': 3, '발급일자': 4, '선적일자': 5,
             'LC': 6, '통화': 8, '환율': 9, '제출외화': 10, '제출원화': 11,
             '당기외화': 12, '당기원화': 13, '미도래외화': 14, '미도래원화': 15}


def read_zero_rate(paths, rep):
    """영세율첨부서류제출명세서 — 템플릿 반영 여부 + 열 매핑 + 월별 합산 검사."""
    tpl = FORMS / '영세율첨부서류제출명세서 양식.xlsx'
    t = _template_struct(tpl) if tpl.exists() else None

    total_file = None
    monthly_sum = 0
    monthly_rows = 0
    for path in paths:
        wb = open_book(path, rep)
        ws = wb.active
        is_total = path.stem.endswith('_전체')

        if t is not None and is_total:
            missing = t['merges'] - {str(r) for r in ws.merged_cells.ranges}
            if ws.max_column < t['max_column'] or missing:
                rep.mismatch(f'{path.name} 헤더 구조',
                             f'최대열 {t["max_column"]}/병합 {len(t["merges"])}개',
                             f'최대열 {ws.max_column}/병합 {len(ws.merged_cells.ranges)}개',
                             '서식 템플릿 미반영',
                             'VERIFICATION D-3 (신고서류 서식 템플릿)',
                             'extra_docs.py:426 — _find_template(base_dir, "영세율첨부서류명세서 양식.xlsx") '
                             '이지만 실제 파일은 forms/영세율첨부서류제출명세서 양식.xlsx ("제출" 누락) '
                             '→ 항상 None → _new_zero_template() 폴백')
            else:
                rep.ok('영세율첨부서류 서식 템플릿 반영',
                       f'최대열 {ws.max_column}/병합 {len(ws.merged_cells.ranges)}개')

        rows = []
        for r in range(3, ws.max_row + 1):
            fx = num(ws.cell(r, ZERO_COLS['제출외화']).value)
            if fx is None:
                continue
            rows.append({r_: ws.cell(r, c).value for r_, c in ZERO_COLS.items()} | {'row': r})
        if not rows:
            rep.broken(f'{path.name} 데이터 행 0건')
            continue

        if is_total:
            bad_pairs = [x['row'] for x in rows
                         if x['제출외화'] != x['당기외화'] or x['제출원화'] != x['당기원화']]
            bad_future = [x['row'] for x in rows
                          if (num(x['미도래외화']) or 0) != 0 or (num(x['미도래원화']) or 0) != 0]
            bad_issuer = [x['row'] for x in rows if not str(x['발급자'] or '').strip()]
            bad_kind = [x['row'] for x in rows if x['구분'] != 1 or x['서류명'] != '소포수령증']
            bad_calc = [x['row'] for x in rows
                        if abs((num(x['제출원화']) or 0)
                               - round((num(x['제출외화']) or 0) * (num(x['환율']) or 0))) > 1]
            for label, bad, basis in (
                ('당기 제출 = 당기 신고 해당분', bad_pairs, 'VERIFICATION D-3'),
                ('미도래 금액 0', bad_future, 'VERIFICATION A-5 (미도래 0)'),
                ('발급자 채워짐', bad_issuer, 'VERIFICATION D-3 (_company_name 제출자)'),
                ('구분=1 / 서류명=소포수령증', bad_kind, 'VERIFICATION D-3'),
                ('원화 = 외화 × 환율', bad_calc, 'VERIFICATION D-3'),
            ):
                if bad:
                    rep.mismatch(f'영세율 {label}', '전건 충족',
                                 f'{len(bad)}건 위반(행 {bad[:5]})', len(bad), basis,
                                 'extra_docs.py:379-382 _write_zero_sheet 열 매핑')
                else:
                    rep.ok(f'영세율 {label}', f'{len(rows)}행')
            total_file = {'rows': rows,
                          'sum_krw': sum(int(num(x['제출원화']) or 0) for x in rows)}
        else:
            monthly_sum += sum(int(num(x['제출원화']) or 0) for x in rows)
            monthly_rows += len(rows)

    if total_file and monthly_rows:
        if monthly_sum != total_file['sum_krw'] or monthly_rows != len(total_file['rows']):
            rep.mismatch('영세율 월별 파일 합산',
                         f'{total_file["sum_krw"]:,}원/{len(total_file["rows"])}행',
                         f'{monthly_sum:,}원/{monthly_rows}행',
                         monthly_sum - total_file['sum_krw'],
                         'VERIFICATION C-1 (월별 분할이 건을 잃지 않는가)',
                         'extra_docs.py:450-454 date_to_month_key 분할')
        else:
            rep.ok('영세율 월별 파일 합산 = 전체 파일', f'{monthly_sum:,}원/{monthly_rows}행')
    return total_file


# ── C-3 중복 (경고 등급) ────────────────────────────────────────

def check_c3(built, rep):
    """같은 원본 행이 두 번 계상되지 않았는지 확인합니다.

    추적번호만으로 판정하면 위양성이 납니다 — 라자다 주문 Excel·쇼피파이는
    라인아이템 단위라 한 소포(한 추적번호)에 동일 금액 행이 여럿인 것이 정상입니다.
    그래서 **원본 행 식별자**(order_item_id / source_file+source_row / row_index)로
    중복을 봅니다. 병합(merge_*_results)이 같은 행을 두 번 넣으면 여기서 잡힙니다.
    """
    ids = {}
    fallback = 0
    for res in ([built['lazada']] if built.get('lazada') else []) + list(built.get('shopify') or []):
        for it in res.get('items', []):
            row_id = it.get('source_row') if it.get('source_row') is not None else it.get('row_index')
            if it.get('order_item_id'):
                key = ('item', it['order_item_id'])
            elif row_id is not None:
                key = ('row', it.get('source_file'), row_id)
            else:
                fallback += 1
                key = ('fb', it.get('tracking_no'), it.get('date'), it.get('amount'))
            ids[key] = ids.get(key, 0) + 1
    for sd in built.get('shopee') or []:
        for tx in sd.get('transactions', []):
            key = ('shopee', sd.get('currency'), tx.get('tracking_no'), tx.get('date'), tx.get('amount'))
            ids[key] = ids.get(key, 0) + 1
    for jr in built.get('joom') or []:
        for it in jr.get('items', []):
            key = ('joom', it.get('order_id'), it.get('tracking_no'), it.get('date'), it.get('amount'))
            ids[key] = ids.get(key, 0) + 1

    dup = {k: v for k, v in ids.items() if v > 1}
    if dup:
        rep.mismatch('C-3 원본 행 중복 계상', '0건',
                     f'{len(dup)}종 {sum(dup.values())}건 (예: {list(dup)[:3]})', len(dup),
                     'VERIFICATION C-3 (중복 계상 금지)',
                     'merge_lazada_results / merge_shopify_results 병합부')
    else:
        rep.ok('C-3 원본 행 중복 없음',
               f'{len(ids)}행' + (f' (식별자 없어 대체키 사용 {fallback}행)' if fallback else ''))

    # 참고: 추적번호 재사용 현황 — 라인아이템 단위 자료에서는 정상입니다.
    tn = {}
    for r in built.get('decl') or []:
        t = str(r.get('tracking_no') or '')
        if t:
            tn[t] = tn.get(t, 0) + 1
    reused = {k: v for k, v in tn.items() if v > 1}
    if reused:
        rep.skip('C-3 추적번호 재사용(참고)',
                 f'{len(reused)}종 {sum(reused.values())}건 — 한 소포 다수 라인아이템으로 정상')


# ── D-4 쇼피파이 원본 재현 ──────────────────────────────────────

def read_csv_table(path):
    """쇼피파이 orders CSV 를 감사기가 독립적으로 다시 읽습니다.

    파서를 경유하지 않되 인코딩 후보와 헤더 행 탐색은 같은 방식을 씁니다
    (shopify_parser._read_csv_rows / _find_header_row).
    """
    raw = Path(path).read_bytes()
    text = None
    for enc in ('utf-8-sig', 'utf-8', 'cp949', 'euc-kr'):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode('utf-8', errors='replace')
    rows = list(csv.reader(io.StringIO(text, newline='')))
    for i, row in enumerate(rows[:20]):
        norm = {str(c).strip().lower() for c in (row or [])}
        if {'name', 'total', 'currency'}.issubset(norm):
            return [str(c).strip() for c in row], [r for r in rows[i + 1:] if any(str(c).strip() for c in r)]
    return None, []


def check_d4(wb, shopify_results, csv_paths, rep):
    """쇼피파이 시트가 원본 CSV 전 건을 설명하는지 확인합니다.

    주의: VERIFICATION.md D-4 는 "headers+rows 로 원본 CSV 전 행을 그대로 재현"
    이라고 적었지만, 커밋 ab82a8f 이후 write_shopify_sheet 는 원본 덤프가 아니라
    간결한 표(10열)를 씁니다. 그래서 '전 건이 설명되는가'로 판정합니다.
    """
    sheets = [ws for ws in wb.worksheets if ws.title.startswith('쇼피파이')]
    if not sheets:
        rep.skip('D-4 쇼피파이 원본 재현', '쇼피파이 시트 없음')
        return
    # 원본 CSV 를 감사기가 직접 다시 읽습니다 (파서를 경유하지 않음)
    csv_names = {}
    csv_rows = 0
    for p in csv_paths:
        headers, rows = read_csv_table(p)
        if not headers:
            rep.broken(f'{p.name} 헤더 행을 찾지 못했습니다')
            continue
        idx = {h: i for i, h in enumerate(headers)}
        name_col = idx.get('Name')
        csv_rows += len(rows)
        for row in rows:
            name = str(row[name_col]).strip() if name_col is not None and name_col < len(row) else ''
            if name:
                csv_names.setdefault(name, row)

    sheet_rows = 0
    unknown = []
    for ws in sheets:
        headers = [ws.cell(3, c).value for c in range(1, 11)]
        if headers != SHOPIFY_HEADERS:
            rep.mismatch(f'{ws.title}!A3:J3', SHOPIFY_HEADERS, headers, '헤더 불일치',
                         'VERIFICATION D-4', 'excel_writer.py:1216-1219')
            continue
        for r in range(4, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if label == '합계(매출 반영)':
                continue
            name = ws.cell(r, 2).value
            if not name:
                continue
            sheet_rows += 1
            if str(name) not in csv_names:
                unknown.append((ws.title, r, name))

    blank = sum(int(sr.get('skipped_blank', 0) or 0) for sr in (shopify_results or []))
    explained = sheet_rows + blank
    if explained != csv_rows:
        rep.mismatch('쇼피파이 전 건 설명', csv_rows, explained, explained - csv_rows,
                     'VERIFICATION D-4/C-1 (원본 전 행이 설명되는가)',
                     f'시트행 {sheet_rows} + 라인아이템 연속행 {blank} ≠ CSV 데이터행 {csv_rows}')
    else:
        rep.ok('쇼피파이 전 건 설명', f'시트행 {sheet_rows} + 라인아이템 {blank} = CSV {csv_rows}')
    if unknown:
        rep.mismatch('쇼피파이 주문번호 원본 확인', '전건 CSV에 존재',
                     f'{len(unknown)}건 미확인({unknown[:3]})', len(unknown),
                     'VERIFICATION D-4', 'shopify_parser order_name 추출')
    else:
        rep.ok('쇼피파이 주문번호 전건 원본 CSV 확인', f'{sheet_rows}건')


# ── A-5 6중 대조 ────────────────────────────────────────────────

def check_a5(values, rep):
    """총집계·월별집계·통화시트·수출실적·영세율·독립재계산 6중 대조 (허용오차 0원)."""
    present = {k: v for k, v in values.items() if v is not None}
    if len(present) < 2:
        rep.skip('A-5 다중 대조', f'비교 가능한 값이 {len(present)}개')
        return
    uniq = set(present.values())
    if len(uniq) == 1:
        rep.ok('A-5 ' + '='.join(present) + ' 일치',
               f'{next(iter(uniq)):,}원 ({len(present)}중 일치)')
        return
    base = present.get('총집계') or next(iter(present.values()))
    detail = ' / '.join(f'{k} {v:,}' for k, v in present.items())
    rep.mismatch('A-5 다중 대조', f'{base:,} (전부 동일)', detail,
                 max(present.values()) - min(present.values()),
                 'VERIFICATION A-5 (총집계=월별집계=통화별시트=신고서류 일치)',
                 '갈라진 쌍이 원인 지문: 총집계≠월별집계→귀속월 상충 / 통화시트≠총집계→'
                 '_get_rate 폴백 / 수출실적≠통화시트→build_declaration_rows 플랫폼 블록 / '
                 '독립재계산만 다름→rates dict 구성')


# ── 케이스 조립 (--build) ───────────────────────────────────────

def quiet(fn, *a, **kw):
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        return fn(*a, **kw)


def parse_input(path):
    suf = path.suffix.lower()
    if suf == '.pdf':
        return quiet(parse_pdf, str(path))
    if suf in ('.csv', '.xlsx', '.xlsm'):
        if quiet(is_shopify_orders_file, path):
            return quiet(parse_shopify_orders, path)
        if suf != '.csv' and quiet(is_lazada_order_excel, path):
            return quiet(parse_lazada_order_excel, path)
    return None


def build_rate_map(daily_curs, monthly_curs):
    """data/ 캐시만으로 rates dict 를 조립합니다 (SMBS 접속 없음).

    일별 캐시는 1통화로 정규화된 값, 월평균 캐시는 SMBS 원문(100통화) 값입니다.
    둘 다 round_applied_rate 가 자릿수로 다시 정규화하므로 그대로 넘깁니다.
    """
    daily = load_rate_cache()
    monthly = load_monthly_rate_cache()
    out = {}
    for cur in sorted(set(daily_curs) | set(monthly_curs)):
        entry = {'currency': cur, 'currency_name': cur, 'period': '', 'average': 0.0,
                 'min': 0.0, 'min_date': '', 'max': 0.0, 'max_date': '', 'range': 0.0,
                 'cross_rate': 0.0, 'display_start': '', 'display_end': '', 'daily': []}
        if cur in daily_curs:
            d = daily[daily['currency'] == cur].sort_values('date')
            entry['daily'] = [{'date': dt.strftime('%Y.%m.%d'), 'rate': float(r), 'change': 0, 'cross': 0}
                              for dt, r in zip(d['date'], d['rate'])]
            if len(d):
                entry['average'] = float(d['rate'].mean())
                entry['period'] = (f"{d['date'].iloc[0]:%Y.%m.%d} ~ {d['date'].iloc[-1]:%Y.%m.%d}")
        if cur in monthly_curs:
            m = monthly[monthly['currency'] == cur].sort_values('year_month')
            entry['monthly'] = [{'year_month': ym, 'rate': float(r)}
                                for ym, r in zip(m['year_month'], m['rate'])]
        out[cur] = entry
    return out


def _all_dates(shopee, lazada, ebay, joom, shopify, qoo10):
    dates = []
    for sd in shopee:
        dates += [t.get('date') for t in sd.get('transactions', [])]
    for res in ([lazada] if lazada else []) + list(ebay) + list(joom) + list(shopify):
        dates += [it.get('date') for it in res.get('items', [])]
    if qoo10:
        dates.append(qoo10.get('period_end') or qoo10.get('write_date'))
    keys = sorted(re.sub(r'\D', '', str(d))[:8] for d in dates if d)
    return [k for k in keys if len(k) == 8]


def build_case(case, outdir, rep):
    """samples/<case>/input 을 조립해 산출물 3종을 생성합니다."""
    from modules.excel_writer import generate_excel
    from modules.extra_docs import (
        build_declaration_rows, company_name_from_results,
        create_export_performance, create_zero_rate_attachments,
    )

    inputs = sorted(p for p in (S / case / 'input').glob('*')
                    if p.is_file() and p.suffix.lower() in ('.pdf', '.csv', '.xlsx', '.xlsm'))
    if not inputs:
        rep.broken(f'samples/{case}/input 에 파일이 없습니다')
        return None

    shopee, lazada_parts, ebay, joom, shopify_parts, qoo10 = [], [], [], [], [], None
    csv_paths = []
    for p in inputs:
        r = parse_input(p)
        if not r:
            rep.warn(f'{p.name} 판별', '알려진 플랫폼', 'unknown/None', 1,
                     'VERIFICATION E-1 (판별 정확도)', 'pdf_parser.detect_pdf_type')
            continue
        t = r.get('type')
        if t == 'shopee':
            shopee.append(r)
        elif t == 'lazada':
            lazada_parts.append(r)
        elif t == 'ebay':
            ebay.append(r)
        elif t == 'joom':
            joom.append(r)
        elif t == 'shopify':
            shopify_parts.append(r)
            csv_paths.append(p)
        elif t == 'qoo10':
            qoo10 = r
        else:
            rep.warn(f'{p.name} 판별', '알려진 플랫폼', t, 1,
                     'VERIFICATION E-1', 'pdf_parser.parse_pdf 반환 type')

    lazada = merge_lazada_results(lazada_parts) if lazada_parts else None
    shopify = merge_shopify_results(shopify_parts) if shopify_parts else []

    daily_curs, monthly_curs = set(), set()
    for sd in shopee:
        if sd.get('currency'):
            daily_curs.add(sd['currency'])
    for res in ([lazada] if lazada else []) + list(joom) + list(shopify):
        daily_curs |= {it.get('currency') for it in res.get('items', []) if it.get('currency')}
    for er in ebay:
        monthly_curs |= {it.get('currency') for it in er.get('items', []) if it.get('currency')}
    if qoo10:
        monthly_curs.add('JPY')
    rates = build_rate_map(daily_curs, monthly_curs)

    keys = _all_dates(shopee, lazada, ebay, joom, shopify, qoo10)
    year, month = (int(keys[-1][:4]), int(keys[-1][4:6])) if keys else (2026, 6)

    xlsx = outdir / f'매출집계_{case}.xlsx'
    quiet(generate_excel, shopee_results=shopee, lazada_result=lazada, qoo10_result=qoo10,
          rates=rates, output_path=str(xlsx), year=year, month=month,
          ebay_results=ebay, joom_results=joom, shopify_results=shopify)

    decl = build_declaration_rows(shopee, lazada, qoo10, rates, ebay_results=ebay,
                                 joom_results=joom, shopify_results=shopify)
    company = company_name_from_results(shopee, lazada, qoo10, ebay_results=ebay,
                                        joom_results=joom, shopify_results=shopify)
    export = create_export_performance(decl, outdir, company, base_dir=ROOT)
    zero = create_zero_rate_attachments(decl, outdir, company, base_dir=ROOT, mode='both')
    return {'xlsx': xlsx, 'export': export, 'zero': zero, 'decl': decl,
            'company': company, 'shopify': shopify, 'csv_paths': csv_paths,
            'shopee': shopee, 'lazada': lazada, 'ebay': ebay, 'joom': joom,
            'qoo10': qoo10, 'rates': rates, 'year': year, 'month': month}


# ── 감사 실행 ───────────────────────────────────────────────────

def audit(xlsx, zero_paths=(), export_path=None, built=None, scope=''):
    built = built or {}
    decl_rows = built.get('decl')
    shopify_results = built.get('shopify')
    csv_paths = built.get('csv_paths') or ()
    rep = Report(scope or Path(xlsx).name)
    wb = open_book(xlsx, rep)
    if rep.struct:
        return rep

    check_d1(wb, rep)
    summary = read_summary(wb, rep)
    monthly = read_monthly(wb, rep)
    sheets = read_currency_sheets(wb, rep)
    check_b2_via_rate_sheet(wb, sheets, rep)
    if csv_paths:
        check_d4(wb, shopify_results, csv_paths, rep)

    export = read_export_performance(Path(export_path), rep) if export_path else None
    zero = read_zero_rate([Path(p) for p in zero_paths], rep) if zero_paths else None

    check_a5({
        '총집계': summary['total_krw'] if summary else None,
        '월별집계': monthly['grand_krw'] if monthly else None,
        '통화시트': sum(v['sum_krw'] for v in sheets.values()) if sheets else None,
        '수출실적': export['sum_krw'] if export else None,
        '영세율': zero['sum_krw'] if zero else None,
        '독립재계산': sum(int(r['krw']) for r in decl_rows) if decl_rows else None,
    }, rep)

    if built:
        check_c3(built, rep)
    return rep


def main():
    ap = argparse.ArgumentParser(description='산출물 자기완결 감사')
    ap.add_argument('xlsx', nargs='?', help='매출집계 xlsx 경로')
    ap.add_argument('--build', help=f'케이스 조립 후 감사 ({", ".join(CASES)}, all)')
    ap.add_argument('--zero', nargs='*', default=[], help='영세율첨부서류제출명세서 경로들')
    ap.add_argument('--export', help='수출실적명세서 경로')
    ap.add_argument('--keep', help='--build 산출물을 이 디렉터리에 남깁니다')
    args = ap.parse_args()

    reports = []
    if args.build:
        cases = CASES if args.build == 'all' else [args.build]
        for case in cases:
            if case not in CASES:
                print(f'알 수 없는 케이스: {case} (가능: {", ".join(CASES)}, all)')
                return 2
            with tempfile.TemporaryDirectory() as td:
                outdir = Path(args.keep) if args.keep else Path(td)
                outdir.mkdir(parents=True, exist_ok=True)
                assert not str(outdir.resolve()).startswith(str(S.resolve())), \
                    'samples/ 하위에는 절대 쓰지 않습니다'
                rep = Report(f'{case} 조립')
                built = build_case(case, outdir, rep)
                if built is None:
                    reports.append(rep)
                    continue
                rep2 = audit(built['xlsx'], zero_paths=built['zero'],
                             export_path=built['export'], built=built,
                             scope=f'{case} ({built["company"]}, {built["year"]}-{built["month"]:02d})')
                rep.lines.extend(rep2.lines)
                rep.n_mismatch += rep2.n_mismatch
                rep.n_warn += rep2.n_warn
                rep.n_pass += rep2.n_pass
                rep.struct.extend(rep2.struct)
                reports.append(rep)
    elif args.xlsx:
        reports.append(audit(args.xlsx, zero_paths=args.zero, export_path=args.export))
    else:
        ap.print_help()
        return 2

    for rep in reports:
        print(rep.render())
        print()

    struct = sum(len(r.struct) for r in reports)
    mism = sum(r.n_mismatch for r in reports)
    warn = sum(r.n_warn for r in reports)
    passed = sum(r.n_pass for r in reports)
    print(f'요약: 일치 {passed} / 불일치 {mism} / 경고 {warn} / 구조붕괴 {struct}')
    if struct:
        print('❌ 감사기가 판정할 수 없는 구조 붕괴가 있습니다.')
        return 2
    if mism:
        print(f'❌ 불일치 {mism}건')
        return 1
    print('✅ 전체 통과' + (f' (경고 {warn}건)' if warn else ''))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
