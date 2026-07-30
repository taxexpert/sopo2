# -*- coding: utf-8 -*-
"""실경로 E2E 검증 — 실제 사용자가 밟는 app.py 경로를 그대로 통과시킵니다.

`quick_check.py` 와 `audit_workbook.py` 는 모듈을 직접 조립하므로 app.py 의
오케스트레이션(app.py:733-961 인라인 블록)을 한 줄도 지나지 않습니다. 이 스크립트는
streamlit `AppTest` 로 app.py 를 스크립트째 실행해 그 구멍을 메웁니다.
app.py 에만 있는 로직 — `_build_qoo10_result()` · `_qoo10_reporting_month` ·
`_daily_rate_period_bounds` · `_needed_currencies` · 파일명 규칙 — 이 검증 대상입니다.

브라우저 자동화는 쓰지 않습니다. Streamlit 드롭존 셀렉터가 버전마다 바뀌고
다운로드 경로가 불안정해서, 실패 원인이 앱 결함인지 자동화 결함인지 갈리지 않습니다.
화면 확인은 `preview_start(name="sopo-streamlit")` 로 따로 봅니다.

사용법:
    python tools/e2e_apptest.py                 # 전 케이스 + 큐텐 10월 반기말 시나리오
    python tools/e2e_apptest.py lazada qoo10

종료코드: 0 통과 / 1 불일치 / 2 감사기 자체 오류(위젯 미발견·예외 등)
"""

import shutil
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import pandas as pd  # noqa: E402
from streamlit.testing.v1 import AppTest  # noqa: E402

from tools.audit_workbook import Report, audit, num  # noqa: E402

S = ROOT / 'samples'
FIXTURES = S / 'fixtures'

MIME = {
    '.pdf': 'application/pdf',
    '.csv': 'text/csv',
    '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    '.xlsm': 'application/vnd.ms-excel.sheet.macroEnabled.12',
}

CASES = ['shopee', 'lazada', 'qoo10', 'shopify', 'joom', 'joom2']

DOC_LABELS = ['매출집계', '영세율첨부서류제출명세서', '수출실적명세서']
PROCESS_BTN = '엑셀 파일 생성'
ZERO_MODE_LABEL = '영세율첨부서류제출명세서 생성 범위'

# 네트워크 진입점 — 하나라도 열려 있으면 SMBS 에 실제로 접속한다
NETWORK_FUNCS = [
    'fetch_smbs_period_rates',
    'try_fetch_std_rates_by_requests',
    'fetch_std_rates_by_selenium',
    'try_fetch_month_avg_rates_by_requests',
    'fetch_month_avg_rates_by_selenium',
]


# ── 가드 ────────────────────────────────────────────────────────

def install_guards(tmp):
    """반드시 AppTest 실행 전에 호출합니다.

    (1) `data/` 캐시 오염 방지 — `get_cached_or_fetch_smbs_period_rates` 는
        `save_rate_cache` 로 `data/` 를 실제로 덮어씁니다. 두 경로 상수는 호출
        시마다 다시 읽히므로 tmp 사본으로 바꿔치기하면 원본이 보호됩니다.
    (2) SMBS 실접속 차단 — 캐시가 모자라면 조용히 진행하지 말고 터뜨립니다.
    (3) 환율 조회 범위 기록 — `_daily_rate_period_bounds` 검증용.
    """
    import modules.exchange_rate as ex

    for name in ('exchange_rate_cache.csv', 'monthly_exchange_rate_cache.csv'):
        src = FIXTURES / name
        if not src.exists():
            raise SystemExit(f'환율 고정본이 없습니다: {src}')
        shutil.copy(src, tmp / name)
    ex.RATE_CACHE_FILE = tmp / 'exchange_rate_cache.csv'
    ex.MONTHLY_RATE_CACHE_FILE = tmp / 'monthly_exchange_rate_cache.csv'

    def blocked(*_a, **_kw):
        raise RuntimeError('NETWORK BLOCKED — 고정본 캐시가 이 구간을 덮지 못합니다')

    for name in NETWORK_FUNCS:
        if hasattr(ex, name):
            setattr(ex, name, blocked)

    calls = []
    orig_daily = ex.get_cached_or_fetch_smbs_period_rates

    def rec_daily(currency, start_date, end_date, quiet=False):
        calls.append(('daily', str(currency).upper(),
                      pd.Timestamp(start_date).normalize(), pd.Timestamp(end_date).normalize()))
        return orig_daily(currency, start_date, end_date, quiet=quiet)

    ex.get_cached_or_fetch_smbs_period_rates = rec_daily

    orig_month = ex.get_cached_or_fetch_month_avg_rates

    def rec_month(currency, start_month, end_month):
        calls.append(('monthly', str(currency).upper(), str(start_month), str(end_month)))
        return orig_month(currency, start_month, end_month)

    ex.get_cached_or_fetch_month_avg_rates = rec_month
    return calls


# ── 위젯 접근 ───────────────────────────────────────────────────

def find_widget(at, kind, label, rep):
    """라벨 문자열로 위젯을 찾습니다.

    인덱스 의존(at.button[0])은 UI 가 바뀌면 조용히 다른 위젯을 누르게 되므로
    라벨로만 찾고, 못 찾으면 구조 붕괴로 명확히 실패시킵니다.
    """
    for w in getattr(at, kind):
        if str(getattr(w, 'label', '') or '').strip() == label:
            return w
    have = [str(getattr(w, 'label', '') or '')[:30] for w in getattr(at, kind)]
    rep.broken(f'{kind} 라벨 "{label}" 를 찾을 수 없습니다 — 현재 목록 {have}')
    return None


def exceptions(at):
    return [str(e.value) for e in at.exception]


# ── 실행 ────────────────────────────────────────────────────────

def drive(at, rep, files=None, qoo10_entries=None):
    """업로드 → 문서 3종 선택 → 생성 버튼까지 실제 경로를 밟습니다."""
    at.run()
    if exceptions(at):
        rep.broken(f'초기 렌더에서 예외: {exceptions(at)[0][:300]}')
        return None

    if files:
        up = find_widget(at, 'file_uploader', 'PDF · Excel · CSV 파일 선택', rep)
        if up is None:
            # label_visibility="collapsed" 라도 라벨은 남지만, 혹시 없으면 단독 위젯을 씁니다
            if len(at.file_uploader) == 1:
                up = at.file_uploader[0]
                rep.struct.clear()
                rep.lines = [ln for ln in rep.lines if '[구조붕괴]' not in ln]
            else:
                return None
        up.set_value(files)
        at.run()
        if exceptions(at):
            rep.broken(f'업로드 직후 예외: {exceptions(at)[0][:300]}')
            return None

    if qoo10_entries is not None:
        at.session_state.qoo10_entries = qoo10_entries
        at.run()

    for label in DOC_LABELS:
        cb = find_widget(at, 'checkbox', label, rep)
        if cb is None:
            return None
        cb.set_value(True)
    at.run()

    radio = find_widget(at, 'radio', ZERO_MODE_LABEL, rep)
    if radio is not None:
        radio.set_value('전체')
        at.run()

    btn = find_widget(at, 'button', PROCESS_BTN, rep)
    if btn is None:
        return None
    btn.click()
    at.run()

    errs = exceptions(at)
    if errs:
        rep.mismatch('처리 중 예외', '예외 0건', errs[0][:400], len(errs),
                     'VERIFICATION E-2 (조용한 오답 금지 — 예외는 드러나야 한다)',
                     'app.py:958-961 except 블록')
        return None

    result = list(at.session_state.result_files or [])
    if not result:
        rep.mismatch('생성 산출물', '3종 이상', '0개', 3,
                     'VERIFICATION D-1 (산출물 생성)', 'app.py:900-953 created/result_files')
        return None
    rep.ok('실경로 산출물 생성', f'{len(result)}개: ' + ', '.join(f["name"] for f in result))
    return result


def check_filenames(result, rep):
    """app.py:894-945 의 파일명 규칙 — 제출자 미추출 재발 감시(2026-07-29 라자다 결함 ②)."""
    names = [f['name'] for f in result]
    sales = [n for n in names if n.startswith('매출집계_')]
    zero = [n for n in names if n.startswith('영세율첨부서류제출명세서_')]
    export = [n for n in names if n.startswith('수출실적명세서_')]
    for label, got in (('매출집계_*', sales), ('영세율첨부서류제출명세서_*', zero),
                       ('수출실적명세서_*', export)):
        if not got:
            rep.mismatch(f'파일명 규칙 {label}', '1개 이상', '0개', 1,
                         'VERIFICATION D-3', 'app.py:906/932/942 파일명 조립')
    placeholder = [n for n in names if '사업자명(사업자번호)' in n or '제출자' in n]
    if placeholder:
        rep.mismatch('제출자 추출', '실제 사업자명', f'플레이스홀더 {placeholder}', len(placeholder),
                     'VERIFICATION D-3 (_company_name 제출자 정상 추출)',
                     'extra_docs.py:118-144 _company_name / 파서 submitter 미반환')
    else:
        rep.ok('파일명 규칙·제출자 추출', ', '.join(names[:3]) + (' …' if len(names) > 3 else ''))


def check_rate_window(calls, at, rep):
    """`_daily_rate_period_bounds` 가 실제 기준일 범위 + 소급 7일을 요청했는지."""
    daily = [c for c in calls if c[0] == 'daily']
    if not daily:
        rep.skip('환율 조회 범위', '일별 환율 요청 없음 (월평균 전용 케이스)')
        return
    starts = {c[2] for c in daily}
    ends = {c[3] for c in daily}
    if len(starts) != 1 or len(ends) != 1:
        rep.mismatch('환율 조회 범위 일관성', '모든 통화 동일 구간',
                     f'start {sorted(starts)} / end {sorted(ends)}', len(starts) + len(ends),
                     'VERIFICATION B-1', 'app.py:849-859 rate_start/rate_end')
        return
    start, end = starts.pop(), ends.pop()
    lookback = (end - start).days
    rep.ok('환율 조회 범위', f'{start:%Y-%m-%d} ~ {end:%Y-%m-%d} '
           f'({len(daily)}통화, 소급 포함 {lookback}일)')

    from modules.exchange_rate import RATE_LOOKBACK_DAYS
    # display 시작일 = start + LOOKBACK 이어야 한다 (app.py:858)
    if (start + pd.Timedelta(days=RATE_LOOKBACK_DAYS)) > end:
        rep.mismatch('환율 소급 구간', f'{RATE_LOOKBACK_DAYS}일 소급 후에도 start <= end',
                     f'{start:%Y-%m-%d} + {RATE_LOOKBACK_DAYS}일 > {end:%Y-%m-%d}', 0,
                     'VERIFICATION B-4 (휴일 직전 영업일 확보)', 'app.py:857-859')


def check_b5_triple(result, at, tmp, rep):
    """큐텐 3중 일치 — app.py 수집 / excel_writer 조회 / extra_docs 신고행."""
    import openpyxl

    entries = list(at.session_state.qoo10_entries or [])
    if not entries:
        rep.skip('B-5 큐텐 3중 일치', '큐텐 데이터 없음')
        return
    app_rates = {(e.get('rate_month'), round(float(e.get('rate') or 0), 4)) for e in entries}
    if not all(e.get('rate') for e in entries):
        rep.mismatch('B-5 app.py 축 환율', '전건 설정', f'{entries}', 0,
                     'VERIFICATION B-5', 'excel_writer.py:2066-2073 이 entries 에 rate 를 채우지 못함')
        return

    sales = next((f for f in result if f['name'].startswith('매출집계_')), None)
    export = next((f for f in result if f['name'].startswith('수출실적명세서_')), None)
    sheet_rates, decl_rates, ship_dates = set(), set(), set()

    if sales:
        p = tmp / sales['name']
        p.write_bytes(sales['bytes'])
        wb = openpyxl.load_workbook(p, data_only=True)
        if 'JPY' in wb.sheetnames:
            ws = wb['JPY']
            for r in range(5, ws.max_row + 1):
                if ws.cell(r, 4).value != 'JPY':
                    continue
                rate = num(ws.cell(r, 5).value)
                if rate:
                    sheet_rates.add(round(rate, 4))
                    ship_dates.add(ws.cell(r, 3).value)
    if export:
        p = tmp / export['name']
        p.write_bytes(export['bytes'])
        wb = openpyxl.load_workbook(p, data_only=True)
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 4).value != 'JPY':
                continue
            rate = num(ws.cell(r, 5).value)
            if rate:
                decl_rates.add(round(rate, 4))

    app_only = {r for _m, r in app_rates}
    if not (app_only == sheet_rates == decl_rates):
        rep.mismatch('B-5 큐텐 3중 일치',
                     f'app.py {sorted(app_only)} 와 동일',
                     f'JPY시트 {sorted(sheet_rates)} / 신고서류 {sorted(decl_rates)}', 0,
                     'VERIFICATION B-5 (큐텐 삼중 일치)',
                     'app.py `_qoo10_reporting_month` / excel_writer 동명 함수 / '
                     'extra_docs.py:266-272 — 세 곳이 같은 월을 반환해야 함')
    else:
        rep.ok('B-5 큐텐 3중 일치',
               f'환율 {sorted(app_only)} (월 {sorted({m for m, _r in app_rates})}), '
               f'선적일자 {sorted(d for d in ship_dates if d)}')

    # 선적일자 = 거래기간 종료일. 종료일이 비어 있을 때만 반기 말일로 보완한다
    # (extra_docs.qoo10_reporting_date 독스트링). 반기말은 '환율 조회월'에만 적용된다.
    import re as _re
    expect_dates = set()
    for e in entries:
        end = _re.sub(r'\D', '', str(e.get('period_end') or ''))[:8]
        if len(end) == 8:
            expect_dates.add(int(end))
        else:
            base = _re.sub(r'\D', '', str(e.get('period_start') or e.get('write_date') or ''))[:8]
            if len(base) >= 6:
                expect_dates.add(int(f'{base[:4]}0630' if int(base[4:6]) <= 6 else f'{base[:4]}1231'))
    got_dates = {int(d) for d in ship_dates if d}
    if expect_dates and got_dates != expect_dates:
        rep.mismatch('큐텐 선적일자', sorted(expect_dates), sorted(got_dates), 0,
                     'VERIFICATION B-5 (선적일자 = 거래기간 종료일, 없으면 반기말)',
                     'extra_docs.qoo10_reporting_date / excel_writer._qoo10_reporting_date')
    elif expect_dates:
        rep.ok('큐텐 선적일자 = 거래기간 종료일', f'{sorted(got_dates)}')
    # 환율 조회월은 반기말 월이어야 한다
    months = {m for m, _r in app_rates if m}
    bad_m = [m for m in months if not str(m).endswith(('-06', '-12'))]
    if bad_m:
        rep.mismatch('큐텐 환율 조회월 반기말', 'YYYY-06 또는 YYYY-12', f'{bad_m}', len(bad_m),
                     'VERIFICATION B-5', 'app.py `_qoo10_reporting_month`')


# ── 케이스 실행 ─────────────────────────────────────────────────

def run_case(case, rep):
    inputs = sorted(p for p in (S / case / 'input').glob('*')
                    if p.is_file() and p.suffix.lower() in MIME)
    if not inputs:
        rep.broken(f'samples/{case}/input 에 파일이 없습니다')
        return
    files = [(p.name, p.read_bytes(), MIME[p.suffix.lower()]) for p in inputs]

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        calls = install_guards(tmp)
        at = AppTest.from_file(str(ROOT / 'app.py'), default_timeout=900)
        result = drive(at, rep, files=files)
        if not result:
            return
        check_filenames(result, rep)
        check_rate_window(calls, at, rep)
        check_b5_triple(result, at, tmp, rep)

        # 실경로 산출물을 방법 1 감사기로 재검증
        sales = next((f for f in result if f['name'].startswith('매출집계_')), None)
        if not sales:
            rep.mismatch('매출집계 산출물', '1개', '0개', 1, 'VERIFICATION D-1', 'app.py:905-921')
            return
        paths = {}
        for f in result:
            p = tmp / f['name']
            p.write_bytes(f['bytes'])
            paths[f['name']] = p
        sub = audit(paths[sales['name']],
                    zero_paths=[v for k, v in paths.items() if k.startswith('영세율')],
                    export_path=next((v for k, v in paths.items() if k.startswith('수출실적')), None),
                    scope=f'{case} 실경로 산출물 감사')
        rep.lines.extend(sub.lines)
        rep.n_mismatch += sub.n_mismatch
        rep.n_warn += sub.n_warn
        rep.n_pass += sub.n_pass
        rep.struct.extend(sub.struct)


def run_qoo10_halfyear_scenario(rep):
    """샘플에 없는 10월 큐텐 건을 session_state 에 직접 넣어 반기말 규칙을 확인합니다.

    VERIFICATION B-5 의 핵심 — 기준일이 10월이어도 환율 조회월은 12월(하반기말),
    선적일자는 그 달 말일이어야 합니다.
    """
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        calls = install_guards(tmp)
        at = AppTest.from_file(str(ROOT / 'app.py'), default_timeout=900)
        entries = [{
            'period_start': '2025-10-01', 'period_end': '2025-10-31',
            'tracking_no': 'K2510319999001', 'qty': 3, 'amount': 100000.0,
            'write_date': '2025-11-05',
        }]
        result = drive(at, rep, qoo10_entries=entries)
        if not result:
            return
        months = {e.get('rate_month') for e in at.session_state.qoo10_entries}
        if months != {'2025-12'}:
            rep.mismatch('큐텐 10월 건 환율 조회월', '2025-12 (하반기말)', f'{months}', 0,
                         'VERIFICATION B-5 (10월 PDF도 환율은 12월 월평균)',
                         'app.py `_qoo10_reporting_month` / excel_writer 동명 함수')
        else:
            rep.ok('큐텐 10월 건 → 환율 12월 월평균', f'{months}')
        check_b5_triple(result, at, tmp, rep)
        monthly_calls = {c[2] for c in calls if c[0] == 'monthly'}
        if monthly_calls and monthly_calls != {'2025-12'}:
            rep.mismatch('큐텐 월평균 요청 월', '2025-12', f'{sorted(monthly_calls)}', 0,
                         'VERIFICATION B-5 (수집 월 = 조회 월)',
                         'app.py `_monthly_rate_requests`')
        else:
            rep.ok('큐텐 월평균 요청 월 = 조회 월', f'{sorted(monthly_calls)}')


def main():
    wanted = sys.argv[1:] or CASES + ['큐텐10월']
    reports = []
    for case in wanted:
        if case == '큐텐10월':
            rep = Report('큐텐 10월 반기말 시나리오 (E2E)')
            run_qoo10_halfyear_scenario(rep)
        elif case in CASES:
            rep = Report(f'{case} (E2E 실경로)')
            run_case(case, rep)
        else:
            print(f'알 수 없는 케이스: {case} (가능: {", ".join(CASES)}, 큐텐10월)')
            return 2
        reports.append(rep)
        print(rep.render())
        print(flush=True)

    struct = sum(len(r.struct) for r in reports)
    mism = sum(r.n_mismatch for r in reports)
    warn = sum(r.n_warn for r in reports)
    passed = sum(r.n_pass for r in reports)
    print(f'요약: 일치 {passed} / 불일치 {mism} / 경고 {warn} / 구조붕괴 {struct}')
    if struct:
        print('❌ 감사기가 판정할 수 없는 구조 붕괴가 있습니다.')
        return 2
    if mism:
        print(f'❌ 불일치 {mism}건')
        return 1
    print('✅ 전체 통과' + (f' (경고 {warn}건)' if warn else ''))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
