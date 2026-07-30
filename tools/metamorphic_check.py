# -*- coding: utf-8 -*-
"""메타모픽 불변식 검증 — 정답지에 의존하지 않는 축.

`samples/*/expected/` 를 **읽지도 쓰지도 않습니다.** 대신 입력을 규칙에 따라 변형한 뒤
산출물이 어떻게 변해야/변하지 않아야 하는지를 판정합니다. 2026-07-29 쇼피 MX 사고
(약 -380만원) 처럼 `expected` 자체가 버그를 등록한 경우 회귀 축은 통과시켜 버리는데,
이 축은 정답을 몰라도 규칙 위반을 잡습니다.

불변식 목록은 docs/VERIFICATION.md 의 A-1(대체)·B-1·B-3·B-4·C-1·C-4·E-1·E-3 에 대응합니다.
성립하지 않아도 정상인 것(같은 파일 2회 투입 시 이중계상, xlsx 바이트 재현성,
외화 완전 일치, 분할 시 시트명 변화)은 판정 대상에서 제외했습니다 — 아래 주석 참조.

사용법:
    python tools/metamorphic_check.py            # 전체
    python tools/metamorphic_check.py MR-4b MR-5

종료코드: 0 통과 / 1 불일치(또는 SKIP 3건 이상) / 2 하니스 자체 오류
"""

import random
import re
import shutil
import sys
import tempfile
import warnings
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

warnings.simplefilter('ignore')

import openpyxl  # noqa: E402
import pandas as pd  # noqa: E402

from tools.audit_workbook import (  # noqa: E402
    Report, CURRENCY_SHEET_RE, parse_input, build_rate_map, quiet, num, read_summary,
)
from tools.e2e_apptest import install_guards  # noqa: E402

S = ROOT / 'samples'
DELIVERED_COL = 16   # deliveredDate
CREATE_COL = 9       # createTime
UPDATE_COL = 10      # updateTime
LZ_DATE_FMT = '%d %b %Y %H:%M'


# ── 하니스 ──────────────────────────────────────────────────────

class Run:
    """한 번의 파이프라인 실행 결과 — 비교 가능한 관측값만 담습니다."""

    def __init__(self, xlsx, detected):
        self.xlsx = xlsx
        self.detected = detected
        wb = openpyxl.load_workbook(xlsx, data_only=True)
        self.sheet_names = set(wb.sheetnames)
        self.krw_by_currency = {}
        self.fx_by_currency = {}
        self.rows_by_currency = {}
        for ws in wb.worksheets:
            if not CURRENCY_SHEET_RE.match(ws.title):
                continue
            header_row = 4 if ws.title == 'JPY' else 6
            krw = 0
            fx = 0.0
            rows = []
            for r in range(header_row + 1, ws.max_row + 1):
                if ws.cell(r, 4).value != ws.title:
                    continue
                v_fx = num(ws.cell(r, 6).value)
                if v_fx is None:
                    continue
                krw += int(num(ws.cell(r, 7).value) or 0)
                fx += v_fx
                rows.append({'ship_date': ws.cell(r, 3).value,
                             'rate': num(ws.cell(r, 5).value), 'fx': v_fx,
                             'krw': num(ws.cell(r, 7).value)})
            self.krw_by_currency[ws.title] = krw
            self.fx_by_currency[ws.title] = round(fx, 2)
            self.rows_by_currency[ws.title] = rows
        self.total_krw = sum(self.krw_by_currency.values())
        self.cell_dump = {(ws.title, c.row, c.column): c.value
                          for ws in wb.worksheets for row in ws.iter_rows() for c in row}


def run_pipeline(paths, tmp, tag, drop_monthly=()):
    """파서 → 환율 조립 → generate_excel 을 한 번 돌립니다.

    `drop_monthly` 에 (통화, 'YYYY-MM') 을 넣으면 그 월평균을 빼고 조립합니다
    (MR-6: 공식 월평균 결손 시 조용히 대체하지 않는지 확인).
    """
    from modules.excel_writer import generate_excel
    from modules.lazada_order_parser import merge_lazada_results
    from modules.shopify_parser import merge_shopify_results
    from modules.pdf_parser import detect_pdf_type

    shopee, lazada_parts, ebay, joom, shopify_parts, qoo10 = [], [], [], [], [], None
    detected = {}
    for p in paths:
        if p.suffix.lower() == '.pdf':
            detected[p.name] = quiet(detect_pdf_type, str(p))
        r = parse_input(p)
        if not r:
            detected.setdefault(p.name, 'unknown')
            continue
        t = r.get('type')
        detected[p.name] = t
        {'shopee': shopee.append, 'lazada': lazada_parts.append, 'ebay': ebay.append,
         'joom': joom.append, 'shopify': shopify_parts.append}.get(t, lambda _x: None)(r)
        if t == 'qoo10':
            qoo10 = r

    lazada = merge_lazada_results(lazada_parts) if lazada_parts else None
    shopify = merge_shopify_results(shopify_parts) if shopify_parts else []

    daily_curs, monthly_curs = set(), set()
    for sd in shopee:
        if sd.get('currency'):
            daily_curs.add(sd['currency'])
    for res in ([lazada] if lazada else []) + list(joom) + list(shopify):
        daily_curs |= {it.get('currency') for it in res.get('items', []) if it.get('currency')}
    for er in ebay:
        monthly_curs |= {it.get('currency') for it in er.get('items', []) if it.get('currency')}
    if qoo10:
        monthly_curs.add('JPY')
    rates = build_rate_map(daily_curs, monthly_curs)
    for cur, month in drop_monthly:
        if cur in rates:
            rates[cur]['monthly'] = [m for m in rates[cur].get('monthly', [])
                                     if str(m.get('year_month')) != month]

    out = tmp / f'{tag}.xlsx'
    quiet(generate_excel, shopee_results=shopee, lazada_result=lazada, qoo10_result=qoo10,
          rates=rates, output_path=str(out), year=2026, month=6,
          ebay_results=ebay, joom_results=joom, shopify_results=shopify)
    return Run(out, detected)


def case_inputs(case, pattern='*'):
    return sorted(p for p in (S / case / 'input').glob(pattern)
                  if p.is_file() and p.suffix.lower() in ('.pdf', '.csv', '.xlsx', '.xlsm'))


# ── 라자다 Excel 변형 ───────────────────────────────────────────

def lazada_copy(src, dst, shift_delivered=None, shift_other=None, set_delivered=None):
    """라자다 주문 Excel 사본을 만들며 날짜 열만 바꿉니다 (원본은 읽기만)."""
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb.active
    changed = 0
    for r in range(2, ws.max_row + 1):
        if set_delivered is not None:
            if ws.cell(r, DELIVERED_COL).value:
                ws.cell(r, DELIVERED_COL, set_delivered)
                changed += 1
            continue
        if shift_delivered is not None:
            v = ws.cell(r, DELIVERED_COL).value
            new = _shift(v, shift_delivered)
            if new:
                ws.cell(r, DELIVERED_COL, new)
                changed += 1
        if shift_other is not None:
            for col in (CREATE_COL, UPDATE_COL):
                new = _shift(ws.cell(r, col).value, shift_other)
                if new:
                    ws.cell(r, col, new)
                    changed += 1
    wb.save(dst)
    return changed


def _shift(value, days):
    if not value:
        return None
    dt = pd.to_datetime(str(value), errors='coerce')
    if pd.isna(dt):
        return None
    return (dt + pd.Timedelta(days=days)).strftime(LZ_DATE_FMT)


# ── 불변식 ──────────────────────────────────────────────────────

def mr1_shuffle(tmp, rep):
    """MR-1 입력 순서 셔플 — 원화 완전 일치, 외화 <=0.01, 시트명 집합 동일."""
    paths = case_inputs('lazada')
    base = run_pipeline(paths, tmp, 'mr1_base')
    rng = random.Random(0)
    for i in range(1, 6):
        shuffled = list(paths)
        rng.shuffle(shuffled)
        run = run_pipeline(shuffled, tmp, f'mr1_{i}')
        if run.krw_by_currency != base.krw_by_currency:
            rep.mismatch(f'MR-1 셔플 #{i} 원화', base.krw_by_currency, run.krw_by_currency,
                         run.total_krw - base.total_krw,
                         'MR-1 / VERIFICATION A-1 (순서 무관 — 원화는 건별 round 후 정수 누적)',
                         'excel_writer 누산 순서 의존 / _shopee_sort_key 정렬')
            return
        off = {c: round(run.fx_by_currency[c] - v, 4)
               for c, v in base.fx_by_currency.items() if abs(run.fx_by_currency[c] - v) > 0.01}
        if off:
            rep.mismatch(f'MR-1 셔플 #{i} 외화', '차이 0.01 이하', off, off,
                         'MR-1', 'excel_writer.py:2123 float 누적')
            return
        if run.sheet_names != base.sheet_names:
            rep.mismatch(f'MR-1 셔플 #{i} 시트 집합', sorted(base.sheet_names),
                         sorted(run.sheet_names), 0, 'MR-1 / VERIFICATION D-1',
                         '_shopee_sheet_names_for_results 정렬 의존')
            return
    rep.ok('MR-1 입력 순서 셔플 5회', f'원화 {base.total_krw:,}원 불변, 시트 {len(base.sheet_names)}개 동일')


def mr2_reproducible(tmp, rep):
    """MR-2 동일 입력 2회 — 셀 값 전체 동일 (xlsx 바이트는 생성 타임스탬프 때문에 항상 다름)."""
    paths = case_inputs('lazada')
    a = run_pipeline(paths, tmp, 'mr2_a')
    b = run_pipeline(paths, tmp, 'mr2_b')
    diff = [k for k in a.cell_dump if a.cell_dump.get(k) != b.cell_dump.get(k)]
    extra = set(b.cell_dump) - set(a.cell_dump)
    if diff or extra:
        rep.mismatch('MR-2 재현성', '셀 값 전체 동일',
                     f'{len(diff)}셀 상이 + {len(extra)}셀 추가 (예: {diff[:3]})',
                     len(diff) + len(extra), 'MR-2 / VERIFICATION E-3 (재현성)',
                     '난수·시각 의존 또는 캐시 상태 변화')
    else:
        rep.ok('MR-2 재현성', f'{len(a.cell_dump):,}셀 전부 동일')


def mr3_split_merge(tmp, rep):
    """MR-3 분할/합침 — 쇼피 THB 1~3월 + 4~6월. 시트명은 의도적으로 달라지므로 총합만 본다."""
    th = case_inputs('shopee', '*_TH_*.pdf')
    if len(th) < 2:
        rep.skip('MR-3 분할/합침', f'쇼피 THB PDF가 {len(th)}개 (2개 필요)')
        return
    both = run_pipeline(th, tmp, 'mr3_both')
    parts = [run_pipeline([p], tmp, f'mr3_p{i}') for i, p in enumerate(th)]
    apart = sum(r.krw_by_currency.get('THB', 0) for r in parts)
    together = both.krw_by_currency.get('THB', 0)
    if apart != together:
        rep.mismatch('MR-3 THB 분할 합산', together, apart, apart - together,
                     'MR-3 / VERIFICATION C-4 (다중 파일 병합)',
                     'excel_writer.py:590-598 통화 시트 합침 / 건별 환율 적용')
    else:
        rep.ok('MR-3 THB 분할=합침', f'{together:,}원 (파일 {len(th)}개)')


def mr4_basis_date(tmp, rep):
    """MR-4a/4b 기준일 규칙 양방향 — 라자다 Excel 의 기준일은 deliveredDate 뿐이어야 한다."""
    src = next((p for p in case_inputs('lazada', '라자다_*.xlsx')), None)
    if src is None:
        rep.skip('MR-4 기준일', '라자다 주문 Excel 없음')
        return
    base = run_pipeline([src], tmp, 'mr4_base')
    cur = next(iter(base.krw_by_currency), None)
    if cur is None:
        rep.skip('MR-4 기준일', '통화 시트 없음')
        return

    # 4a — 기준일이 아닌 날짜열(createTime·updateTime)만 바꿔도 원화는 그대로여야 한다
    a = tmp / 'mr4a.xlsx'
    n = lazada_copy(src, a, shift_other=-30)
    run_a = run_pipeline([a], tmp, 'mr4a_out')
    if run_a.krw_by_currency != base.krw_by_currency:
        rep.mismatch('MR-4a 비기준일 변경', base.krw_by_currency, run_a.krw_by_currency,
                     run_a.total_krw - base.total_krw,
                     'MR-4a / VERIFICATION B-1 (라자다 Excel 기준일 = deliveredDate)',
                     'extra_docs.py:98 lazada_item_date / excel_writer._lazada_item_date 가 '
                     'createTime·updateTime 을 보고 있을 가능성')
    else:
        rep.ok('MR-4a 비기준일(createTime·updateTime) 변경 → 원화 불변',
               f'{n}셀 변경, {base.total_krw:,}원 유지')

    # 4b — 기준일을 바꾸면 원화가 반드시 변해야 한다 (단방향 검사는 무의미)
    b = tmp / 'mr4b.xlsx'
    lazada_copy(src, b, shift_delivered=-7)
    run_b = run_pipeline([b], tmp, 'mr4b_out')
    if run_b.krw_by_currency == base.krw_by_currency:
        rep.mismatch('MR-4b 기준일 -7일 변경', '원화가 변해야 함',
                     f'불변 {base.total_krw:,}원', 0,
                     'MR-4b / VERIFICATION B-1 (양방향)',
                     'deliveredDate 를 무시하고 문서 기준일 폴백을 쓰고 있을 가능성 — '
                     'extra_docs.py:105-115 lazada_item_rate')
    else:
        rep.ok('MR-4b 기준일 -7일 변경 → 원화 변동',
               f'{base.total_krw:,} → {run_b.total_krw:,} '
               f'({run_b.total_krw - base.total_krw:+,}원)')


def mr5_holiday_ffill(tmp, rep):
    """MR-5 휴일 ffill — 토요일 기준일에는 직전 금요일 환율이 적용돼야 한다."""
    from modules.exchange_rate import load_rate_cache, round_applied_rate
    src = next((p for p in case_inputs('lazada', '라자다_*.xlsx')), None)
    if src is None:
        rep.skip('MR-5 휴일 ffill', '라자다 주문 Excel 없음')
        return
    base = run_pipeline([src], tmp, 'mr5_base')
    cur = next(iter(base.krw_by_currency), None)
    if cur is None:
        rep.skip('MR-5 휴일 ffill', '통화 시트 없음')
        return

    # 캐시는 fill_missing_dates 로 휴일까지 채워져 있으므로, 먼저 캐시 자체의
    # ffill 불변식(토요일 값 = 직전 금요일 값)을 확인한 뒤 산출물에 적용된 값을 봅니다.
    cache = load_rate_cache(cur).sort_values('date')
    sat = fri = fri_rate = None
    for cand in reversed([d for d in cache['date'] if d.weekday() == 5]):
        prev = cand - pd.Timedelta(days=1)
        rs = cache[cache['date'] == cand]['rate']
        rf = cache[cache['date'] == prev]['rate']
        if len(rs) and len(rf):
            if float(rs.iloc[0]) != float(rf.iloc[0]):
                rep.mismatch(f'MR-5 {cur} 캐시 ffill', float(rf.iloc[0]), float(rs.iloc[0]),
                             float(rs.iloc[0]) - float(rf.iloc[0]),
                             f'MR-5 / VERIFICATION B-4 (토 {cand:%Y-%m-%d} = 금 {prev:%Y-%m-%d})',
                             'exchange_rate.fill_missing_dates ffill')
                return
            sat, fri, fri_rate = cand, prev, float(rf.iloc[0])
            break
    if sat is None:
        rep.skip('MR-5 휴일 ffill', f'{cur} 캐시에 (토·금) 쌍이 없음')
        return

    mod = tmp / 'mr5.xlsx'
    lazada_copy(src, mod, set_delivered=sat.strftime(LZ_DATE_FMT))
    run = run_pipeline([mod], tmp, 'mr5_out')
    expect = round_applied_rate(cur, fri_rate)
    rows = run.rows_by_currency.get(cur, [])
    target = int(sat.strftime('%Y%m%d'))
    bad = [r for r in rows if r['ship_date'] == target and abs((r['rate'] or 0) - expect) > 1e-6]
    if not rows:
        rep.skip('MR-5 휴일 ffill', '변형 후 데이터 행 없음')
    elif bad:
        rep.mismatch(f'MR-5 {cur} 토요일({sat:%Y-%m-%d}) 적용환율', expect, bad[0]['rate'],
                     round((bad[0]['rate'] or 0) - expect, 6),
                     f'MR-5 / VERIFICATION B-4 (직전 영업일 {fri:%Y-%m-%d})',
                     'exchange_rate.fill_missing_dates ffill / get_rate_for_date bisect')
    else:
        rep.ok(f'MR-5 토요일 → 직전 금요일 환율', f'{cur} {sat:%Y-%m-%d} → '
               f'{fri:%Y-%m-%d} {expect} ({len(rows)}행)')


def mr6_monthly_missing(tmp, rep):
    """MR-6 공식 월평균 결손 — 자체 평균으로 대체하지 않고 중단해야 한다."""
    paths = case_inputs('qoo10')
    if not paths:
        rep.skip('MR-6 월평균 결손', '큐텐 샘플 없음')
        return
    ok_run = run_pipeline(paths, tmp, 'mr6_ok')
    month = None
    for c in ok_run.rows_by_currency.get('JPY', []):
        d = str(c['ship_date'] or '')
        if len(d) == 8:
            month = f'{d[:4]}-06' if int(d[4:6]) <= 6 else f'{d[:4]}-12'
            break
    if not month:
        rep.skip('MR-6 월평균 결손', 'JPY 기준월을 못 찾음')
        return
    try:
        run_pipeline(paths, tmp, 'mr6_drop', drop_monthly=[('JPY', month)])
    except RuntimeError as e:
        rep.ok('MR-6 월평균 결손 → 중단', f'JPY {month} 제거 시 RuntimeError: {str(e)[:70]}')
        return
    except Exception as e:  # noqa: BLE001
        rep.mismatch('MR-6 월평균 결손', 'RuntimeError', f'{type(e).__name__}: {e}', 0,
                     'MR-6 / VERIFICATION B-3', 'exchange_rate.monthly_avg_rate_for_month')
        return
    rep.mismatch('MR-6 월평균 결손', 'RuntimeError 로 중단', '조용히 계속 진행', 0,
                 'MR-6 / VERIFICATION B-3 (월평균 공식값만 사용) · E-2',
                 'exchange_rate.py:1474-1487 monthly_avg_rate_for_month 가 대체값을 쓰고 있음')


def mr7_detection(tmp, rep):
    """MR-7a 무관 파일 투입 / MR-7b 파일명 변경 — 판별은 내용 기반이어야 한다."""
    from modules.pdf_parser import detect_pdf_type
    paths = case_inputs('joom')
    if not paths:
        rep.skip('MR-7 판별', 'Joom 샘플 없음')
        return
    base = run_pipeline(paths, tmp, 'mr7_base')

    # 7a — PDF 확장자지만 내용이 PDF가 아닌 파일. unknown 으로 떨어지고 합계는 불변이어야 한다
    junk = tmp / '무관한자료.pdf'
    junk.write_bytes('이것은 소포수령증이 아닙니다.\n임의의 텍스트 파일입니다.\n'.encode('utf-8'))
    run_a = run_pipeline(paths + [junk], tmp, 'mr7a')
    kind = run_a.detected.get(junk.name)
    if kind not in ('unknown', None):
        rep.mismatch('MR-7a 무관 파일 판별', 'unknown', kind, 0,
                     'MR-7a / VERIFICATION E-1 (미지 형식은 수동 선택으로 유도)',
                     'pdf_parser.detect_pdf_type / _detect_pdf_type_from_text')
    elif run_a.krw_by_currency != base.krw_by_currency:
        rep.mismatch('MR-7a 무관 파일 투입 후 합계', base.krw_by_currency,
                     run_a.krw_by_currency, run_a.total_krw - base.total_krw,
                     'MR-7a / VERIFICATION E-1', 'app.py/process.py 미지 파일 처리')
    else:
        rep.ok('MR-7a 무관 파일 → unknown, 합계 불변', f'{base.total_krw:,}원 유지')

    # 7b — 파일명만 바꾼 사본. 판별과 원화가 모두 그대로여야 한다 (내용 기반 판별)
    renamed = []
    for i, p in enumerate(case_inputs('shopee', '*_TH_*.pdf') + paths):
        dst = tmp / f'doc{i}{p.suffix}'
        shutil.copy(p, dst)
        renamed.append((p, dst))
    if not renamed:
        rep.skip('MR-7b 파일명 변경', '대상 파일 없음')
        return
    mismatched = []
    for orig, dst in renamed:
        a = quiet(detect_pdf_type, str(orig))
        b = quiet(detect_pdf_type, str(dst))
        if a != b:
            mismatched.append((orig.name, a, dst.name, b))
    if mismatched:
        rep.mismatch('MR-7b 파일명 무관 판별', '판별 결과 동일',
                     f'{len(mismatched)}건 상이 (예: {mismatched[0]})', len(mismatched),
                     'MR-7b / VERIFICATION E-1 (판별은 내용 기반)',
                     'pdf_parser.detect_pdf_type 이 파일명 규칙에만 의존')
    else:
        origs = run_pipeline([o for o, _d in renamed], tmp, 'mr7b_orig')
        news = run_pipeline([d for _o, d in renamed], tmp, 'mr7b_new')
        if origs.krw_by_currency != news.krw_by_currency:
            rep.mismatch('MR-7b 파일명 변경 후 원화', origs.krw_by_currency,
                         news.krw_by_currency, news.total_krw - origs.total_krw,
                         'MR-7b / VERIFICATION E-1', 'pdf_parser 파일명 기반 통화·유형 추정')
        else:
            rep.ok('MR-7b 파일명 변경 → 판별·원화 불변',
                   f'{len(renamed)}개 사본, {origs.total_krw:,}원 유지')


def mr8_one_by_one(tmp, rep):
    """MR-8 단일 처리 합 vs 배치 처리 합 — MR-3 의 일반형."""
    paths = case_inputs('shopee')
    if len(paths) < 2:
        rep.skip('MR-8 단일 vs 배치', f'쇼피 PDF {len(paths)}개')
        return
    batch = run_pipeline(paths, tmp, 'mr8_batch')
    one = {}
    for i, p in enumerate(paths):
        r = run_pipeline([p], tmp, f'mr8_one{i}')
        for cur, krw in r.krw_by_currency.items():
            one[cur] = one.get(cur, 0) + krw
    if one != batch.krw_by_currency:
        diff = {c: one.get(c, 0) - batch.krw_by_currency.get(c, 0)
                for c in set(one) | set(batch.krw_by_currency)
                if one.get(c, 0) != batch.krw_by_currency.get(c, 0)}
        rep.mismatch('MR-8 단일 합 vs 배치', batch.krw_by_currency, one, diff,
                     'MR-8 / VERIFICATION C-1·C-4', 'excel_writer 통화 시트 병합 / 환율 조회 범위')
    else:
        rep.ok('MR-8 단일 처리 합 = 배치 처리 합',
               f'{sum(one.values()):,}원 (파일 {len(paths)}개)')


# ── 실행 ────────────────────────────────────────────────────────

MRS = {
    'MR-1': mr1_shuffle,
    'MR-2': mr2_reproducible,
    'MR-3': mr3_split_merge,
    'MR-4': mr4_basis_date,
    'MR-5': mr5_holiday_ffill,
    'MR-6': mr6_monthly_missing,
    'MR-7': mr7_detection,
    'MR-8': mr8_one_by_one,
}


def main():
    wanted = [a for a in sys.argv[1:]] or list(MRS)
    for w in wanted:
        if w not in MRS:
            print(f'알 수 없는 불변식: {w} (가능: {", ".join(MRS)})')
            return 2

    rep = Report('메타모픽 불변식 (정답지 비의존)')
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        assert not str(tmp.resolve()).startswith(str(S.resolve())), 'samples/ 에는 쓰지 않습니다'
        install_guards(tmp)   # data/ 캐시 보호 + SMBS 차단
        for name in wanted:
            try:
                MRS[name](tmp, rep)
            except Exception as e:  # noqa: BLE001
                rep.broken(f'{name} 하니스 오류: {type(e).__name__}: {e}')

    print(rep.render())
    print()
    skipped = sum(1 for ln in rep.lines if '[건너뜀]' in ln)
    print(f'요약: 일치 {rep.n_pass} / 불일치 {rep.n_mismatch} / 건너뜀 {skipped} / '
          f'구조붕괴 {len(rep.struct)}')
    if rep.struct:
        print('❌ 하니스 자체 오류가 있습니다.')
        return 2
    if rep.n_mismatch:
        print(f'❌ 불일치 {rep.n_mismatch}건')
        return 1
    if skipped >= 3:
        print(f'❌ 건너뜀 {skipped}건 — 조용히 통과한 것처럼 보이지 않도록 실패로 취급합니다.')
        return 1
    print('✅ 전체 통과' + (f' (건너뜀 {skipped}건)' if skipped else ''))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
